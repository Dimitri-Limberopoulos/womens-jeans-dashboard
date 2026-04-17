#!/usr/bin/env python3
"""
Production Target PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  pdp_urls.csv  (one URL per line, no header — same folder as this script)
Output: target_pdp_results.xlsx + target_pdp_progress.json

Architecture: 5 browser contexts × 2 tabs each = 10 parallel workers
              30-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

Each product page is fetched ONCE. The embedded JSON is parsed to produce
one row per color option, each with its own specs, image, sizes, and TCIN.
"""

import asyncio, json, os, random, re, time, html as html_mod
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 5
TABS_PER_CTX = 2
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 10 total workers
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 30 * 60  # 30 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# ── Target Owned Brands ─────────────────────────────────────────────────────
TARGET_OWNED_BRANDS = {
    'universal thread', 'wild fable', 'ava & viv', 'a new day',
}


def classify_brand(brand_name):
    """Returns True if Target owned brand, False for national brand."""
    if brand_name and brand_name.strip().lower() in TARGET_OWNED_BRANDS:
        return True
    return False


# ── Target Data Extraction ──────────────────────────────────────────────────

def extract_child_pricing(page_html):
    """
    Extract per-child TCIN pricing from the children array.
    Returns {tcin: {'current_retail': float, 'formatted_price': str, 'price_type': str, 'reg_retail': float}}

    Handles both double-escaped (own brands) and single-escaped (marketplace) JSON.
    """
    pricing = {}

    def _store(tcin, formatted, ptype, cr, rr):
        if tcin in pricing:
            return
        if formatted:
            pricing[tcin] = {
                'formatted_price': formatted,
                'price_type': ptype or '',
                'current_retail': cr,
                'reg_retail': rr,
            }
        elif cr is not None:
            pricing[tcin] = {
                'formatted_price': f"${cr:.2f}",
                'price_type': ptype or '',
                'current_retail': cr,
                'reg_retail': rr,
            }

    # ── Pass 1: Double-escaped children array ──
    children_start = page_html.find('children\\":[')
    if children_start >= 0:
        chunk = page_html[children_start:children_start + 300000]
        child_blocks = list(re.finditer(r'__typename\\":\\"Product\\",\\"tcin\\":\\"(\d+)', chunk))
        for i, cb in enumerate(child_blocks):
            tcin = cb.group(1)
            end = child_blocks[i + 1].start() if i + 1 < len(child_blocks) else cb.end() + 20000
            cc = chunk[cb.end():end]

            pm = re.search(r'formatted_current_price\\":\\"([^"\\\\]+)', cc)
            tm = re.search(r'formatted_current_price_type\\":\\"([^"\\\\]+)', cc)
            rm = re.search(r'current_retail\\":(\d+\.?\d*)', cc)
            rg = re.search(r'reg_retail\\":(\d+\.?\d*)', cc)

            _store(tcin,
                   pm.group(1) if pm else None,
                   tm.group(1) if tm else None,
                   float(rm.group(1)) if rm else None,
                   float(rg.group(1)) if rg else None)

    # ── Pass 2: Single-escaped children array (marketplace) ──
    children_start2 = page_html.find('"children":[')
    if children_start2 >= 0:
        chunk2 = page_html[children_start2:children_start2 + 300000]
        child_blocks2 = list(re.finditer(r'"__typename":"Product","tcin":"(\d+)"', chunk2))
        for i, cb in enumerate(child_blocks2):
            tcin = cb.group(1)
            end = child_blocks2[i + 1].start() if i + 1 < len(child_blocks2) else cb.end() + 20000
            cc = chunk2[cb.end():end]

            pm = re.search(r'"formatted_current_price":"([^"]+)"', cc)
            tm = re.search(r'"formatted_current_price_type":"([^"]*)"', cc)
            rm = re.search(r'"current_retail":(\d+\.?\d*)', cc)
            rg = re.search(r'"reg_retail":(\d+\.?\d*)', cc)

            _store(tcin,
                   pm.group(1) if pm else None,
                   tm.group(1) if tm else None,
                   float(rm.group(1)) if rm else None,
                   float(rg.group(1)) if rg else None)

    # ── Pass 3: Full-page TCIN scan (both encodings) ──
    for pattern in [r'tcin\\":\\"(\d+)\\"', r'"tcin":"(\d+)"']:
        for m in re.finditer(pattern, page_html):
            tcin = m.group(1)
            if tcin in pricing:
                continue
            window = page_html[m.end():m.end() + 2000]
            # Don't cross into another TCIN block
            for tp in [r'tcin\\":\\"(\d+)\\"', r'"tcin":"(\d+)"']:
                nt = re.search(tp, window)
                if nt:
                    window = window[:nt.start()]

            # Try double-escaped
            rm = re.search(r'current_retail\\":(\d+\.?\d*)', window)
            pm = re.search(r'formatted_current_price\\":\\"([^"\\\\]+)', window)
            tm = re.search(r'formatted_current_price_type\\":\\"([^"\\\\]+)', window)
            rg = re.search(r'reg_retail\\":(\d+\.?\d*)', window)
            # Try single-escaped
            if not rm:
                rm = re.search(r'"current_retail":(\d+\.?\d*)', window)
            if not pm:
                pm = re.search(r'"formatted_current_price":"([^"]+)"', window)
            if not tm:
                tm = re.search(r'"formatted_current_price_type":"([^"]*)"', window)
            if not rg:
                rg = re.search(r'"reg_retail":(\d+\.?\d*)', window)

            if rm or pm:
                _store(tcin,
                       pm.group(1) if pm else None,
                       tm.group(1) if tm else None,
                       float(rm.group(1)) if rm else None,
                       float(rg.group(1)) if rg else None)

    return pricing


def parse_price_range(price_str):
    """Parse '$69.99 - $74.99' into (69.99, 74.99). Single price returns (28.0, 28.0)."""
    if not price_str:
        return None, None
    prices = re.findall(r'\$?([\d,.]+)', str(price_str))
    if len(prices) >= 2:
        return float(prices[0].replace(',', '')), float(prices[1].replace(',', ''))
    elif len(prices) == 1:
        p = float(prices[0].replace(',', ''))
        return p, p
    return None, None


def extract_color_variation_data(page_html):
    """
    Parse the variation_hierarchy to build a per-color data structure.

    Handles THREE JSON encodings found on Target:
      1. Double-escaped (own-brand):  name\\":\\"Color\\"
      2. Single-escaped (marketplace): name":"color"  (inside script tags)
      3. Inverted hierarchy:  Size → Color (some products like KBB)

    Returns {color_name: {swatch_url, primary_image_url, tcins, first_tcin,
                          size_groups, sizes, all_size_variants, buy_url}}
    """
    colors = {}

    def _ensure_color(name, swatch=''):
        """Get or create a color record."""
        if name not in colors:
            colors[name] = {
                'swatch_url': swatch,
                'primary_image_url': '',
                'tcins': [],
                'first_tcin': '',
                'size_groups': [],
                'sizes': [],
                'all_size_variants': [],
                'buy_url': '',
            }
        elif swatch and not colors[name]['swatch_url']:
            colors[name]['swatch_url'] = swatch
        return colors[name]

    def _add_tcin(c, tcin, img='', buy=''):
        if tcin and tcin not in c['tcins']:
            c['tcins'].append(tcin)
        if img and not c['primary_image_url']:
            c['primary_image_url'] = img
        if buy and not c['buy_url']:
            c['buy_url'] = buy

    # ── Pattern 1: Double-escaped JSON (Target own brands, national brands) ──
    # "name\\":\\"Color\\",\\"value\\":\\"<name>\\",\\"swatch_image_url\\":\\"<url>\\"
    dbl_color_entries = list(re.finditer(
        r'name\\":\\"Color\\",\\"value\\":\\"([^"\\\\]+)\\",\\"(?:tcin|swatch)',
        page_html, re.IGNORECASE
    ))

    for ci, cm in enumerate(dbl_color_entries):
        color_name = cm.group(1).strip()
        end_pos = dbl_color_entries[ci + 1].start() if ci + 1 < len(dbl_color_entries) else cm.end() + 100000
        color_chunk = page_html[cm.start():end_pos]

        swatch_m = re.search(r'swatch_image_url\\":\\"((?:[^"\\]|\\.)*?)\\"', color_chunk)
        swatch_url = swatch_m.group(1).replace('\\u0026', '&') if swatch_m else ''

        c = _ensure_color(color_name, swatch_url)

        # Nested Size Group → Size entries
        for g in re.findall(r'name\\":\\"Size Group\\",\\"value\\":\\"([^"\\\\]+)', color_chunk):
            g = g.strip()
            if g not in c['size_groups']:
                c['size_groups'].append(g)

        for size_val, tcin, img, buy in re.findall(
            r'name\\":\\"Size\\",\\"value\\":\\"([^"\\\\]+)\\",\\"tcin\\":\\"(\d+)\\"'
            r'(?:,\\"primary_image_url\\":\\"([^"\\\\]+)\\")?'
            r'(?:,\\"buy_url\\":\\"([^"\\\\]+)\\")?',
            color_chunk
        ):
            _add_tcin(c, tcin, img, buy)
            sv = size_val.strip()
            if sv not in c['all_size_variants']:
                c['all_size_variants'].append(sv)

        # Leaf TCIN directly on color entry (inverted or flat layout)
        leaf = re.search(r'tcin\\":\\"(\d+)\\"', color_chunk[:300])
        leaf_img = re.search(r'primary_image_url\\":\\"([^"\\\\]+)\\"', color_chunk[:600])
        leaf_buy = re.search(r'buy_url\\":\\"([^"\\\\]+)\\"', color_chunk[:600])
        if leaf:
            _add_tcin(c, leaf.group(1),
                       leaf_img.group(1) if leaf_img else '',
                       leaf_buy.group(1) if leaf_buy else '')

    # ── Pattern 2: Single-escaped JSON (marketplace sellers) ──
    # "name":"color","value":"<name>","tcin":"<id>","swatch_image_url":"<url>"
    sgl_color_entries = list(re.finditer(
        r'"name":"[Cc]olor","value":"([^"]+)","tcin":"(\d+)"',
        page_html
    ))

    for m in sgl_color_entries:
        color_name = m.group(1).strip()
        tcin = m.group(2)

        # Grab swatch and image from nearby context
        ctx = page_html[m.end():m.end() + 500]
        swatch_m = re.search(r'"swatch_image_url":"([^"]+)"', ctx)
        img_m = re.search(r'"primary_image_url":"([^"]+)"', ctx)
        buy_m = re.search(r'"buy_url":"([^"]+)"', ctx)

        swatch_url = swatch_m.group(1).replace('\\u0026', '&') if swatch_m else ''
        c = _ensure_color(color_name, swatch_url)
        _add_tcin(c, tcin,
                  img_m.group(1) if img_m else '',
                  buy_m.group(1) if buy_m else '')

    # Also look for single-escaped color entries with swatch before tcin
    for m in re.finditer(
        r'"name":"[Cc]olor","value":"([^"]+)","swatch_image_url":"([^"]+)"',
        page_html
    ):
        color_name = m.group(1).strip()
        swatch_url = m.group(2).replace('\\u0026', '&')
        ctx = page_html[m.end():m.end() + 500]

        c = _ensure_color(color_name, swatch_url)

        # Look for nested size/tcin entries in single-escaped format
        for sv, tcin in re.findall(r'"name":"[Ss]ize","value":"([^"]+)","tcin":"(\d+)"', ctx):
            _add_tcin(c, tcin)
            sv = sv.strip()
            if sv not in c['all_size_variants']:
                c['all_size_variants'].append(sv)

        # Or a direct tcin on the color entry
        tcin_m = re.search(r'"tcin":"(\d+)"', ctx[:200])
        if tcin_m:
            img_m = re.search(r'"primary_image_url":"([^"]+)"', ctx)
            buy_m = re.search(r'"buy_url":"([^"]+)"', ctx)
            _add_tcin(c, tcin_m.group(1),
                      img_m.group(1) if img_m else '',
                      buy_m.group(1) if buy_m else '')

    # ── Finalize ──
    for color_name, c in colors.items():
        c['first_tcin'] = c['tcins'][0] if c['tcins'] else ''
        c['sizes'] = [s for s in c['all_size_variants']
                      if 'Short' not in s and 'Long' not in s]

    return colors


def extract_child_specs(page_html, tcin):
    """Extract specs from a specific child TCIN's product_description in the children array.
    Handles both double-escaped and single-escaped JSON."""
    specs = {}

    # Double-escaped
    match = re.search(rf'tcin\\":\\"{tcin}\\".*?bullet_descriptions\\":\[(.*?)\]', page_html)
    if match:
        for key, val in re.findall(r'\\u003cB\\u003e([^\\]+):\\u003c/B\\u003e\s*([^"\\\\]+)', match.group(1)):
            clean_key = key.strip()
            if clean_key not in specs:
                specs[clean_key] = val.strip()
        return specs

    # Single-escaped
    match = re.search(rf'"tcin":"{tcin}".*?"bullet_descriptions":\[(.*?)\]', page_html)
    if match:
        for key, val in re.findall(r'<B>([^<]+):</B>\s*([^"<]+)', match.group(1)):
            clean_key = key.strip()
            if clean_key not in specs:
                specs[clean_key] = val.strip()

    return specs


def _rx(page_html, dbl_pattern, sgl_pattern=None):
    """Try double-escaped regex first, then single-escaped. Returns match or None."""
    m = re.search(dbl_pattern, page_html)
    if m:
        return m
    if sgl_pattern:
        return re.search(sgl_pattern, page_html)
    return None


def _extract_color_from_title(title):
    """Extract color/wash from title for products with no color variation data.
    E.g. 'Jeans - Denim, 18P' → 'Denim', 'Jeans Black 1X' → 'Black'"""
    if not title:
        return ''
    # Pattern: after brand name or dash, before size — common color words
    # Try "- ColorName," or "- ColorName Size" patterns
    m = re.search(r'[-–]\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*(?:,|\d|$)', title)
    if m:
        candidate = m.group(1).strip()
        # Filter out non-color words
        skip = {'Women', 'Petite', 'Plus', 'Size', 'Length', 'Short', 'Long', 'Regular'}
        if candidate not in skip:
            return candidate
    # Try title ending pattern: "Jeans ColorName SizeCode"
    m = re.search(r'(?:Jeans?|Denim|Pants?)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?(?:\s*/\s*[A-Z][a-z]+)?)\s+\d', title)
    if m:
        candidate = m.group(1).strip()
        skip = {'Women', 'Petite', 'Plus', 'Size', 'Length', 'Short', 'Long', 'Regular', 'Mid', 'High', 'Low'}
        if candidate not in skip:
            return candidate
    return ''


def parse_target_pdp(page_html, url, js_pricing=None):
    """
    Parse a Target PDP page and return a LIST of result dicts — one per color.
    If a product has no color variations, returns a single-element list.

    js_pricing: optional dict of {tcin: pricing_info} extracted via page.evaluate()
                Used as supplemental source when HTML regex misses per-child pricing.
    """
    rows = []

    try:
        # ── Shared / parent-level data ──────────────────────────────────
        parent = {}

        # Title
        title_match = re.search(r'<title[^>]*>(.*?)</title>', page_html, re.DOTALL)
        raw_title = html_mod.unescape(title_match.group(1)).replace(' : Target', '').strip() if title_match else ''
        parent['title'] = raw_title

        # Clean title: remove color/size suffix for parent-level name
        parent['title_clean'] = re.sub(r'\s+\d+$', '', raw_title).strip()

        # Parent TCIN from URL
        tcin_url_match = re.search(r'/A-(\d+)', url)
        parent['parent_tcin'] = tcin_url_match.group(1) if tcin_url_match else ''

        # Preselected TCIN from URL
        preselect_match = re.search(r'preselect=(\d+)', url)
        parent['preselected_tcin'] = preselect_match.group(1) if preselect_match else ''

        # Brand (try double then single-escaped)
        brand_match = _rx(page_html,
                          r'primary_brand.*?name\\":\\"([^"\\\\]+)',
                          r'"primary_brand".*?"name":"([^"]+)"')
        parent['brand'] = brand_match.group(1) if brand_match else ''
        parent['owned_brand'] = classify_brand(parent['brand'])
        parent['brand_type'] = 'Owned Brand' if parent['owned_brand'] else 'National Brand'

        # ── Pricing (parent level) ──────────────────────────────────────
        price_match = _rx(page_html,
                          r'formatted_current_price\\":\\"([^"\\\\]+)',
                          r'"formatted_current_price":"([^"]+)"')
        parent['current_price'] = price_match.group(1) if price_match else ''

        price_type_match = _rx(page_html,
                               r'formatted_current_price_type\\":\\"([^"\\\\]+)',
                               r'"formatted_current_price_type":"([^"]*)"')
        parent['price_type'] = price_type_match.group(1) if price_type_match else ''

        retail_min = _rx(page_html,
                         r'current_retail_min\\":(\d+\.?\d*)',
                         r'"current_retail_min":(\d+\.?\d*)')
        parent['current_retail_min'] = float(retail_min.group(1)) if retail_min else ''

        retail_max = _rx(page_html,
                         r'reg_retail_max\\":(\d+\.?\d*)',
                         r'"reg_retail_max":(\d+\.?\d*)')
        parent['reg_retail_max'] = float(retail_max.group(1)) if retail_max else ''

        # Sale/clearance pricing
        save_dollar = _rx(page_html,
                          r'save_dollar\\":(\d+\.?\d*)',
                          r'"save_dollar":(\d+\.?\d*)')
        parent['save_dollar'] = float(save_dollar.group(1)) if save_dollar else ''

        save_pct = _rx(page_html,
                       r'save_percent\\":(\d+\.?\d*)',
                       r'"save_percent":(\d+\.?\d*)')
        parent['save_percent'] = f"{float(save_pct.group(1)):.0f}%" if save_pct else ''

        # Original price — only populate when item is on sale/clearance
        parent['original_price'] = ''
        if parent['price_type'] in ('sale', 'clearance', 'reduced'):
            reg_price = _rx(page_html, r'reg_retail\\":(\d+\.?\d*)', r'"reg_retail":(\d+\.?\d*)')
            if reg_price:
                parent['original_price'] = f"${float(reg_price.group(1)):.2f}"
            elif parent['reg_retail_max'] and parent['current_retail_min']:
                if float(parent['reg_retail_max']) > float(parent['current_retail_min']):
                    parent['original_price'] = f"${float(parent['reg_retail_max']):.2f}"

        parent['discount_percent'] = calc_discount(
            parent.get('original_price', ''), parent.get('current_price', '')
        )

        # ── Ratings & Reviews ───────────────────────────────────────────
        rating_avg = _rx(page_html, r'average\\":(\d+\.?\d*)', r'"average":(\d+\.?\d*)')
        parent['rating_avg'] = float(rating_avg.group(1)) if rating_avg else ''

        rating_count = _rx(page_html,
                           r'average\\":\d+\.?\d*,\\"count\\":(\d+)',
                           r'"average":\d+\.?\d*,"count":(\d+)')
        parent['review_count'] = int(rating_count.group(1)) if rating_count else ''

        for star in [1, 2, 3, 4, 5]:
            star_match = _rx(page_html,
                             rf'rating{star}\\":(\d+)',
                             rf'"rating{star}":(\d+)')
            parent[f'stars_{star}'] = int(star_match.group(1)) if star_match else ''

        for attr in ['comfort', 'sizing', 'quality', 'value']:
            attr_match = _rx(page_html,
                             rf'id\\":\\"{attr}\\".*?value\\":(\d+\.?\d*)',
                             rf'"id":"{attr}".*?"value":(\d+\.?\d*)')
            parent[f'rating_{attr}'] = float(attr_match.group(1)) if attr_match else ''

        sizing_selector = _rx(page_html,
                              r'id\\":\\"sizing\\".*?selector_text\\":\\"([^"\\\\]+)',
                              r'"id":"sizing".*?"selector_text":"([^"]+)"')
        parent['sizing_feedback'] = sizing_selector.group(1) if sizing_selector else ''

        question_count = _rx(page_html, r'question_count\\":(\d+)', r'"question_count":(\d+)')
        parent['question_count'] = int(question_count.group(1)) if question_count else ''

        # ── Description ─────────────────────────────────────────────────
        desc_match = _rx(page_html,
                         r'downstream_description\\":\\"(.*?)(?:\\",|\\"})',
                         r'"downstream_description":"(.*?)(?:",|"})')
        if desc_match:
            desc = desc_match.group(1)
            desc = desc.replace('\\u003cbr /\\u003e', ' ').replace('\\u003cbr/\\u003e', ' ')
            desc = desc.replace('\\u003cbr\\u003e', ' ').replace('\\u0026', '&')
            desc = re.sub(r'\\u003c[^\\]*?\\u003e', '', desc)
            desc = re.sub(r'<[^>]+>', '', desc)  # clean single-escaped HTML tags
            parent['description'] = desc.strip()
        else:
            parent['description'] = ''

        # Highlights
        soft_match = _rx(page_html,
                         r'"bullets\\":\[(.*?)\]',
                         r'"bullets":\[(.*?)\]')
        if soft_match:
            bullets_raw = soft_match.group(1)
            bullets = re.findall(r'"([^"]{3,})"', bullets_raw)
            if not bullets:
                bullets = re.findall(r'\\"([^"\\\\]{3,})\\"', bullets_raw)
            parent['highlights'] = ' | '.join(bullets)
        else:
            parent['highlights'] = ''

        # ── Origin ──────────────────────────────────────────────────────
        origin_match = _rx(page_html,
                           r'import_designation_description\\":\\"([^"\\\\]+)',
                           r'"import_designation_description":"([^"]+)"')
        parent['origin'] = origin_match.group(1) if origin_match else ''

        # ── Department & Classification ─────────────────────────────────
        dept_match = _rx(page_html,
                         r'department_name\\":\\"([^"\\\\]+)',
                         r'"department_name":"([^"]+)"')
        parent['department'] = dept_match.group(1) if dept_match else ''

        dept_id_match = _rx(page_html, r'department_id\\":(\d+)', r'"department_id":(\d+)')
        parent['department_id'] = int(dept_id_match.group(1)) if dept_id_match else ''

        class_id_match = _rx(page_html, r'class_id\\":(\d+)', r'"class_id":(\d+)')
        parent['class_id'] = int(class_id_match.group(1)) if class_id_match else ''

        prod_type_match = _rx(page_html,
                              r'product_type_name\\":\\"([^"\\\\]+)',
                              r'"product_type_name":"([^"]+)"')
        parent['product_type'] = prod_type_match.group(1) if prod_type_match else ''

        # Category breadcrumbs
        cats = re.findall(r'category_id\\":\\"[^"\\\\]+\\",\\"name\\":\\"([^"\\\\]+)', page_html)
        if not cats:
            cats = re.findall(r'"category_id":"[^"]+","name":"([^"]+)"', page_html)
        seen = set()
        unique_cats = []
        for c in cats:
            if c not in seen and c != 'target':
                seen.add(c)
                unique_cats.append(c)
        parent['breadcrumb'] = ' > '.join(unique_cats)
        for i in range(5):
            parent[f'category_{i+1}'] = unique_cats[i] if i < len(unique_cats) else ''

        # OG metadata
        og_img = re.search(r'(?:property="og:image"\s+content|content)="(https://target\.scene7\.com[^"]+)"', page_html)
        if not og_img:
            og_img = re.search(r'og:image[^>]+content="([^"]+)"', page_html)
        parent['og_image_url'] = og_img.group(1) if og_img else ''

        og_desc = re.search(r'(?:name="description"\s+content|content)="(Shop [^"]*)"', page_html)
        if not og_desc:
            og_desc = re.search(r'name="description"[^>]+content="([^"]*)"', page_html)
        parent['meta_description'] = html_mod.unescape(og_desc.group(1)).strip() if og_desc else ''

        canonical = re.search(r'rel="canonical"\s+[^>]*href="([^"]+)"', page_html)
        if not canonical:
            canonical = re.search(r'href="([^"]+)"\s+rel="canonical"', page_html)
        parent['canonical_url'] = canonical.group(1) if canonical else ''

        # Image count
        alt_images = re.findall(r'alternate_images.*?url\\":\\"([^"\\\\]+)', page_html[:200000])
        parent['image_count'] = len(set(alt_images)) + 1 if alt_images else 1

        # Return policy
        ret_match = re.search(r'return_policies_guest_message\\":\\"(.*?)(?:\\\\"|")', page_html)
        parent['return_policy'] = ret_match.group(1)[:200] if ret_match else ''

        # Bought last month
        bought_match = re.search(r'(\d+[kK]?\+?\s*bought in (?:last|past) month)', page_html, re.IGNORECASE)
        parent['bought_last_month'] = bought_match.group(1) if bought_match else ''

        # Package dimensions
        dims_block = re.search(
            r'dimensions\\":\{[^}]*?depth\\":(\d+\.?\d*)[^}]*?weight\\":(\d+\.?\d*)[^}]*?'
            r'height\\":(\d+\.?\d*)[^}]*?width\\":(\d+\.?\d*)',
            page_html
        )
        if dims_block:
            depth, weight, height, width = dims_block.groups()
            parent['package_weight_lbs'] = float(weight)
            if float(height) < 50 and float(width) < 50:
                parent['package_dimensions'] = f"{height}×{width}×{depth} in"
            else:
                parent['package_dimensions'] = ''
        else:
            parent['package_weight_lbs'] = ''
            parent['package_dimensions'] = ''

        # Shop the look
        stl_match = re.findall(r'shop_the_look\\":\[([^\]]+)\]', page_html)
        if stl_match:
            related = re.findall(r'\\"(\d+)\\"', stl_match[0])
            parent['shop_the_look_tcins'] = ', '.join(related)
        else:
            parent['shop_the_look_tcins'] = ''

        # ── Per-color expansion ─────────────────────────────────────────
        color_data = extract_color_variation_data(page_html)
        child_pricing = extract_child_pricing(page_html)

        # Merge JS-extracted pricing (fills gaps the regex approach missed)
        if js_pricing:
            for tcin, pdata in js_pricing.items():
                if tcin not in child_pricing:
                    child_pricing[tcin] = pdata

        # Get total color count for the parent
        all_color_names = list(color_data.keys())
        parent['total_colors_in_product'] = len(all_color_names)
        parent['all_colors'] = ', '.join(all_color_names)

        # Parse parent price range into min/max
        parent_price_min, parent_price_max = parse_price_range(parent.get('current_price', ''))

        # Detect whether parent price is a range or single value
        parent_is_range = ' - ' in parent.get('current_price', '')

        def _apply_color_pricing(row, tcins):
            """
            Set per-color pricing on a row.
            Priority:
              1. Child-level pricing from children array (best — specific to this TCIN)
              2. Parent single price (if not a range, all colors share it)
              3. Parent range with min/max (last resort)
            """
            color_price = None
            for t in tcins:
                if t in child_pricing:
                    color_price = child_pricing[t]
                    break

            if color_price:
                # ── Found child-level pricing for this color ──
                cp = color_price.get('formatted_price', '')
                row['color_current_price'] = cp
                row['color_price_type'] = color_price.get('price_type', '')
                row['color_current_retail'] = color_price.get('current_retail', '')
                row['color_reg_retail'] = color_price.get('reg_retail', '')

                # Override top-level price fields with color-specific values
                row['current_price'] = cp
                row['price_type'] = color_price.get('price_type', '') or parent.get('price_type', '')

                # Original price: only when on sale/clearance
                cr = color_price.get('current_retail')
                rr = color_price.get('reg_retail')
                if row['price_type'] in ('sale', 'clearance', 'reduced') and rr and cr and rr > cr:
                    row['original_price'] = f"${rr:.2f}"
                    row['discount_percent'] = calc_discount(row['original_price'], cp)
                    row['save_dollar'] = round(rr - cr, 2)
                    row['save_percent'] = f"{((rr - cr) / rr * 100):.0f}%"
                elif row['price_type'] not in ('sale', 'clearance', 'reduced'):
                    row['original_price'] = ''
                    row['discount_percent'] = ''
                    row['save_dollar'] = ''
                    row['save_percent'] = ''

                pmin, pmax = parse_price_range(cp)
                row['price_min'] = pmin
                row['price_max'] = pmax

            elif not parent_is_range:
                # ── Parent price is a single value — all colors share it ──
                pp = parent.get('current_price', '')
                row['color_current_price'] = pp
                row['color_price_type'] = parent.get('price_type', '')
                row['color_current_retail'] = parent.get('current_retail_min', '')
                row['color_reg_retail'] = parent.get('reg_retail_max', '')
                row['price_min'] = parent_price_min
                row['price_max'] = parent_price_max

            else:
                # ── Parent price is a range and no child pricing found ──
                # Fill with parent range so the field is never empty
                row['color_current_price'] = parent.get('current_price', '')
                row['color_price_type'] = parent.get('price_type', '')
                row['color_current_retail'] = ''
                row['color_reg_retail'] = ''
                row['price_min'] = parent_price_min
                row['price_max'] = parent_price_max

        if not color_data:
            # No color variations found — single row with parent-level specs
            specs = {}
            # Try double-escaped bullet specs
            bullet_descs = re.findall(
                r'\\u003cB\\u003e([^\\]+):\\u003c/B\\u003e\s*([^"\\\\]+)',
                page_html
            )
            if not bullet_descs:
                # Try single-escaped (marketplace)
                bullet_descs = re.findall(
                    r'<B>([^<]+):</B>\s*([^"<]+)',
                    page_html
                )
            for key, val in bullet_descs:
                clean_key = key.strip()
                if clean_key not in specs:
                    specs[clean_key] = val.strip()

            # Try to extract color from title for single-color products
            title_color = _extract_color_from_title(parent.get('title', ''))

            row = {**parent}
            row['url'] = url
            row['color'] = title_color
            row['color_swatch_url'] = ''
            row['color_image_url'] = parent.get('og_image_url', '')
            row['color_buy_url'] = ''
            row['color_tcin'] = parent.get('preselected_tcin', '') or parent.get('parent_tcin', '')
            row['color_sizes'] = ''
            row['color_num_sizes'] = ''
            row['color_size_groups'] = ''
            row['color_all_size_variants'] = ''
            row['color_num_skus'] = ''
            # Use preselected TCIN for pricing lookup
            preselect_tcins = [parent.get('preselected_tcin', ''), parent.get('parent_tcin', '')]
            preselect_tcins = [t for t in preselect_tcins if t]
            _apply_color_pricing(row, preselect_tcins)
            row.update(_specs_to_fields(specs))
            row['non_basic'] = detect_non_basic(
                parent.get('title', ''), title_color, parent.get('highlights', '') + ' ' + parent.get('description', '')
            )
            row['timestamp'] = datetime.now().isoformat()
            rows.append(row)
        else:
            # One row per color
            for color_name, cdata in color_data.items():
                row = {**parent}
                row['url'] = url
                row['color'] = color_name
                row['color_swatch_url'] = cdata.get('swatch_url', '')
                row['color_image_url'] = cdata.get('primary_image_url', '')
                row['color_buy_url'] = cdata.get('buy_url', '')
                row['color_tcin'] = cdata.get('first_tcin', '')
                row['color_sizes'] = ', '.join(cdata.get('sizes', []))
                row['color_num_sizes'] = len(cdata.get('sizes', []))
                row['color_size_groups'] = ', '.join(cdata.get('size_groups', []))
                row['color_all_size_variants'] = ', '.join(cdata.get('all_size_variants', []))
                row['color_num_skus'] = len(cdata.get('tcins', []))

                # Per-color pricing from children array
                _apply_color_pricing(row, cdata.get('tcins', []))

                # Per-color specs from children array
                first_tcin = cdata.get('first_tcin', '')
                if first_tcin:
                    specs = extract_child_specs(page_html, first_tcin)
                else:
                    specs = {}

                # Fallback to parent-level specs if child has none
                if not specs:
                    bullet_descs = re.findall(
                        r'\\u003cB\\u003e([^\\]+):\\u003c/B\\u003e\s*([^"\\\\]+)',
                        page_html
                    )
                    if not bullet_descs:
                        bullet_descs = re.findall(
                            r'<B>([^<]+):</B>\s*([^"<]+)',
                            page_html
                        )
                    for key, val in bullet_descs:
                        clean_key = key.strip()
                        if clean_key not in specs:
                            specs[clean_key] = val.strip()

                row.update(_specs_to_fields(specs))

                # Non-basic detection (per color — some colors are "destroy", etc.)
                row['non_basic'] = detect_non_basic(
                    parent.get('title', ''),
                    color_name,
                    parent.get('highlights', '') + ' ' + parent.get('description', '')
                )
                row['timestamp'] = datetime.now().isoformat()
                rows.append(row)

    except Exception as e:
        # On parse error, return a single error row
        rows = [{
            'url': url,
            'error': f'Parse error: {str(e)[:200]}',
            'timestamp': datetime.now().isoformat(),
        }]

    return rows


def _specs_to_fields(specs):
    """Convert a specs dict into standardized field names."""
    mat = specs.get('Material', '')
    return {
        'sizing': specs.get('Sizing', ''),
        'material': mat,
        'fabric_parsed': parse_fabric(mat),
        'pct_cotton': calc_pct_cotton(mat),
        'pct_natural': calc_pct_natural(mat),
        'pct_recycled': calc_pct_recycled(mat),
        'inseam_length': specs.get('Inseam Length', ''),
        'garment_length': specs.get('Garment Length', ''),
        'closure_style': specs.get('Closure Style', ''),
        'rise': specs.get('Rise', ''),
        'fit': specs.get('Fit', ''),
        'fabric_name': specs.get('Fabric Name', ''),
        'garment_details': specs.get('Garment Details', ''),
        'fabric_weight_type': specs.get('Fabric Weight Type', ''),
        'stretch': specs.get('Stretch', ''),
        'care_and_cleaning': specs.get('Care and Cleaning', ''),
        'hem_style': specs.get('Hem Style', ''),
        'package_quantity': specs.get('Package Quantity', ''),
    }


# ── Helper functions ────────────────────────────────────────────────────────

def parse_fabric(raw):
    if not raw:
        return ""
    m = re.search(r'(\d+\s*%\s*\w+(?:\s*,\s*\d+\s*%\s*[\w\s-]+)*)', raw, re.IGNORECASE)
    return m.group(1).strip() if m else raw[:150]


def calc_pct_cotton(mat):
    if not mat:
        return ""
    m = re.search(r'(\d+)\s*%\s*(?:Recycled\s+)?Cotton', mat, re.IGNORECASE)
    return f"{m.group(1)}%" if m else "0%"


def calc_pct_natural(mat):
    if not mat:
        return ""
    ns = re.findall(r'(\d+)\s*%\s*(?:Recycled\s+)?(?:Cotton|Wool|Silk|Linen|Hemp|Lyocell|Tencel)', mat, re.IGNORECASE)
    return f"{sum(int(n) for n in ns)}%" if ns else "0%"


def calc_pct_recycled(mat):
    if not mat:
        return ""
    ns = re.findall(r'(\d+)\s*%\s*Recycled\s+\w+', mat, re.IGNORECASE)
    return f"{sum(int(n) for n in ns)}%" if ns else "0%"


def detect_non_basic(title, color, text):
    kws = [
        'print', 'printed', 'graphic', 'pattern', 'floral', 'stripe', 'striped',
        'plaid', 'camo', 'camouflage', 'tie-dye', 'tie dye', 'leopard', 'animal',
        'paisley', 'tropical', 'abstract', 'geometric', 'polka dot', 'checkered',
        'check', 'tartan', 'houndstooth', 'argyle', 'logo', 'embroidered',
        'embroidery', 'novelty', 'snake', 'zebra', 'cheetah', 'destroy',
        'distressed', 'destructed', 'patchwork', 'colorblock',
    ]
    combined = f"{title} {color} {text}".lower()
    return any(kw in combined for kw in kws)


def calc_discount(original, current):
    if not original or not current:
        return ""
    try:
        o = float(str(original).replace('$', '').replace(',', ''))
        c = float(str(current).replace('$', '').replace(',', ''))
        return f"{((o - c) / o) * 100:.1f}%" if o > c > 0 else ""
    except (ValueError, ZeroDivisionError):
        return ""


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 1.5
        self.current_delay = 1.5
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    🐌 Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    🏎️  Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    🔄 Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'URL', 'Parent TCIN', 'Title', 'Brand', 'Owned Brand', 'Brand Type',
    'Color', 'Color Swatch URL', 'Color Image URL', 'Color Buy URL', 'Color TCIN',
    'All Colors in Product', 'Total Colors',
    'Current Price', 'Price Type', 'Original Price', 'Save $', 'Save %', 'Discount %',
    'Price Min', 'Price Max',
    'Color Current Price', 'Color Price Type', 'Color Current Retail', 'Color Reg Retail',
    'Current Retail Min', 'Reg Retail Max',
    'Color Sizes', '# Sizes', 'Color Size Groups', 'All Size Variants', '# SKUs for Color',
    'Sizing', 'Material', 'Fabric Parsed', '% Cotton', '% Natural Fiber', '% Recycled',
    'Inseam Length', 'Garment Length', 'Rise', 'Fit', 'Closure Style',
    'Fabric Name', 'Fabric Weight Type', 'Stretch', 'Hem Style',
    'Garment Details', 'Care and Cleaning', 'Origin',
    'Rating', 'Review Count', 'Stars 5', 'Stars 4', 'Stars 3', 'Stars 2', 'Stars 1',
    'Comfort Rating', 'Sizing Rating', 'Quality Rating', 'Value Rating', 'Sizing Feedback',
    'Question Count', 'Bought Last Month',
    'Description', 'Highlights', 'Non-Basic',
    'Department', 'Dept ID', 'Class ID', 'Product Type',
    'Breadcrumb', 'Category 1', 'Category 2', 'Category 3', 'Category 4', 'Category 5',
    'Primary Image URL', 'Image Count',
    'Package Weight (lbs)', 'Package Dimensions',
    'Return Policy', 'Shop The Look TCINs',
    'Canonical URL', 'Meta Description',
    'Retries', 'Timestamp', 'Error',
]

EXCEL_FIELDS = [
    'url', 'parent_tcin', 'title', 'brand', 'owned_brand', 'brand_type',
    'color', 'color_swatch_url', 'color_image_url', 'color_buy_url', 'color_tcin',
    'all_colors', 'total_colors_in_product',
    'current_price', 'price_type', 'original_price', 'save_dollar', 'save_percent', 'discount_percent',
    'price_min', 'price_max',
    'color_current_price', 'color_price_type', 'color_current_retail', 'color_reg_retail',
    'current_retail_min', 'reg_retail_max',
    'color_sizes', 'color_num_sizes', 'color_size_groups', 'color_all_size_variants', 'color_num_skus',
    'sizing', 'material', 'fabric_parsed', 'pct_cotton', 'pct_natural', 'pct_recycled',
    'inseam_length', 'garment_length', 'rise', 'fit', 'closure_style',
    'fabric_name', 'fabric_weight_type', 'stretch', 'hem_style',
    'garment_details', 'care_and_cleaning', 'origin',
    'rating_avg', 'review_count', 'stars_5', 'stars_4', 'stars_3', 'stars_2', 'stars_1',
    'rating_comfort', 'rating_sizing', 'rating_quality', 'rating_value', 'sizing_feedback',
    'question_count', 'bought_last_month',
    'description', 'highlights', 'non_basic',
    'department', 'department_id', 'class_id', 'product_type',
    'breadcrumb', 'category_1', 'category_2', 'category_3', 'category_4', 'category_5',
    'og_image_url', 'image_count',
    'package_weight_lbs', 'package_dimensions',
    'return_policy', 'shop_the_look_tcins',
    'canonical_url', 'meta_description',
    'retries', 'timestamp', 'error',
]


def load_progress(d):
    p = os.path.join(d, "target_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "target_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            field_map = dict(zip(EXCEL_HEADERS, EXCEL_FIELDS))
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i] in field_map:
                        rec[field_map[headers[i]]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   📂 Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   ⚠️  Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    with open(os.path.join(d, "target_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    out = os.path.join(d, "target_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Target PDP Results'

    hfill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_FIELDS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    ws.freeze_panes = 'A2'
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    await ctx.add_cookies([
        {"name": "GuestLocation", "value": "90210", "domain": ".target.com", "path": "/"},
        {"name": "fiatsCookie", "value": "DSI_2788|DSN_Beverly Hills|DSZ_90210", "domain": ".target.com", "path": "/"},
        {"name": "sapphire", "value": "1", "domain": ".target.com", "path": "/"},
    ])
    return ctx


# ── Scraping ────────────────────────────────────────────────────────────────

async def extract_js_pricing(page):
    """
    Evaluate JavaScript on the page to extract per-TCIN pricing from
    Target's client-side data store. Returns {tcin: {'formatted_price': str,
    'price_type': str, 'current_retail': float, 'reg_retail': float}}.
    Falls back to empty dict on any error.
    """
    try:
        result = await page.evaluate('''() => {
            const pricing = {};
            try {
                // Target stores product data in __TGT_DATA__ or window.__PRELOADED_QUERIES__
                // Walk through all script tags looking for the product JSON
                const scripts = document.querySelectorAll('script');
                for (const script of scripts) {
                    const text = script.textContent || '';
                    // Look for the preloaded queries pattern
                    if (text.includes('__TGT_DATA__') || text.includes('Product') && text.includes('tcin')) {
                        // Extract all tcin + price blocks using a targeted approach
                        const tcinPriceRegex = /"tcin":"(\\d+)"[^}]*?"price":\\s*\\{[^}]*?"formatted_current_price":"([^"]+)"[^}]*?(?:"formatted_current_price_type":"([^"]*)")?[^}]*?(?:"current_retail":(\\d+\\.?\\d*))?[^}]*?(?:"reg_retail":(\\d+\\.?\\d*))?/g;
                        let match;
                        while ((match = tcinPriceRegex.exec(text)) !== null) {
                            const [, tcin, fmtPrice, priceType, currentRetail, regRetail] = match;
                            if (tcin && fmtPrice && !pricing[tcin]) {
                                pricing[tcin] = {
                                    formatted_price: fmtPrice,
                                    price_type: priceType || '',
                                    current_retail: currentRetail ? parseFloat(currentRetail) : null,
                                    reg_retail: regRetail ? parseFloat(regRetail) : null,
                                };
                            }
                        }
                    }
                }

                // Also try to find pricing in the React fiber tree via __PRELOADED_QUERIES__
                if (window.__TGT_DATA__) {
                    const data = window.__TGT_DATA__;
                    const walk = (obj) => {
                        if (!obj || typeof obj !== 'object') return;
                        if (obj.tcin && obj.price && obj.price.formatted_current_price) {
                            if (!pricing[obj.tcin]) {
                                pricing[obj.tcin] = {
                                    formatted_price: obj.price.formatted_current_price,
                                    price_type: obj.price.formatted_current_price_type || '',
                                    current_retail: obj.price.current_retail || null,
                                    reg_retail: obj.price.reg_retail || null,
                                };
                            }
                        }
                        for (const key of Object.keys(obj)) {
                            if (typeof obj[key] === 'object') walk(obj[key]);
                        }
                    };
                    walk(data);
                }
            } catch(e) {}
            return pricing;
        }''')
        return result if isinstance(result, dict) else {}
    except Exception:
        return {}


async def scrape_target_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single Target PDP page. Returns a LIST of result dicts (one per color)."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
        status = resp.status if resp else 0

        if status == 403 or status == 429 or status == 503:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        page_url = page.url.lower()
        if 'captcha' in page_url or 'blocked' in page_url or 'denied' in page_url:
            await stats.inc_blocked()
            wait_time = 60 * (retries + 1) + random.uniform(10, 30)
            print(f"    ⚠️  Blocked on {url[:60]}... backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
            if retries < MAX_RETRIES:
                await asyncio.sleep(wait_time)
                return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
            return [{'url': url, 'error': 'Blocked/CAPTCHA', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        await asyncio.sleep(random.uniform(1.5, 3.0))
        html_content = await page.content()

        if 'enrichment' not in html_content and 'product_description' not in html_content:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': 'No product data in page', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Extract per-TCIN pricing from JS before parsing HTML
        js_pricing = await extract_js_pricing(page)

        color_rows = parse_target_pdp(html_content, url, js_pricing=js_pricing)
        for row in color_rows:
            row['retries'] = retries
            row['url'] = url

        await stats.inc_ok(len(color_rows))
        await throttle.on_ok()
        return color_rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_target_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  ⏰ W{wid}: 30-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url[:80]}...")

        await throttle.wait()
        color_rows = await scrape_target_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(color_rows)
            progress['processed'].add(url)

        colors_found = len(color_rows)
        has_error = any(r.get('error') for r in color_rows)
        if has_error:
            print(f"    ❌ W{wid}: Error — {color_rows[0].get('error', '')[:80]}")
        else:
            print(f"    ✅ W{wid}: {colors_found} color rows extracted")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  💾 Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   📊 {len(results)} existing rows, {len(remaining)} URLs remaining")

    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-gpu",
    ])

    async def block_resources(route):
        await route.abort()

    all_pages = []
    all_contexts = []
    print(f"🌐 Launching {NUM_CONTEXTS} contexts × {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            await pg.route('**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,ico,webp}', block_resources)
            all_pages.append(pg)

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\n⚠️  Browser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  💾 Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    uf = os.path.join(sdir, "pdp_urls.csv")
    if not os.path.exists(uf):
        print(f"❌ pdp_urls.csv not found at {sdir}")
        print(f"   Create pdp_urls.csv with one Target URL per line (no header)")
        return

    with open(uf, encoding="utf-8-sig") as f:
        urls = [l.strip() for l in f if l.strip() and l.strip().startswith('http')]
    print(f"🎯 Loaded {len(urls)} Target URLs")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 20

    crash_count = 0
    while True:
        remaining = [u for u in urls if u not in progress['processed']]
        if not remaining:
            print("✅ All URLs processed!")
            break
        print(f"\n{'=' * 60}")
        print(f"▶️  Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
        print(f"{'=' * 60}")

        was_crash = False
        try:
            pw = await async_playwright().start()
            finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
            try:
                await pw.stop()
            except Exception:
                pass
            if finished:
                break
        except Exception as e:
            print(f"\n⚠️  Playwright/browser error: {str(e)[:150]}")
            was_crash = True
            try:
                await pw.stop()
            except Exception:
                pass

        throttle.reset()
        save_progress(progress, sdir)
        save_to_excel(results, sdir)

        if was_crash:
            crash_count += 1
            if crash_count >= max_crashes:
                print(f"❌ Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                break
            wait = min(30, 10 * crash_count)
            print(f"🔄 Crash restart in {wait}s... (crash #{crash_count})")
            await asyncio.sleep(wait)
        else:
            crash_count = 0
            print(f"🔄 Fresh restart in 5s...")
            await asyncio.sleep(5)

    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"✅ Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"   Results: {out}")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
