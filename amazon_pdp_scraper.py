#!/usr/bin/env python3
"""
Production Amazon PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  amazon.csv  (URLs in first column 'a-link-normal href')
Output: amazon_pdp_results.xlsx + amazon_pdp_progress.json

Architecture: 3 browser contexts × 1 tab each = 3 parallel workers
              20-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

Extracts from Amazon PDP pages via Playwright evaluate() to produce
one row per color variant with pricing, specs, ratings, and fabric.
"""

import asyncio, json, os, random, re, time, csv
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 3
TABS_PER_CTX = 1
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 3 workers (small URL set)
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 20 * 60  # 20 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# ── Amazon Owned Brands ────────────────────────────────────────────────────
AMAZON_OWNED_BRANDS = {
    'amazon essentials', 'the drop',
}


def classify_brand(brand_name):
    """Returns 'OB' if Amazon owned brand, 'NB' for national brand."""
    if brand_name and brand_name.strip().lower() in AMAZON_OWNED_BRANDS:
        return 'OB'
    return 'NB'


# ── Non-Basic Detection ───────────────────────────────────────────────────
NON_BASIC_KEYWORDS = [
    'print', 'printed', 'graphic', 'pattern', 'floral', 'stripe', 'striped',
    'plaid', 'camo', 'camouflage', 'tie-dye', 'tie dye', 'leopard', 'logo',
    'embroidered', 'embroidery', 'novelty', 'character', 'varsity', 'retro',
    'animal', 'abstract', 'geometric', 'paisley', 'polka', 'check', 'checked',
    'tropical', 'botanical', 'ditsy', 'patchwork', 'colorblock', 'ombre',
]


def is_non_basic(title, colors_str, bullets_str):
    """Returns 1 if product is non-basic (graphic/patterned), 0 otherwise."""
    combined = f"{title} {colors_str} {bullets_str}".lower()
    for kw in NON_BASIC_KEYWORDS:
        if kw in combined:
            return 1
    return 0


# ── Rise / Leg / Fit Parsing from Title ────────────────────────────────────

def parse_rise(title):
    """Extract rise from product title."""
    t = title.lower()
    if 'super high' in t or 'super-high' in t:
        return 'Super High Rise'
    if 'high-rise' in t or 'high rise' in t:
        return 'High Rise'
    if 'mid-rise' in t or 'mid rise' in t:
        return 'Mid Rise'
    if 'low-rise' in t or 'low rise' in t:
        return 'Low Rise'
    return ''


def parse_leg_shape(title):
    """Extract leg shape from product title."""
    t = title.lower()
    if 'wide' in t and 'leg' in t:
        return 'Wide Leg'
    if 'wide' in t:
        return 'Wide Leg'
    if 'flare' in t:
        return 'Flare'
    if 'bootcut' in t or 'boot cut' in t or 'boot-cut' in t:
        return 'Bootcut'
    if 'straight' in t:
        return 'Straight'
    if 'slim' in t and 'straight' not in t:
        return 'Slim'
    if 'skinny' in t:
        return 'Skinny'
    if 'tapered' in t or 'taper' in t:
        return 'Tapered'
    if 'relaxed' in t:
        return 'Relaxed'
    if 'boyfriend' in t:
        return 'Boyfriend'
    if 'jegging' in t:
        return 'Jegging'
    if 'crop' in t:
        return 'Crop'
    return ''


def parse_fit(title):
    """Extract fit from product title."""
    t = title.lower()
    if 'relaxed' in t:
        return 'Relaxed'
    if 'slim' in t:
        return 'Slim'
    if 'curvy' in t:
        return 'Curvy'
    if 'regular' in t:
        return 'Regular'
    return ''


# ── Material Parsing ──────────────────────────────────────────────────────

def parse_material(fabric_raw):
    """
    Parse fabric composition string and return (fabric_parsed, pct_cotton, pct_natural).
    Handles: '98% Cotton, 2% Spandex' or 'Front: 90% Cotton | Back: 88% Cotton'
    Only sums first panel to avoid double-counting.
    """
    if not fabric_raw:
        return ('', '', '')

    clean = fabric_raw.strip()
    # Only take first panel if multi-panel
    first_panel = clean.split('|')[0].strip()

    cotton_pct = 0
    natural_pct = 0
    # Pattern: number% MaterialName
    matches = re.findall(r'(\d+(?:\.\d+)?)\s*%\s*([A-Za-z\s]+)', first_panel)
    for pct_str, name in matches:
        pct = float(pct_str)
        name_lower = name.strip().lower()
        if 'cotton' in name_lower:
            cotton_pct += pct
        # Natural fibers
        if any(f in name_lower for f in ['cotton', 'linen', 'silk', 'wool', 'hemp', 'bamboo', 'tencel', 'modal', 'lyocell', 'cashmere']):
            natural_pct += pct

    cotton_str = f"{cotton_pct:.0f}%" if cotton_pct > 0 else ''
    natural_str = f"{natural_pct:.0f}%" if natural_pct > 0 else ''
    return (clean, cotton_str, natural_str)


# ── Amazon Page Extraction (via Playwright evaluate) ──────────────────────

# Quick price-only extraction — used after clicking each color swatch
EXTRACT_PRICE_JS = """
() => {
    function extractPriceFromEl(priceEl) {
        if (!priceEl) return '';
        const offscreen = priceEl.querySelector('.a-offscreen');
        if (offscreen) {
            const t = offscreen.textContent.trim();
            if (t && /^\\$[\\d,.]+$/.test(t)) return t;
        }
        const whole = priceEl.querySelector('.a-price-whole');
        const fraction = priceEl.querySelector('.a-price-fraction');
        if (whole) {
            return '$' + whole.textContent.replace(/[^\\d]/g, '') + '.' + (fraction ? fraction.textContent.replace(/[^\\d]/g, '') : '00');
        }
        return '';
    }

    const result = {};

    // Current price
    result.current_price = '';
    const priceSelectors = [
        '#corePrice_feature_div .priceToPay',
        '#corePriceDisplay_desktop_feature_div .priceToPay',
        '.apexPriceToPay',
        '.priceToPay'
    ];
    for (const sel of priceSelectors) {
        const el = document.querySelector(sel);
        if (el) {
            const p = extractPriceFromEl(el);
            if (p) { result.current_price = p; break; }
        }
    }
    if (!result.current_price) {
        const allPrices = document.querySelectorAll('.a-price:not([data-a-strike])');
        for (const ap of allPrices) {
            const p = extractPriceFromEl(ap);
            if (p) { result.current_price = p; break; }
        }
    }

    // Original / strike-through price
    result.original_price = '';
    const strikeEl = document.querySelector('.a-price[data-a-strike] .a-offscreen');
    if (strikeEl) {
        const t = strikeEl.textContent.trim();
        if (t && /^\\$[\\d,.]+$/.test(t)) result.original_price = t;
    }
    if (!result.original_price) {
        const basisEl = document.querySelector('#corePriceDisplay_desktop_feature_div .basisPrice .a-offscreen');
        if (basisEl) {
            const t = basisEl.textContent.trim();
            if (t && /^\\$[\\d,.]+$/.test(t)) result.original_price = t;
        }
    }

    // Also grab the selected color name to confirm which color we're seeing
    result.selected_color = '';
    const colorLabelEl = document.querySelector(
        '#inline-twister-expanded-dimension-text-color_name .a-color-secondary, ' +
        '#variation_color_name .selection'
    );
    if (colorLabelEl) {
        let selText = colorLabelEl.textContent.trim();
        selText = selText.replace(/^Color:\\s*/i, '').trim();
        if (selText) result.selected_color = selText;
    }

    return result;
}
"""

EXTRACT_JS = """
() => {
    const result = {};

    // --- Title ---
    const titleEl = document.getElementById('productTitle');
    result.title = titleEl ? titleEl.textContent.trim() : '';

    // --- Brand ---
    const brandEl = document.getElementById('bylineInfo');
    result.brand = '';
    if (brandEl) {
        let bt = brandEl.textContent.trim();
        bt = bt.replace(/^Visit the\\s+/i, '').replace(/^Brand:\\s*/i, '').replace(/\\s+Store$/i, '');
        result.brand = bt;
    }

    // --- ASIN ---
    result.asin = '';
    const asinInput = document.querySelector('input[name="ASIN"]');
    if (asinInput) result.asin = asinInput.value;
    if (!result.asin) {
        // Try URL
        const m = window.location.pathname.match(/\\/dp\\/([A-Z0-9]{10})/);
        if (m) result.asin = m[1];
    }

    // --- Price extraction ---
    function extractPriceFromEl(priceEl) {
        if (!priceEl) return '';
        const offscreen = priceEl.querySelector('.a-offscreen');
        if (offscreen) {
            const t = offscreen.textContent.trim();
            if (t && /^\\$[\\d,.]+$/.test(t)) return t;
        }
        const whole = priceEl.querySelector('.a-price-whole');
        const fraction = priceEl.querySelector('.a-price-fraction');
        if (whole) {
            return '$' + whole.textContent.replace(/[^\\d]/g, '') + '.' + (fraction ? fraction.textContent.replace(/[^\\d]/g, '') : '00');
        }
        return '';
    }

    // Current price — try multiple selectors in priority order
    result.current_price = '';
    const priceSelectors = [
        '#corePrice_feature_div .priceToPay',
        '#corePriceDisplay_desktop_feature_div .priceToPay',
        '.apexPriceToPay',
        '.priceToPay'
    ];
    for (const sel of priceSelectors) {
        const el = document.querySelector(sel);
        if (el) {
            const p = extractPriceFromEl(el);
            if (p) { result.current_price = p; break; }
        }
    }
    // Fallback: first .a-price without data-a-strike
    if (!result.current_price) {
        const allPrices = document.querySelectorAll('.a-price:not([data-a-strike])');
        for (const ap of allPrices) {
            const p = extractPriceFromEl(ap);
            if (p) { result.current_price = p; break; }
        }
    }

    // Original / strike-through price
    result.original_price = '';
    const strikeEl = document.querySelector('.a-price[data-a-strike] .a-offscreen');
    if (strikeEl) {
        const t = strikeEl.textContent.trim();
        if (t && /^\\$[\\d,.]+$/.test(t)) result.original_price = t;
    }
    // Also check "Typical:" or "List:" price
    if (!result.original_price) {
        const basisEl = document.querySelector('#corePriceDisplay_desktop_feature_div .basisPrice .a-offscreen');
        if (basisEl) {
            const t = basisEl.textContent.trim();
            if (t && /^\\$[\\d,.]+$/.test(t)) result.original_price = t;
        }
    }

    // --- Rating ---
    result.rating = '';
    const ratingEl = document.querySelector('#acrPopover .a-icon-alt');
    if (ratingEl) {
        const m = ratingEl.textContent.match(/([\\d.]+)/);
        if (m) result.rating = m[1];
    }

    // --- Review count ---
    result.review_count = '';
    const rcEl = document.getElementById('acrCustomerReviewCount');
    if (rcEl) {
        const m = rcEl.textContent.match(/([\\d,]+)/);
        if (m) result.review_count = m[1].replace(/,/g, '');
    }

    // --- Bought past month ---
    result.bought_past_month = '';
    const bpmEl = document.getElementById('social-proofing-faceout-title-tk_bought');
    if (bpmEl) {
        result.bought_past_month = bpmEl.textContent.trim();
    }

    // --- Breadcrumbs ---
    result.breadcrumb = '';
    const bcEls = document.querySelectorAll('#wayfinding-breadcrumbs_container .a-link-normal');
    if (bcEls.length) {
        result.breadcrumb = Array.from(bcEls).map(e => e.textContent.trim()).filter(t => t).join(' > ');
    }

    // --- Colors ---
    result.colors = [];
    result.selected_color = '';

    // Helper: clean a color name and deduplicate
    function cleanColor(name) {
        if (!name) return '';
        name = name.trim().replace(/\\s+/g, ' ');
        if (!name || name.length < 2) return '';
        // Filter out size-like values and pure numbers
        if (/^\\d+[xX]?\\s*(short|regular|long|tall)?$/i.test(name)) return '';
        return name;
    }

    // Get selected color from the "Color:" label on page
    // Try inline twister label first, then classic variation
    const colorLabelEl = document.querySelector(
        '#inline-twister-expanded-dimension-text-color_name .a-color-secondary, ' +
        '#variation_color_name .selection, ' +
        'label[for*="color"] + span'
    );
    if (colorLabelEl) {
        let selText = colorLabelEl.textContent.trim();
        // Remove "Color:" prefix if present
        selText = selText.replace(/^Color:\\s*/i, '').trim();
        if (selText) result.selected_color = selText;
    }

    // PRIMARY METHOD: Inline twister ul li img (confirmed working on current Amazon)
    // The inline twister has a <ul> with role="radiogroup" inside #tp-inline-twister-dim-values-container
    // Each color swatch is an <li> containing <img> elements with alt = color name
    // Each color appears TWICE (thumbnail + display), so we deduplicate
    const inlineTwister = document.querySelector('#tp-inline-twister-dim-values-container');
    if (inlineTwister) {
        // Find the color radiogroup specifically (not size)
        const colorGroup = inlineTwister.querySelector('ul[data-a-button-group*="color_name"]');
        if (colorGroup) {
            const imgs = colorGroup.querySelectorAll('li img');
            imgs.forEach(function(img) {
                const name = cleanColor(img.alt || '');
                if (name && result.colors.indexOf(name) === -1) result.colors.push(name);
            });
        }
        // Fallback: get ALL img alts from inline twister and deduplicate
        if (!result.colors.length) {
            const allImgs = inlineTwister.querySelectorAll('ul li img');
            allImgs.forEach(function(img) {
                const name = cleanColor(img.alt || '');
                if (name && result.colors.indexOf(name) === -1) result.colors.push(name);
            });
        }
    }

    // FALLBACK 1: Classic #variation_color_name container (older Amazon layout)
    if (!result.colors.length) {
        const colorContainer = document.getElementById('variation_color_name');
        if (colorContainer) {
            if (!result.selected_color) {
                const selSpan = colorContainer.querySelector('.selection');
                if (selSpan) result.selected_color = selSpan.textContent.trim();
            }
            const lis = colorContainer.querySelectorAll('li[data-defaultasin]');
            if (lis.length) {
                lis.forEach(function(li) {
                    const img = li.querySelector('img');
                    let name = '';
                    if (img && img.alt) name = img.alt.trim();
                    if (!name) {
                        const btn = li.querySelector('button');
                        if (btn && btn.getAttribute('aria-label')) name = btn.getAttribute('aria-label').trim();
                    }
                    if (!name) name = li.textContent.trim();
                    const clean = cleanColor(name);
                    if (clean && result.colors.indexOf(clean) === -1) result.colors.push(clean);
                });
            }
            if (!result.colors.length) {
                colorContainer.querySelectorAll('li').forEach(function(li) {
                    const img = li.querySelector('img');
                    let name = '';
                    if (img && img.alt) name = img.alt.trim();
                    if (!name) name = li.textContent.trim().substring(0, 60);
                    const clean = cleanColor(name);
                    if (clean && result.colors.indexOf(clean) === -1) result.colors.push(clean);
                });
            }
        }
    }

    // FALLBACK 2: Classic #twister section
    if (!result.colors.length) {
        const twister = document.getElementById('twister');
        if (twister) {
            const swatchImgs = twister.querySelectorAll('li img, button img');
            swatchImgs.forEach(function(img) {
                const name = cleanColor(img.alt || '');
                if (name && result.colors.indexOf(name) === -1) result.colors.push(name);
            });
        }
    }

    // --- Sizes from #variation_size_name ---
    result.sizes = [];
    const sizeContainer = document.getElementById('variation_size_name');
    if (sizeContainer) {
        const lis = sizeContainer.querySelectorAll('li');
        lis.forEach(li => {
            const btn = li.querySelector('button');
            let name = '';
            if (btn && btn.getAttribute('aria-label')) name = btn.getAttribute('aria-label').trim();
            if (!name) name = li.textContent.trim();
            if (name) result.sizes.push(name);
        });
    }

    // --- Feature bullets ---
    result.feature_bullets = [];
    const bulletEls = document.querySelectorAll('#feature-bullets .a-list-item');
    bulletEls.forEach(el => {
        const t = el.textContent.trim();
        if (t && t.length > 5 && !t.startsWith('›')) result.feature_bullets.push(t);
    });

    // --- Product details table ---
    result.product_details = {};
    // Method 1: #productDetails_detailBullets_sections1
    const detailRows = document.querySelectorAll('#productDetails_detailBullets_sections1 tr, #detailBullets_feature_div .content li');
    detailRows.forEach(row => {
        const th = row.querySelector('th, .a-text-bold');
        const td = row.querySelector('td, span:not(.a-text-bold)');
        if (th && td) {
            const key = th.textContent.trim().replace(/[:\\s]+$/g, '').replace(/\\u200e/g, '');
            const val = td.textContent.trim().replace(/\\u200e/g, '');
            if (key && val) result.product_details[key] = val;
        }
    });
    // Method 2: #detailBulletsWrapper_feature_div
    const detailBullets = document.querySelectorAll('#detailBulletsWrapper_feature_div .a-list-item');
    detailBullets.forEach(el => {
        const bold = el.querySelector('.a-text-bold');
        if (bold) {
            const key = bold.textContent.trim().replace(/[:\\s]+$/g, '').replace(/\\u200e/g, '');
            // Get text after the bold element
            let val = el.textContent.replace(bold.textContent, '').trim().replace(/\\u200e/g, '');
            if (key && val) result.product_details[key] = val;
        }
    });

    // --- Bestseller rank ---
    result.bestseller_rank = '';
    const bsrEl = document.querySelector('#SalesRank, #detailBulletsWrapper_feature_div');
    if (bsrEl) {
        const bsrText = bsrEl.textContent;
        const bsrMatch = bsrText.match(/#([\\d,]+)\\s+in\\s+([^(\\n]+)/);
        if (bsrMatch) result.bestseller_rank = '#' + bsrMatch[1] + ' in ' + bsrMatch[2].trim();
    }
    // Also check product details
    if (!result.bestseller_rank && result.product_details['Best Sellers Rank']) {
        result.bestseller_rank = result.product_details['Best Sellers Rank'].split('\\n')[0].trim();
    }

    // --- Department ---
    result.department = result.product_details['Department'] || '';

    // --- Manufacturer ---
    result.manufacturer = result.product_details['Manufacturer'] || '';

    // --- Fabric / Material ---
    result.fabric_raw = '';
    // Check product details first
    const fabricKeys = ['Fabric type', 'Material', 'Fabric Type', 'Material type', 'Outer Material'];
    for (const k of fabricKeys) {
        if (result.product_details[k]) { result.fabric_raw = result.product_details[k]; break; }
    }
    // Fallback: feature bullets
    if (!result.fabric_raw) {
        for (const b of result.feature_bullets) {
            const m = b.match(/(\\d+%\\s*\\w+[^.]*)/i);
            if (m && /cotton|polyester|spandex|elastane|rayon|nylon|linen|lyocell|tencel/i.test(m[1])) {
                result.fabric_raw = m[1];
                break;
            }
        }
    }

    // --- Image count ---
    result.image_count = 0;
    const thumbs = document.querySelectorAll('#altImages .a-button-thumbnail img');
    result.image_count = thumbs.length || 1;

    // --- Has Prime ---
    result.has_prime = !!document.querySelector('#prime-badge, .a-icon-prime');

    // --- Deal badge ---
    result.deal_badge = '';
    const dealEl = document.querySelector('#dealBadge_feature_div .a-badge-text, .deal-badge');
    if (dealEl) result.deal_badge = dealEl.textContent.trim();

    // --- Delivery ---
    result.delivery = '';
    const delEl = document.querySelector('#deliveryBlockMessage .a-text-bold, #mir-layout-DELIVERY_BLOCK .a-text-bold');
    if (delEl) result.delivery = delEl.textContent.trim();

    // --- Pack size ---
    result.pack_size = '';
    const packKeys = ['Number of Items', 'Package Quantity', 'Item Package Quantity', 'Count', 'Unit Count'];
    for (const k of packKeys) {
        if (result.product_details[k]) { result.pack_size = result.product_details[k]; break; }
    }

    return result;
}
"""


def parse_amazon_pdp(raw_data, url):
    """
    Take the raw JS extraction result and build a LIST of row dicts — one per color.
    If no colors found, returns single row.
    """
    rows = []

    try:
        title = raw_data.get('title', '')
        brand = raw_data.get('brand', '')
        asin = raw_data.get('asin', '')

        # Pricing
        current_price_str = raw_data.get('current_price', '')
        original_price_str = raw_data.get('original_price', '')

        def price_to_float(p):
            if not p:
                return None
            return float(re.sub(r'[^0-9.]', '', p)) if re.search(r'[\d]', p) else None

        current_price = price_to_float(current_price_str)
        original_price = price_to_float(original_price_str)

        on_sale = 0
        discount_pct = ''
        if original_price and current_price and original_price > current_price:
            on_sale = 1
            discount_pct = f"{((original_price - current_price) / original_price) * 100:.1f}%"

        if not original_price:
            original_price = current_price
            original_price_str = current_price_str

        # Rating
        rating = raw_data.get('rating', '')
        review_count = raw_data.get('review_count', '')
        bought_past_month = raw_data.get('bought_past_month', '')

        # Category
        breadcrumb = raw_data.get('breadcrumb', '')

        # Colors / Sizes
        colors = raw_data.get('colors', [])
        sizes = raw_data.get('sizes', [])
        selected_color = raw_data.get('selected_color', '')
        total_colors = len(colors) if colors else 1
        total_sizes = len(sizes)
        sizes_list = ', '.join(sizes) if sizes else ''

        # Rise / Leg / Fit from title
        rise = parse_rise(title)
        leg_shape = parse_leg_shape(title)
        fit = parse_fit(title)

        # Material
        fabric_raw = raw_data.get('fabric_raw', '')
        fabric_parsed, pct_cotton, pct_natural = parse_material(fabric_raw)

        # Feature bullets
        bullets = raw_data.get('feature_bullets', [])
        bullets_str = ' | '.join(bullets) if bullets else ''

        # Non-basic detection
        colors_str = ', '.join(colors) if colors else selected_color
        non_basic = is_non_basic(title, colors_str, bullets_str)

        # Other fields
        bestseller_rank = raw_data.get('bestseller_rank', '')
        department = raw_data.get('department', '')
        manufacturer = raw_data.get('manufacturer', '')
        image_count = raw_data.get('image_count', 0)
        has_prime = 'Yes' if raw_data.get('has_prime') else 'No'
        deal_badge = raw_data.get('deal_badge', '')
        delivery = raw_data.get('delivery', '')
        pack_size = raw_data.get('pack_size', '')
        product_details_raw = json.dumps(raw_data.get('product_details', {}))

        # Per-color prices (populated by Phase 2 clicking)
        per_color_prices = raw_data.get('per_color_prices', {})

        # Build rows — one per color, with per-color pricing
        def make_row(color_name):
            # Use per-color price if available, otherwise fall back to default page price
            color_cur_str = current_price_str
            color_orig_str = original_price_str
            color_cur = current_price
            color_orig = original_price
            color_on_sale = on_sale
            color_discount = discount_pct

            if color_name in per_color_prices:
                pcp = per_color_prices[color_name]
                if pcp.get('current_price'):
                    color_cur_str = pcp['current_price']
                    color_cur = price_to_float(color_cur_str)
                if pcp.get('original_price'):
                    color_orig_str = pcp['original_price']
                    color_orig = price_to_float(color_orig_str)
                # Recalculate discount for this color
                color_on_sale = 0
                color_discount = ''
                if color_orig and color_cur and color_orig > color_cur:
                    color_on_sale = 1
                    color_discount = f"{((color_orig - color_cur) / color_orig) * 100:.1f}%"
                if not color_orig:
                    color_orig = color_cur
                    color_orig_str = color_cur_str

            return {
                'url': url,
                'asin': asin,
                'product_name': title,
                'brand': brand,
                'brand_type': classify_brand(brand),
                'color': color_name,
                'current_price': color_cur,
                'current_price_formatted': color_cur_str,
                'original_price': color_orig,
                'original_price_formatted': color_orig_str,
                'on_sale': color_on_sale,
                'discount_pct': color_discount,
                'rise': rise,
                'leg_shape': leg_shape,
                'fit': fit,
                'fabric_raw': fabric_raw,
                'fabric_parsed': fabric_parsed,
                'pct_cotton': pct_cotton,
                'pct_natural_fiber': pct_natural,
                'non_basic': non_basic,
                'breadcrumb': breadcrumb,
                'department': department,
                'manufacturer': manufacturer,
                'total_colors': total_colors,
                'total_sizes': total_sizes,
                'sizes_list': sizes_list,
                'average_rating': rating,
                'review_count': review_count,
                'bought_past_month': bought_past_month,
                'bestseller_rank': bestseller_rank,
                'image_count': image_count,
                'has_prime': has_prime,
                'deal_badge': deal_badge,
                'delivery': delivery,
                'pack_size': pack_size,
                'feature_bullets': bullets_str,
                'product_details_raw': product_details_raw,
                'timestamp': datetime.now().isoformat(),
            }

        if colors:
            for c in colors:
                rows.append(make_row(c))
        else:
            rows.append(make_row(selected_color))

    except Exception as e:
        rows = [{
            'url': url,
            'error': f'Parse error: {str(e)[:200]}',
            'timestamp': datetime.now().isoformat(),
        }]

    return rows


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 3.0     # slightly higher for Amazon
        self.current_delay = 3.0
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'url', 'asin', 'product_name', 'brand', 'brand_type', 'color',
    'current_price', 'current_price_formatted', 'original_price', 'original_price_formatted',
    'on_sale', 'discount_pct',
    'rise', 'leg_shape', 'fit',
    'fabric_raw', 'fabric_parsed', 'pct_cotton', 'pct_natural_fiber', 'non_basic',
    'breadcrumb', 'department', 'manufacturer',
    'total_colors', 'total_sizes', 'sizes_list',
    'average_rating', 'review_count', 'bought_past_month',
    'bestseller_rank', 'image_count', 'has_prime', 'deal_badge', 'delivery', 'pack_size',
    'feature_bullets', 'product_details_raw',
    'retries', 'timestamp', 'error',
]


def load_progress(d):
    """Load progress and results from files."""
    p = os.path.join(d, "amazon_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "amazon_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        rec[headers[i]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    """Save progress to JSON file."""
    with open(os.path.join(d, "amazon_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    """Save results to Excel file with formatting."""
    out = os.path.join(d, "amazon_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Amazon PDP Results'

    # Dark header styling
    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_HEADERS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    # Freeze header
    ws.freeze_panes = 'A2'

    # Column widths
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Create a browser context with anti-detection and US locale/currency cookies."""
    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    # Amazon cookies for US pricing
    await ctx.add_cookies([
        {"name": "sp-cdn", "value": '"L5Z9:US"', "domain": ".amazon.com", "path": "/"},
        {"name": "lc-main", "value": "en_US", "domain": ".amazon.com", "path": "/"},
        {"name": "i18n-prefs", "value": "USD", "domain": ".amazon.com", "path": "/"},
    ])
    return ctx


async def set_us_zip(page):
    """
    Set the delivery zip code to 90402 (Santa Monica) for US pricing.
    Amazon sometimes shows local currency without this.
    """
    try:
        # Click location popup trigger
        loc_link = page.locator('#nav-global-location-popover-link')
        if await loc_link.count() > 0:
            await loc_link.click()
            await asyncio.sleep(1.5)

            # Type zip code char-by-char (Amazon validates keystroke-by-keystroke)
            # IMPORTANT: use .type() NOT .fill() — Amazon validates each keystroke
            zi = page.locator('#GLUXZipUpdateInput')
            if await zi.count() > 0:
                await zi.click()
                # Select all existing text and delete, then type char-by-char
                await zi.press('Control+a')
                await zi.press('Backspace')
                await zi.type('90402', delay=50)
                await asyncio.sleep(0.5)

                # Click Apply (use wrapper, not inner submit)
                apply_btn = page.locator('#GLUXZipUpdate')
                if await apply_btn.count() > 0:
                    await apply_btn.click(force=True)
                    await asyncio.sleep(1.5)

                    # Click "Continue" on confirmation (NOT "Done")
                    cont_btn = page.locator('.a-popover-footer .a-button-primary button, #GLUXConfirmClose')
                    if await cont_btn.count() > 0:
                        await cont_btn.first.click()
                        await asyncio.sleep(1)

            # Reload to apply zip
            await page.reload(wait_until='domcontentloaded')
            await asyncio.sleep(2)

            # Verify
            glow = page.locator('#glow-ingress-line2')
            if await glow.count() > 0:
                glow_text = await glow.text_content()
                if '90402' in glow_text or 'Santa Monica' in glow_text:
                    print("    US zip 90402 set successfully")
                    return True
                else:
                    print(f"    Zip verification unclear: {glow_text}")
            return True
    except Exception as e:
        print(f"    Could not set zip code: {str(e)[:80]}")
    return False


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_amazon_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single Amazon PDP page. Returns a LIST of result dicts (one per color)."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=25000)
        status = resp.status if resp else 0

        if status == 503 or status == 429:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Check for CAPTCHA page
        page_content = ''
        try:
            page_content = await page.content()
        except Exception:
            pass

        if 'captcha' in page_content.lower() or 'robot' in page_content.lower()[:2000]:
            await stats.inc_blocked()
            await throttle.on_block()
            wait_time = 60 * (retries + 1) + random.uniform(15, 30)
            print(f"    CAPTCHA on {url[:60]}... backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
            if retries < MAX_RETRIES:
                await asyncio.sleep(wait_time)
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            return [{'url': url, 'error': 'CAPTCHA', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Check for "Continue shopping" block page (only if no product title)
        has_product = '#productTitle' in page_content
        if not has_product and 'continue shopping' in page_content.lower():
            # Try clicking the continue link
            try:
                cont_link = page.locator('a:has-text("Continue shopping")')
                if await cont_link.count() > 0:
                    await cont_link.first.click()
                    await asyncio.sleep(3)
                    page_content = await page.content()
                    has_product = '#productTitle' in page_content
            except Exception:
                pass

            if not has_product:
                if retries < MAX_RETRIES:
                    await asyncio.sleep(random.uniform(5, 15))
                    return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
                await stats.inc_fail()
                return [{'url': url, 'error': 'Block page - no product', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Wait for product title first
        try:
            await page.wait_for_selector('#productTitle', timeout=8000)
        except Exception:
            pass

        # Wait for color variation widget to load (critical for multi-color extraction)
        # Inline twister is the primary layout on current Amazon (2024+)
        color_selectors = [
            '#tp-inline-twister-dim-values-container',
            '#variation_color_name',
            '#twister',
        ]
        for cs in color_selectors:
            try:
                await page.wait_for_selector(cs, timeout=5000)
                break
            except Exception:
                continue
        # Extra wait for inline twister images to load
        try:
            await page.wait_for_selector('#tp-inline-twister-dim-values-container ul li img', timeout=3000)
        except Exception:
            pass

        # Human-like behavior: scroll a bit, wait for dynamic content
        await asyncio.sleep(random.uniform(2, 4))
        try:
            await page.evaluate('window.scrollBy(0, Math.floor(Math.random() * 400 + 200))')
            await asyncio.sleep(random.uniform(1, 2))
        except Exception:
            pass

        # Phase 1: Extract shared product data (title, brand, rating, etc.) + color list
        try:
            raw_data = await page.evaluate(EXTRACT_JS)
        except Exception as eval_err:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'Evaluate error: {str(eval_err)[:150]}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Verify we got useful data
        if not raw_data.get('title'):
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': 'No product title found', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Phase 2: Click each color swatch to get per-color pricing
        # The initial extraction captures the default color's price.
        # We'll store {color_name: {current_price, original_price}} for each color.
        colors = raw_data.get('colors', [])
        per_color_prices = {}

        if colors and len(colors) > 1:
            print(f"      Clicking {len(colors)} color swatches for per-color pricing...")
            # Record the default/selected color's price
            selected = raw_data.get('selected_color', '')
            if selected:
                per_color_prices[selected] = {
                    'current_price': raw_data.get('current_price', ''),
                    'original_price': raw_data.get('original_price', ''),
                }

            for color_name in colors:
                # Skip if we already have this color's price (it was selected by default)
                if color_name in per_color_prices:
                    continue

                try:
                    # Find and click the swatch for this color
                    # Use img alt text to locate the right swatch, then click its parent li/button
                    clicked = await page.evaluate("""
                        (targetColor) => {
                            // Try inline twister first
                            const container = document.querySelector('#tp-inline-twister-dim-values-container') ||
                                              document.getElementById('variation_color_name') ||
                                              document.getElementById('twister');
                            if (!container) return false;

                            const imgs = container.querySelectorAll('li img, button img');
                            for (const img of imgs) {
                                if (img.alt && img.alt.trim() === targetColor) {
                                    // Click the closest clickable parent (li or button)
                                    const clickTarget = img.closest('li') || img.closest('button') || img;
                                    clickTarget.click();
                                    return true;
                                }
                            }
                            return false;
                        }
                    """, color_name)

                    if clicked:
                        # Wait for price to update after color click
                        await asyncio.sleep(random.uniform(1.0, 2.0))

                        # Extract just the price for this color
                        price_data = await page.evaluate(EXTRACT_PRICE_JS)
                        per_color_prices[color_name] = {
                            'current_price': price_data.get('current_price', ''),
                            'original_price': price_data.get('original_price', ''),
                        }
                    else:
                        # Couldn't click — use default price
                        per_color_prices[color_name] = {
                            'current_price': raw_data.get('current_price', ''),
                            'original_price': raw_data.get('original_price', ''),
                        }
                except Exception as click_err:
                    # On error, use default price for this color
                    per_color_prices[color_name] = {
                        'current_price': raw_data.get('current_price', ''),
                        'original_price': raw_data.get('original_price', ''),
                    }

        # Attach per-color prices to raw_data for the parser
        raw_data['per_color_prices'] = per_color_prices

        color_rows = parse_amazon_pdp(raw_data, url)
        for row in color_rows:
            row['retries'] = retries
            row['url'] = url

        await stats.inc_ok(len(color_rows))
        await throttle.on_ok()
        return color_rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_amazon_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    """Worker coroutine that processes URLs from the queue."""
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  W{wid}: 20-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url[:80]}...")

        await throttle.wait()
        color_rows = await scrape_amazon_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(color_rows)
            progress['processed'].add(url)

        colors_found = len(color_rows)
        has_error = any(r.get('error') for r in color_rows)
        if has_error:
            print(f"    W{wid}: Error — {color_rows[0].get('error', '')[:80]}")
        else:
            print(f"    W{wid}: {colors_found} color rows extracted")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    """Run a batch of URLs with a fresh browser and contexts."""
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   {len(results)} existing rows, {len(remaining)} URLs remaining")

    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-gpu",
        "--window-size=1440,900",
    ])

    async def block_resources(route):
        """Block images, fonts, CSS for speed (Amazon doesn't need them like Walmart)."""
        await route.abort()

    all_pages = []
    all_contexts = []
    print(f"Launching {NUM_CONTEXTS} contexts x {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            # Block images, fonts, SVG for speed
            await pg.route('**/*.{png,jpg,jpeg,gif,svg,ico,webp,woff,woff2,ttf}', block_resources)
            all_pages.append(pg)

    # Set US zip on first page to establish session
    print("Setting US zip code (90402) for pricing...")
    try:
        await all_pages[0].goto('https://www.amazon.com', wait_until='domcontentloaded', timeout=15000)
        await asyncio.sleep(2)
        await set_us_zip(all_pages[0])
    except Exception as e:
        print(f"    Zip setup skipped: {str(e)[:80]}")

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\nBrowser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(sdir, "amazon.csv")

    if not os.path.exists(csv_path):
        print(f"CSV not found at {csv_path}")
        return

    # Load URLs from CSV (first column is 'a-link-normal href')
    urls = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)  # skip header
            for row in reader:
                if row:
                    url = row[0].strip()
                    if url.startswith('http'):
                        # Clean URL — strip ref/tracking params for dedup
                        clean = url.split('?')[0] if '?' in url else url
                        # But keep original URL for scraping
                        urls.append(url)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return

    print(f"Loaded {len(urls)} Amazon URLs")
    print(f"  Amazon Essentials: {sum(1 for u in urls if 'Amazon-Essentials' in u)}")
    print(f"  The Drop: {sum(1 for u in urls if 'Drop' in u)}")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 10

    crash_count = 0
    try:
        while True:
            remaining = [u for u in urls if u not in progress['processed']]
            if not remaining:
                print("All URLs processed!")
                break
            print(f"\n{'=' * 60}")
            print(f"Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
            print(f"{'=' * 60}")

            was_crash = False
            try:
                pw = await async_playwright().start()
                finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
                try:
                    await pw.stop()
                except Exception:
                    pass
                if finished:
                    break
            except Exception as e:
                print(f"\nPlaywright/browser error: {str(e)[:150]}")
                was_crash = True
                try:
                    await pw.stop()
                except Exception:
                    pass

            throttle.reset()
            save_progress(progress, sdir)
            save_to_excel(results, sdir)

            if was_crash:
                crash_count += 1
                if crash_count >= max_crashes:
                    print(f"Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                    break
                wait = min(30, 10 * crash_count)
                print(f"Crash restart in {wait}s... (crash #{crash_count})")
                await asyncio.sleep(wait)
            else:
                crash_count = 0
                print(f"Fresh restart in 5s...")
                await asyncio.sleep(5)

    except KeyboardInterrupt:
        print(f"\n\n{'=' * 60}")
        print(f"Ctrl+C detected — saving progress before exit...")
        print(f"{'=' * 60}")

    # Always save on exit (normal completion, crash, or Ctrl+C)
    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"Saved {len(results)} rows to: {out}")
    print(f"{len(progress['processed'])} URLs processed — run again to resume remaining")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
