#!/usr/bin/env python3
"""
Data preparation script for Target jeans dashboard.
Reads Excel, cleans all fields per spec, outputs data.json
"""

import openpyxl
import json
import re
from collections import defaultdict

def parse_price(price_str):
    """Parse '$31.99' -> 31.99 or None"""
    if not price_str or price_str == '':
        return None
    if isinstance(price_str, str):
        # Extract numeric value, handle % signs
        match = re.search(r'\d+\.?\d*', str(price_str).replace('$', '').replace('%', ''))
        if match:
            return float(match.group())
    elif isinstance(price_str, (int, float)):
        return float(price_str)
    return None

def parse_fit(fit_str):
    """Parse 'Straight Leg with a Regular Fit' -> leg_shape, fit_style"""
    if not fit_str:
        return None, None

    fit_str = str(fit_str).lower().strip()

    # Parse leg shape
    leg_shape = None
    if 'straight' in fit_str:
        leg_shape = 'Straight'
    elif 'skinny' in fit_str:
        leg_shape = 'Skinny'
    elif 'jegging' in fit_str:
        leg_shape = 'Jegging'
    elif 'tapered' in fit_str:
        leg_shape = 'Tapered'
    elif 'slim' in fit_str:
        leg_shape = 'Slim'
    elif 'wide' in fit_str:
        leg_shape = 'Wide'
    elif 'flare' in fit_str or 'bell' in fit_str:
        leg_shape = 'Flare'
    elif 'bootcut' in fit_str or 'boot cut' in fit_str:
        leg_shape = 'Bootcut'
    elif 'barrel' in fit_str:
        leg_shape = 'Barrel'
    elif 'baggy' in fit_str:
        leg_shape = 'Baggy'
    elif 'crop' in fit_str:
        leg_shape = 'Crop'
    elif 'capri' in fit_str:
        leg_shape = 'Capri'
    elif 'ankle' in fit_str:
        leg_shape = 'Ankle'
    elif 'mom' in fit_str:
        leg_shape = 'Mom'
    elif 'boyfriend' in fit_str or 'girlfriend' in fit_str:
        leg_shape = 'Boyfriend'
    elif 'relaxed' in fit_str:
        leg_shape = 'Relaxed'
    else:
        leg_shape = 'Other'

    # Parse fit style
    fit_style = None
    if 'regular' in fit_str or 'classic' in fit_str or 'standard' in fit_str:
        fit_style = 'Regular'
    elif 'contemporary' in fit_str:
        fit_style = 'Contemporary/Slim'
    elif 'slim' in fit_str and 'fit' in fit_str:
        fit_style = 'Contemporary/Slim'
    elif 'casual' in fit_str or 'relaxed' in fit_str or 'easy' in fit_str:
        fit_style = 'Casual/Relaxed'
    elif 'curvy' in fit_str:
        fit_style = 'Curvy'
    elif 'loose' in fit_str or 'boyfriend' in fit_str or 'girlfriend' in fit_str:
        fit_style = 'Loose'
    elif 'straight' in fit_str and 'fit' in fit_str:
        fit_style = 'Straight'
    elif 'stretch' in fit_str:
        fit_style = 'Stretch'
    elif 'tailored' in fit_str:
        fit_style = 'Contemporary/Slim'
    else:
        fit_style = 'Other'

    return leg_shape, fit_style

def standardize_rise(rise_str):
    """Standardize rise names"""
    if not rise_str:
        return None

    rise_str = str(rise_str).lower().strip()

    if 'ultra' in rise_str or 'super high' in rise_str or 'extremely high' in rise_str:
        return 'Ultra-High Rise'
    elif 'high' in rise_str:
        return 'High Rise'
    elif 'mid' in rise_str or 'classic' in rise_str or 'regular' in rise_str:
        return 'Mid Rise'
    elif 'low' in rise_str:
        return 'Low Rise'
    else:
        return 'Other'

def color_to_wash_category(color_str):
    """Map 500+ color values to wash categories"""
    if not color_str:
        return None  # will be excluded from charts (no data)

    color_str = str(color_str).lower().strip()
    if not color_str or color_str == 'none':
        return None

    # Black
    if 'black' in color_str:
        return 'Black'

    # White/Cream
    if any(x in color_str for x in ['white', 'cream', 'ivory', 'off-white']):
        return 'White/Cream'

    # Grey
    if any(x in color_str for x in ['grey', 'gray', 'charcoal', 'gunmetal', 'steel']):
        return 'Grey'

    # Dark Wash
    if any(x in color_str for x in ['dark', 'indigo', 'rinse', 'deep', 'midnight',
                                      'navy', 'heritage']):
        return 'Dark Wash'

    # Light Wash
    if any(x in color_str for x in ['light', 'bleach', 'faded', 'acid']):
        return 'Light Wash'

    # Medium Wash (mid, stonewash, plain "blue", "denim", "wash")
    if any(x in color_str for x in ['medium', 'stonewash', 'mid wash', 'mid-wash']):
        return 'Medium Wash'

    # Plain "blue" or "denim" without other qualifiers = Medium Wash
    if color_str in ('blue', 'denim', 'denim blue', 'blue denim'):
        return 'Medium Wash'

    # Brown/Tan tones
    if any(x in color_str for x in ['brown', 'chocolate', 'toffee', 'bourbon',
                                      'tan', 'camel', 'cognac', 'coffee']):
        return 'Brown/Tan'

    # Khaki/Olive/Earth
    if any(x in color_str for x in ['khaki', 'olive', 'pine', 'sage', 'moss',
                                      'army', 'camo', 'sand', 'stone', 'natural']):
        return 'Earth Tones'

    # Everything else is truly a non-denim color
    return 'Color'

def parse_inseam(inseam_str):
    """Parse '30 Inches' -> 30 or None"""
    if not inseam_str or 'no' in str(inseam_str).lower():
        return None

    match = re.search(r'(\d+)', str(inseam_str))
    if match:
        return int(match.group(1))
    return None

def parse_cotton_percent(cotton_pct, material_str, fabric_parsed=None):
    """Parse % Cotton, fallback to Material field.
    If fabric contains cotton but we only have 0%, treat as unknown (None)."""
    pct = None

    if cotton_pct:
        match = re.search(r'(\d+)', str(cotton_pct))
        if match:
            pct = int(match.group(1))

    # Fallback to Material for specific percentages
    if pct is None and material_str:
        mat_lower = str(material_str).lower()
        if '100% cotton' in mat_lower or mat_lower.strip() == 'cotton':
            pct = 100
        else:
            match = re.search(r'(\d+)%\s*cotton', mat_lower)
            if match:
                pct = int(match.group(1))

    # If cotton % is 0 but fabric clearly contains cotton, treat as unknown
    if pct == 0:
        fab = str(fabric_parsed or '').lower()
        mat = str(material_str or '').lower()
        if 'cotton' in fab or 'cotton' in mat:
            return None  # we know it has cotton, just don't know how much

    return pct

def standardize_garment_length(gl_str):
    """Group garment lengths into fewer buckets."""
    if not gl_str:
        return None
    gl = str(gl_str).lower().strip()
    if gl in ('none', ''):
        return None
    if gl == 'full':
        return 'Full'
    elif gl == 'ankle':
        return 'Ankle'
    elif gl in ('crop', 'at calf', 'low calf', '7/8'):
        return 'Crop'
    elif gl in ('capri', 'at knee', 'below knee'):
        return 'Capri'
    elif gl == 'short':
        return 'Short'
    else:
        return gl.title()


def standardize_fabric_weight(fw_str):
    """Standardize fabric weight labels."""
    if not fw_str:
        return None
    fw = str(fw_str).lower().strip()
    if 'year round' in fw:
        return 'Midweight'
    elif 'midweight' in fw or 'mid weight' in fw or 'medium' in fw:
        return 'Midweight'
    elif 'extra light' in fw:
        return 'Extra Lightweight'
    elif 'lightweight' in fw or 'light weight' in fw:
        return 'Lightweight'
    elif 'heavyweight' in fw or 'heavy weight' in fw:
        return 'Heavyweight'
    else:
        return fw_str  # keep original if not recognized


def load_and_clean_data():
    """Load Excel, clean all fields, return list of cleaned rows"""

    wb = openpyxl.load_workbook('/sessions/eloquent-kind-feynman/mnt/Womens jeans scraper/target_pdp_results.xlsx')
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    col_indices = {h: i for i, h in enumerate(headers)}

    print("Loading " + str(ws.max_row - 1) + " rows from Excel...")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_cells = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(len(headers))]

        # Extract key fields
        brand = row_cells[col_indices.get('Brand', 0)]
        owned_brand_text = row_cells[col_indices.get('Owned Brand', 0)]
        brand_type = row_cells[col_indices.get('Brand Type', 0)]
        color = row_cells[col_indices.get('Color', 0)]
        current_price_str = row_cells[col_indices.get('Current Price', 0)]
        original_price_str = row_cells[col_indices.get('Original Price', 0)]
        price_type = row_cells[col_indices.get('Price Type', 0)]
        cotton_pct = row_cells[col_indices.get('% Cotton', 0)]
        material = row_cells[col_indices.get('Material', 0)]
        fabric_parsed = row_cells[col_indices.get('Fabric Parsed', 0)]
        inseam_str = row_cells[col_indices.get('Inseam Length', 0)]
        fit_str = row_cells[col_indices.get('Fit', 0)]
        rise_str = row_cells[col_indices.get('Rise', 0)]
        garment_length_raw = row_cells[col_indices.get('Garment Length', 0)]
        garment_length = standardize_garment_length(garment_length_raw)
        fabric_weight_raw = row_cells[col_indices.get('Fabric Weight Type', 0)]
        # Standardize fabric weight
        fabric_weight = standardize_fabric_weight(fabric_weight_raw)
        stretch = row_cells[col_indices.get('Stretch', 0)]
        rating = row_cells[col_indices.get('Rating', 0)]
        review_count = row_cells[col_indices.get('Review Count', 0)]
        total_colors = row_cells[col_indices.get('Total Colors', 0)]
        num_sizes = row_cells[col_indices.get('# Sizes', 0)]
        color_sizes = row_cells[col_indices.get('Color Sizes', 0)]

        # Parse prices (numeric)
        current_price = parse_price(current_price_str)
        original_price = parse_price(original_price_str)

        # If no original price (product not on sale), use current price
        if original_price is None and current_price is not None:
            original_price = current_price

        # Parse fit -> leg_shape, fit_style
        leg_shape, fit_style = parse_fit(fit_str)

        # Standardize rise
        rise = standardize_rise(rise_str)

        # Color -> wash category
        wash_category = color_to_wash_category(color)

        # Parse inseam
        inseam = parse_inseam(inseam_str)

        # Parse cotton %
        cotton = parse_cotton_percent(cotton_pct, material, fabric_parsed)

        # Normalize owned brand flag
        is_owned = owned_brand_text == 'Yes' if owned_brand_text else False

        # Ensure numeric fields
        try:
            rating = float(rating) if rating else 0
        except:
            rating = 0
        try:
            review_count = int(review_count) if review_count else 0
        except:
            review_count = 0
        try:
            total_colors = int(total_colors) if total_colors else 0
        except:
            total_colors = 0
        try:
            num_sizes = int(num_sizes) if num_sizes else 0
        except:
            num_sizes = 0

        cleaned_row = {
            'brand': brand,
            'is_owned_brand': is_owned,
            'brand_type': brand_type,
            'color': color,
            'wash_category': wash_category,
            'current_price': current_price,
            'original_price': original_price,
            'price_type': price_type,
            'leg_shape': leg_shape,
            'fit_style': fit_style,
            'rise': rise,
            'garment_length': garment_length,
            'fabric_weight': fabric_weight,
            'stretch': stretch,
            'inseam': inseam,
            'cotton_percent': cotton,
            'rating': rating,
            'review_count': review_count,
            'total_colors': total_colors,
            'num_sizes': num_sizes,
            'color_sizes': color_sizes,
        }

        rows.append(cleaned_row)

    return rows

def compute_aggregates(rows):
    """Compute brand-level aggregates"""

    brand_stats = defaultdict(lambda: {
        'count': 0,
        'is_owned': False,
        'prices': [],
        'original_prices': [],
        'ratings': [],
        'review_counts': [],
        'color_combos': 0,
        'rises': defaultdict(int),
        'leg_shapes': defaultdict(int),
        'garment_lengths': defaultdict(int),
        'fabric_weights': defaultdict(int),
        'wash_categories': defaultdict(int),
        'inseams': [],
        'cotton_percents': [],
    })

    for row in rows:
        brand = row['brand']

        brand_stats[brand]['count'] += 1
        brand_stats[brand]['is_owned'] = row['is_owned_brand']

        if row['current_price'] is not None:
            brand_stats[brand]['prices'].append(row['current_price'])
        if row['original_price'] is not None:
            brand_stats[brand]['original_prices'].append(row['original_price'])
        if row['rating'] > 0:
            brand_stats[brand]['ratings'].append(row['rating'])
        if row['review_count'] > 0:
            brand_stats[brand]['review_counts'].append(row['review_count'])

        brand_stats[brand]['color_combos'] += 1

        if row['rise']:
            brand_stats[brand]['rises'][row['rise']] += 1
        if row['leg_shape']:
            brand_stats[brand]['leg_shapes'][row['leg_shape']] += 1
        if row['garment_length']:
            brand_stats[brand]['garment_lengths'][row['garment_length']] += 1
        if row['fabric_weight']:
            brand_stats[brand]['fabric_weights'][row['fabric_weight']] += 1
        if row['wash_category']:
            brand_stats[brand]['wash_categories'][row['wash_category']] += 1

        if row['inseam'] is not None:
            brand_stats[brand]['inseams'].append(row['inseam'])
        if row['cotton_percent'] is not None:
            brand_stats[brand]['cotton_percents'].append(row['cotton_percent'])

    return brand_stats

def main():
    rows = load_and_clean_data()
    print("Cleaned " + str(len(rows)) + " rows")

    brand_stats = compute_aggregates(rows)
    print("Found " + str(len(brand_stats)) + " unique brands")

    # Count owned vs national
    owned_count = sum(1 for b in brand_stats.values() if b['is_owned'])
    print("  " + str(owned_count) + " Owned Brands")
    print("  " + str(len(brand_stats) - owned_count) + " National Brands")

    # Output data.json
    output = {
        'rows': rows,
        'brand_stats': {
            k: {
                'count': v['count'],
                'is_owned': v['is_owned'],
                'prices': v['prices'],
                'original_prices': v['original_prices'],
                'ratings': v['ratings'],
                'review_counts': v['review_counts'],
                'color_combos': v['color_combos'],
                'rises': dict(v['rises']),
                'leg_shapes': dict(v['leg_shapes']),
                'garment_lengths': dict(v['garment_lengths']),
                'fabric_weights': dict(v['fabric_weights']),
                'wash_categories': dict(v['wash_categories']),
                'inseams': v['inseams'],
                'cotton_percents': v['cotton_percents'],
            }
            for k, v in brand_stats.items()
        }
    }

    with open('/sessions/eloquent-kind-feynman/mnt/Womens jeans scraper/data.json', 'w') as f:
        json.dump(output, f, indent=2)

    print("Wrote data.json")

if __name__ == '__main__':
    main()
