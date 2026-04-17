#!/usr/bin/env python3
"""
Production Levi's PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  levi.csv  (one URL per line in first column 'cell-image-link href')
Output: levis_pdp_results.xlsx + levis_pdp_progress.json

Architecture: 5 browser contexts × 1 tab each = 5 parallel workers
              20-minute auto-restart for fresh sessions
              Adaptive throttle (base=3.0s for Cloudflare), crash recovery, progress resume

Each product page is fetched ONCE. The __NEXT_DATA__ JSON is parsed to produce
one row per color option with per-color pricing, specs, images, and variants.

All Levi's direct = brand_type 'NB' (National Brand).
"""

import asyncio, json, os, random, re, time, html as html_mod
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 5
TABS_PER_CTX = 1
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 5 total workers
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 20 * 60  # 20 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]


# ── Helper Functions: Parse specs ────────────────────────────────────────────

def parse_rise(rise_str):
    """Parse rise from spec: 'High Rise', 'Low Rise', etc."""
    if not rise_str:
        return ""
    rise_lower = str(rise_str).lower()
    if 'high' in rise_lower:
        return 'High Rise'
    elif 'mid' in rise_lower:
        return 'Mid Rise'
    elif 'low' in rise_lower:
        return 'Low Rise'
    elif 'super' in rise_lower or 'ultra' in rise_lower:
        return 'Super/Ultra Rise'
    return rise_str.strip()


def parse_leg_shape(leg_shape_str):
    """Parse leg shape: 'Straight', 'Skinny', 'Flare', 'Wide Leg', etc."""
    if not leg_shape_str:
        return ""
    leg_lower = str(leg_shape_str).lower()
    shapes = ['straight', 'skinny', 'slim', 'flare', 'wide leg', 'bootcut', 'boyfriend', 'mom', 'wedgie', 'pencil']
    for shape in shapes:
        if shape in leg_lower:
            return leg_lower.replace(leg_lower.split()[0], leg_lower.split()[0].capitalize())
    return leg_shape_str.strip()


def parse_fit(fit_str):
    """Parse fit name: '726 High Rise Straight', '311 Shaping Skinny', etc."""
    if not fit_str:
        return ""
    return fit_str.strip()


def parse_material(raw_material):
    """Extract fabric composition from raw material string."""
    if not raw_material:
        return ""
    m = re.search(r'(\d+\s*%\s*\w+(?:\s*,\s*\d+\s*%\s*[\w\s-]+)*)', raw_material, re.IGNORECASE)
    return m.group(1).strip() if m else raw_material[:150]


def is_non_basic(title, color, text):
    """Detect non-basic styling (prints, distressed, patterns, etc.)."""
    kws = [
        'print', 'printed', 'graphic', 'pattern', 'floral', 'stripe', 'striped',
        'plaid', 'camo', 'camouflage', 'tie-dye', 'tie dye', 'leopard', 'animal',
        'paisley', 'tropical', 'abstract', 'geometric', 'polka dot', 'checkered',
        'check', 'tartan', 'houndstooth', 'argyle', 'logo', 'embroidered',
        'embroidery', 'novelty', 'snake', 'zebra', 'cheetah', 'destroy',
        'distressed', 'destructed', 'patchwork', 'colorblock',
    ]
    combined = f"{title} {color} {text}".lower()
    return any(kw in combined for kw in kws)


def calc_pct_cotton(mat):
    """Extract cotton percentage from material string."""
    if not mat:
        return ""
    m = re.search(r'(\d+)\s*%\s*(?:Recycled\s+)?Cotton', mat, re.IGNORECASE)
    return f"{m.group(1)}%" if m else "0%"


def calc_pct_natural_fiber(mat):
    """Sum natural fiber percentages (Cotton, Wool, Silk, Linen, Hemp, Lyocell, Tencel)."""
    if not mat:
        return ""
    ns = re.findall(r'(\d+)\s*%\s*(?:Recycled\s+)?(?:Cotton|Wool|Silk|Linen|Hemp|Lyocell|Tencel)', mat, re.IGNORECASE)
    return f"{sum(int(n) for n in ns)}%" if ns else "0%"


def calc_discount(original, current):
    """Calculate discount percentage between original and current price."""
    if not original or not current:
        return ""
    try:
        o = float(str(original).replace('$', '').replace(',', ''))
        c = float(str(current).replace('$', '').replace(',', ''))
        return f"{((o - c) / o) * 100:.1f}%" if o > c > 0 else ""
    except (ValueError, ZeroDivisionError):
        return ""


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 3.0  # Higher for Cloudflare
        self.current_delay = 3.0
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    🐌 Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    🏎️  Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    🔄 Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Levi's PDP Extraction ───────────────────────────────────────────────────

def extract_from_next_data(next_data_json, url, page_text=''):
    """
    Extract product data from __NEXT_DATA__ JSON structure.
    Returns a LIST of result dicts — one per color.

    Expected structure:
      __NEXT_DATA__.props.pageProps.product = {
        id, name, brand, price: {current, original},
        attributes: {fit, rise, legShape, inseam},
        variants: [{color, colorName, prices, images, sizes}, ...],
        images, reviews
      }
    """
    rows = []

    try:
        data = json.loads(next_data_json)
    except (json.JSONDecodeError, TypeError):
        return [{'url': url, 'error': 'Invalid JSON in __NEXT_DATA__', 'timestamp': datetime.now().isoformat()}]

    try:
        # Navigate to product
        product = data.get('props', {}).get('pageProps', {}).get('product', {})
        if not product:
            return [{'url': url, 'error': 'No product in __NEXT_DATA__', 'timestamp': datetime.now().isoformat()}]

        # ── Parent-level data ──────────────────────────────────────────
        parent = {}
        parent['url'] = url
        parent['product_id'] = product.get('id', '')
        parent['product_name'] = product.get('name', '')
        parent['brand'] = product.get('brand', '') or 'Levi\'s'
        parent['brand_type'] = 'NB'  # All Levi's direct = National Brand
        parent['page_text'] = page_text  # Full page text for keyword searching

        # Pricing (parent level)
        price_data = product.get('price', {})
        if isinstance(price_data, dict):
            parent['current_price'] = price_data.get('current', '') or price_data.get('currentPrice', '')
            parent['original_price'] = price_data.get('original', '') or price_data.get('originalPrice', '')
        else:
            parent['current_price'] = ''
            parent['original_price'] = ''

        # Convert prices to string format if numeric
        if parent['current_price'] and isinstance(parent['current_price'], (int, float)):
            parent['current_price'] = f"${parent['current_price']:.2f}"
        if parent['original_price'] and isinstance(parent['original_price'], (int, float)):
            parent['original_price'] = f"${parent['original_price']:.2f}"

        parent['on_sale'] = bool(parent['original_price'] and parent['current_price'] and
                                 float(str(parent['original_price']).replace('$', '')) >
                                 float(str(parent['current_price']).replace('$', '')))
        parent['discount_pct'] = calc_discount(parent['original_price'], parent['current_price'])

        # Attributes
        attrs = product.get('attributes', {})
        if isinstance(attrs, dict):
            parent['fit'] = attrs.get('fit', '')
            parent['rise'] = parse_rise(attrs.get('rise', ''))
            parent['leg_shape'] = parse_leg_shape(attrs.get('legShape', ''))
            parent['inseam'] = attrs.get('inseam', '')
        else:
            parent['fit'] = ''
            parent['rise'] = ''
            parent['leg_shape'] = ''
            parent['inseam'] = ''

        # Material/Fabric
        material = product.get('material', '') or product.get('fabric', '') or ''
        parent['fabric_raw'] = material
        parent['fabric_parsed'] = parse_material(material)
        parent['pct_cotton'] = calc_pct_cotton(material)
        parent['pct_natural_fiber'] = calc_pct_natural_fiber(material)

        # Images
        images = product.get('images', [])
        parent['image_count'] = len(images) if isinstance(images, list) else 0
        primary_image = images[0] if isinstance(images, list) and images else ''
        if isinstance(primary_image, dict):
            primary_image = primary_image.get('url', '')
        parent['primary_image'] = primary_image

        # Reviews/ratings
        reviews = product.get('reviews', {})
        if isinstance(reviews, dict):
            parent['average_rating'] = reviews.get('averageRating', '') or reviews.get('average', '')
            parent['review_count'] = reviews.get('count', '') or reviews.get('reviewCount', '')
        else:
            parent['average_rating'] = ''
            parent['review_count'] = ''

        # Breadcrumb & metadata
        parent['breadcrumb'] = ' > '.join(product.get('breadcrumbs', [])) if product.get('breadcrumbs') else ''
        parent['meta_description'] = product.get('description', '')

        # Feature bullets
        bullets = product.get('features', []) or product.get('bullets', [])
        parent['feature_bullets'] = ' | '.join(bullets) if isinstance(bullets, list) else str(bullets)[:500]

        # Product details (new field)
        product_details = product.get('productDetails', []) or product.get('details', [])
        parent['product_details'] = ' | '.join(product_details) if isinstance(product_details, list) else str(product_details)[:500]

        # Size and fit section (new field)
        size_and_fit = product.get('sizeAndFit', []) or product.get('sizeAndFitData', [])
        parent['size_and_fit'] = ' | '.join(size_and_fit) if isinstance(size_and_fit, list) else str(size_and_fit)[:500]

        # Length hit (e.g., "hits at ankle") parsed from size_and_fit (new field)
        length_hit = ""
        size_fit_text = parent['size_and_fit'].lower()
        if 'ankle' in size_fit_text:
            length_hit = 'hits at ankle'
        elif 'knee' in size_fit_text:
            length_hit = 'hits at knee'
        elif 'mid-calf' in size_fit_text or 'mid calf' in size_fit_text:
            length_hit = 'hits at mid-calf'
        elif 'calf' in size_fit_text:
            length_hit = 'hits at calf'
        elif 'floor' in size_fit_text:
            length_hit = 'hits at floor'
        parent['length_hit'] = length_hit

        # Variants (colors)
        variants = product.get('variants', [])
        if not isinstance(variants, list):
            variants = []

        parent['total_colors'] = len(variants)

        if not variants:
            # No color variants — single row with parent data
            row = {**parent}
            row['color'] = ''
            row['total_sizes'] = 0
            row['sizes_list'] = ''
            row['non_basic'] = is_non_basic(parent['product_name'], '', parent.get('meta_description', ''))
            row['retries'] = 0
            row['timestamp'] = datetime.now().isoformat()
            rows.append(row)
        else:
            # One row per color
            for variant in variants:
                if not isinstance(variant, dict):
                    continue

                row = {**parent}

                # Color info
                color_name = variant.get('color', '') or variant.get('colorName', '')
                row['color'] = color_name

                # Per-color pricing
                color_price = variant.get('price', {})
                if isinstance(color_price, dict):
                    row['current_price'] = color_price.get('current', '') or color_price.get('currentPrice', '')
                    row['original_price'] = color_price.get('original', '') or color_price.get('originalPrice', '')
                else:
                    row['current_price'] = parent.get('current_price', '')
                    row['original_price'] = parent.get('original_price', '')

                # Convert to string if numeric
                if row['current_price'] and isinstance(row['current_price'], (int, float)):
                    row['current_price'] = f"${row['current_price']:.2f}"
                if row['original_price'] and isinstance(row['original_price'], (int, float)):
                    row['original_price'] = f"${row['original_price']:.2f}"

                row['on_sale'] = bool(row['original_price'] and row['current_price'] and
                                     float(str(row['original_price']).replace('$', '')) >
                                     float(str(row['current_price']).replace('$', '')))
                row['discount_pct'] = calc_discount(row['original_price'], row['current_price'])

                # Sizes available for this color
                sizes = variant.get('sizes', []) or variant.get('availableSizes', [])
                if isinstance(sizes, list):
                    size_values = []
                    for size in sizes:
                        if isinstance(size, dict):
                            size_values.append(size.get('size', ''))
                        else:
                            size_values.append(str(size))
                    row['sizes_list'] = ', '.join([s for s in size_values if s])
                    row['total_sizes'] = len([s for s in size_values if s])
                else:
                    row['sizes_list'] = ''
                    row['total_sizes'] = 0

                # Color-specific image
                color_images = variant.get('images', [])
                if isinstance(color_images, list) and color_images:
                    img = color_images[0]
                    if isinstance(img, dict):
                        row['primary_image'] = img.get('url', row.get('primary_image', ''))
                    else:
                        row['primary_image'] = str(img)

                row['non_basic'] = is_non_basic(row['product_name'], color_name, parent.get('meta_description', ''))
                row['retries'] = 0
                row['timestamp'] = datetime.now().isoformat()
                rows.append(row)

        return rows

    except Exception as e:
        return [{'url': url, 'error': f'Parse error: {str(e)[:200]}', 'timestamp': datetime.now().isoformat()}]


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Setup a browser context for Levi's — works with Firefox or Chromium."""
    ctx = await browser.new_context(
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    try:
        await ctx.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
    except Exception:
        pass
    return ctx


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_levis_pdp(page, url, stats, throttle, retries=0):
    """
    Scrape a single Levi's PDP page.

    Strategy:
      1. Navigate with headless=False to bypass Cloudflare
      2. Wait for JS to load __NEXT_DATA__
      3. Try to extract from __NEXT_DATA__ (primary)
      4. Fallback to DOM scraping if JSON fails
      5. Return LIST of result dicts (one per color)
    """
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=25000)
        status = resp.status if resp else 0

        # ── Hard block detection — bail immediately on CDN blocks ──
        if status == 403:
            try:
                body_text = await page.text_content('body')
                body_lower = (body_text or '').lower()[:1000]
            except Exception:
                body_lower = ''
            if 'access denied' in body_lower or 'reference #' in body_lower or 'cloudflare' in body_lower:
                print(f"    HARD BLOCK — Access Denied (HTTP 403). Skipping immediately.")
                await stats.inc_blocked()
                return [{'url': url, 'product_name': 'Access Denied', 'error': 'CDN hard block (403)', 'retries': 0, 'timestamp': datetime.now().isoformat()}]
            # Non-CDN 403 — limited retry
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < 1:
                await asyncio.sleep(random.uniform(5, 10))
                return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP 403', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status == 429 or status == 503:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Human-like scroll
        await asyncio.sleep(random.uniform(2, 4))
        await page.evaluate("window.scrollBy(0, window.innerHeight * 0.3)")
        await asyncio.sleep(random.uniform(1, 2))

        # Extract page text for keyword searching (first 10000 chars)
        page_text = ''
        try:
            page_text = await page.evaluate('document.body.innerText.substring(0, 10000)')
        except Exception:
            pass

        # Extract __NEXT_DATA__ from HTML
        html_content = await page.content()
        next_data_match = re.search(r'<script id="__NEXT_DATA__"[^>]*type="application/json"[^>]*>(.*?)</script>', html_content, re.DOTALL)

        if next_data_match:
            next_data_json = next_data_match.group(1)
            color_rows = extract_from_next_data(next_data_json, url, page_text)
        else:
            # Fallback: try DOM extraction
            color_rows = await extract_from_dom(page, url)
            if not color_rows:
                if retries < MAX_RETRIES:
                    await asyncio.sleep(random.uniform(3, 8))
                    return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
                await stats.inc_fail()
                return [{'url': url, 'error': 'No __NEXT_DATA__ or product data found', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        for row in color_rows:
            row['retries'] = retries
            row['url'] = url

        await stats.inc_ok(len(color_rows))
        await throttle.on_ok()
        return color_rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_levis_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def extract_from_dom(page, url):
    """
    Fallback: extract product data from rendered DOM via page.evaluate().
    """
    try:
        data = await page.evaluate('''() => {
            try {
                const productName = document.querySelector('h1, [data-testid*="title"], .product-name')?.textContent?.trim() || '';
                const priceEl = document.querySelector('[data-testid*="price"], .price, [class*="Price"]');
                const price = priceEl?.textContent?.trim() || '';
                const descEl = document.querySelector('[data-testid*="description"], .description, [class*="description"]');
                const description = descEl?.textContent?.trim() || '';

                const colorElements = document.querySelectorAll('[data-testid*="color"], .color-option, [class*="ColorOption"]');
                const colors = Array.from(colorElements).map(el => el.textContent?.trim()).filter(c => c);

                const sizeElements = document.querySelectorAll('[data-testid*="size"], .size-option, [class*="SizeOption"]');
                const sizes = Array.from(sizeElements).map(el => el.textContent?.trim()).filter(s => s);

                const imageElements = document.querySelectorAll('img[alt*="product"], [class*="ProductImage"] img');
                const imageCount = imageElements.length;

                const pageText = document.body.innerText.substring(0, 10000) || '';

                return {
                    productName,
                    price,
                    description,
                    colors: [...new Set(colors)],
                    sizes: [...new Set(sizes)],
                    imageCount,
                    pageText,
                    hasMaterialInfo: document.body.textContent.includes('%') && document.body.textContent.includes('cotton'),
                };
            } catch (e) {
                return null;
            }
        }''')

        if not data or not data.get('productName'):
            return []

        rows = []
        colors = data.get('colors', [])
        if not colors:
            # Single row with product data
            row = {
                'url': url,
                'product_name': data.get('productName', ''),
                'brand': 'Levi\'s',
                'brand_type': 'NB',
                'color': '',
                'current_price': data.get('price', ''),
                'original_price': '',
                'on_sale': False,
                'discount_pct': '',
                'total_colors': 0,
                'total_sizes': len(data.get('sizes', [])),
                'sizes_list': ', '.join(data.get('sizes', [])),
                'image_count': data.get('imageCount', 0),
                'page_text': data.get('pageText', ''),
                'product_details': '',
                'size_and_fit': '',
                'length_hit': '',
                'inseam': '',
                'retries': 0,
                'timestamp': datetime.now().isoformat(),
            }
            rows.append(row)
        else:
            # One row per color
            for color in colors:
                row = {
                    'url': url,
                    'product_name': data.get('productName', ''),
                    'brand': 'Levi\'s',
                    'brand_type': 'NB',
                    'color': color,
                    'current_price': data.get('price', ''),
                    'original_price': '',
                    'on_sale': False,
                    'discount_pct': '',
                    'total_colors': len(colors),
                    'total_sizes': len(data.get('sizes', [])),
                    'sizes_list': ', '.join(data.get('sizes', [])),
                    'image_count': data.get('imageCount', 0),
                    'page_text': data.get('pageText', ''),
                    'product_details': '',
                    'size_and_fit': '',
                    'length_hit': '',
                    'inseam': '',
                    'retries': 0,
                    'timestamp': datetime.now().isoformat(),
                }
                rows.append(row)

        return rows

    except Exception as e:
        return []


# ── Worker ──────────────────────────────────────────────────────────────────

async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  ⏰ W{wid}: 20-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url[:80]}...")

        await throttle.wait()
        color_rows = await scrape_levis_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(color_rows)
            progress['processed'].add(url)

        colors_found = len(color_rows)
        has_error = any(r.get('error') for r in color_rows)
        if has_error:
            print(f"    ❌ W{wid}: Error — {color_rows[0].get('error', '')[:80]}")
        else:
            print(f"    ✅ W{wid}: {colors_found} color rows extracted")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  💾 Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


# ── Batch Run ───────────────────────────────────────────────────────────────

async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   📊 {len(results)} existing rows, {len(remaining)} URLs remaining")

    # Headed mode for Cloudflare — use Chromium with stealth flags
    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
    ])

    async def block_resources(route):
        """Block images only — keep JS/CSS for Cloudflare."""
        if route.request.resource_type in ('image', 'media'):
            await route.abort()
        else:
            await route.continue_()

    all_pages = []
    all_contexts = []
    print(f"🌐 Launching {NUM_CONTEXTS} contexts × {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            await pg.route('**/*', block_resources)
            all_pages.append(pg)

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\n⚠️  Browser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  💾 Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


# ── Excel Output ────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'URL', 'Product ID', 'Product Name', 'Brand', 'Brand Type',
    'Color', 'Current Price', 'Original Price', 'On Sale', 'Discount %',
    'Rise', 'Leg Shape', 'Fit', 'Inseam',
    'Fabric (Raw)', 'Fabric (Parsed)', '% Cotton', '% Natural Fiber',
    'Total Colors', 'Total Sizes', 'Sizes List',
    'Primary Image', 'Image Count',
    'Average Rating', 'Review Count',
    'Breadcrumb', 'Description', 'Feature Bullets',
    'Page Text', 'Product Details', 'Size and Fit', 'Length Hit',
    'Non-Basic', 'Retries', 'Timestamp', 'Error',
]

EXCEL_FIELDS = [
    'url', 'product_id', 'product_name', 'brand', 'brand_type',
    'color', 'current_price', 'original_price', 'on_sale', 'discount_pct',
    'rise', 'leg_shape', 'fit', 'inseam',
    'fabric_raw', 'fabric_parsed', 'pct_cotton', 'pct_natural_fiber',
    'total_colors', 'total_sizes', 'sizes_list',
    'primary_image', 'image_count',
    'average_rating', 'review_count',
    'breadcrumb', 'meta_description', 'feature_bullets',
    'page_text', 'product_details', 'size_and_fit', 'length_hit',
    'non_basic', 'retries', 'timestamp', 'error',
]


def load_progress(d):
    p = os.path.join(d, "levis_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "levis_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            field_map = dict(zip(EXCEL_HEADERS, EXCEL_FIELDS))
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i] in field_map:
                        rec[field_map[headers[i]]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   📂 Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   ⚠️  Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    with open(os.path.join(d, "levis_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    out = os.path.join(d, "levis_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Levi\'s PDP Results'

    # Dark header fill (1a1a1a)
    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_FIELDS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    ws.freeze_panes = 'A2'
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    wb.save(out)
    return out


# ── Main ────────────────────────────────────────────────────────────────────

async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    uf = os.path.join(sdir, "levi.csv")
    if not os.path.exists(uf):
        print(f"❌ levi.csv not found at {sdir}")
        print(f"   Create levi.csv with Levi's URLs in first column 'cell-image-link href'")
        return

    # Read CSV — URLs in first column
    urls = []
    try:
        import csv
        with open(uf, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)  # Skip header
            for row in reader:
                if row and row[0].strip().startswith('http'):
                    urls.append(row[0].strip())
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return

    print(f"🎯 Loaded {len(urls)} Levi's URLs")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 20

    crash_count = 0
    while True:
        remaining = [u for u in urls if u not in progress['processed']]
        if not remaining:
            print("✅ All URLs processed!")
            break
        print(f"\n{'=' * 60}")
        print(f"▶️  Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
        print(f"{'=' * 60}")

        was_crash = False
        try:
            pw = await async_playwright().start()
            finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
            try:
                await pw.stop()
            except Exception:
                pass
            if finished:
                break
        except Exception as e:
            print(f"\n⚠️  Playwright/browser error: {str(e)[:150]}")
            was_crash = True
            try:
                await pw.stop()
            except Exception:
                pass

        throttle.reset()
        save_progress(progress, sdir)
        save_to_excel(results, sdir)

        if was_crash:
            crash_count += 1
            if crash_count >= max_crashes:
                print(f"❌ Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                break
            wait = min(30, 10 * crash_count)
            print(f"🔄 Crash restart in {wait}s... (crash #{crash_count})")
            await asyncio.sleep(wait)
        else:
            crash_count = 0
            print(f"🔄 Fresh restart in 5s...")
            await asyncio.sleep(5)

    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"✅ Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"   Results: {out}")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
