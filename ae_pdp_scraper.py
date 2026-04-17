#!/usr/bin/env python3
"""
American Eagle PDP Scraper — 1 Row per URL (each URL = 1 color)
================================================================
Input:  ae.csv  (URLs in column 'x-link-to href')
Output: ae_pdp_results.xlsx + ae_pdp_progress.json

Architecture: 5 browser contexts × 2 tabs each = 10 parallel workers
              30-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

AE encodes each color as a separate URL/SKU, so 1 URL = 1 CC.
Data comes from JSON-LD (schema.org Product) embedded in the page HTML.
"""

import asyncio, csv, json, os, re, time, random
from datetime import datetime
from html import unescape
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 5
TABS_PER_CTX = 2
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 10 total workers
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 30 * 60  # 30 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# AE is all one brand — all National Brand for competitive analysis
AE_BRAND = "American Eagle"
AE_BRAND_TYPE = "NB"


# ── Parsing helpers ─────────────────────────────────────────────────────────

def strip_html(text):
    """Remove HTML tags and decode entities."""
    if not text:
        return ''
    clean = re.sub(r'<[^>]+>', '', text)
    return unescape(clean).strip()


def parse_rise(name):
    """Extract rise from product name."""
    n = name.lower()
    if 'super high' in n or 'ultra high' in n:
        return 'Super High Rise'
    elif 'high-waisted' in n or 'high-rise' in n or 'high rise' in n:
        return 'High Rise'
    elif 'mid-rise' in n or 'mid rise' in n:
        return 'Mid Rise'
    elif 'low-rise' in n or 'low rise' in n:
        return 'Low Rise'
    return ''


def parse_leg_shape(name):
    """Extract leg shape from product name."""
    n = name.lower()
    if 'ultra wide' in n or 'ultra-wide' in n:
        return 'Ultra Wide Leg'
    if 'wide' in n:
        return 'Wide Leg'
    if 'bootcut' in n or 'boot cut' in n:
        return 'Bootcut'
    if 'flare' in n:
        return 'Flare'
    if 'skinny' in n:
        return 'Skinny'
    if 'jegging' in n:
        return 'Jegging'
    if 'straight' in n:
        return 'Straight'
    if 'barrel' in n:
        return 'Barrel'
    if 'mom' in n:
        return 'Mom'
    if 'tapered' in n:
        return 'Tapered'
    if 'baggy' in n:
        return 'Baggy'
    if 'crop' in n:
        return 'Crop'
    if 'slim' in n:
        return 'Slim'
    return ''


def parse_fit(name):
    """Extract fit style from product name."""
    n = name.lower()
    if 'curvy' in n:
        return 'Curvy'
    if 'relaxed' in n:
        return 'Relaxed'
    if 'loose' in n:
        return 'Loose'
    if 'baggy' in n:
        return 'Baggy'
    if 'slim' in n and 'slim jean' not in n:
        return 'Slim'
    if 'stretch' in n:
        return 'Stretch'
    if 'strigid' in n:
        return 'Rigid'
    if 'regular' in n:
        return 'Regular'
    return ''


def parse_category_from_url(url):
    """Extract category info from AE URL path."""
    parts = url.rstrip('/').split('/')
    category = ''
    subcategory = ''
    for i, part in enumerate(parts):
        if part == 'p' and i + 3 < len(parts):
            category = parts[i + 2] if i + 2 < len(parts) else ''
            subcategory = parts[i + 3] if i + 3 < len(parts) else ''
            break
    return category, subcategory.replace('-', ' ').title()


def parse_inseam_from_description(desc, name):
    """Try to extract inseam from description or name."""
    text = (desc or '') + ' ' + (name or '')
    m = re.search(r'(\d{2}(?:\.\d)?)["\u201d\u2033]?\s*(?:inch|in\.?|inseam)', text, re.IGNORECASE)
    if m:
        return m.group(1) + ' in'
    return ''


def parse_length_hit(size_and_fit_text):
    """Extract length description (e.g., 'hits at ankle') from size & fit text."""
    if not size_and_fit_text:
        return ''
    text = size_and_fit_text.lower()
    # Look for common length descriptors
    if 'hits at ankle' in text:
        return 'hits at ankle'
    elif 'crops at' in text:
        m = re.search(r'crops at\s+(\w+)', text)
        if m:
            return f'crops at {m.group(1)}'
    elif 'hits at' in text:
        m = re.search(r'hits at\s+(\w+)', text)
        if m:
            return f'hits at {m.group(1)}'
    elif 'full length' in text:
        return 'full length'
    elif 'ankle length' in text:
        return 'ankle length'
    elif 'cropped' in text:
        return 'cropped'
    return ''


def parse_inseam_from_size_fit(size_and_fit_text):
    """Extract inseam measurements from size & fit section."""
    if not size_and_fit_text:
        return ''
    # Look for patterns like "32-inch inseam" or "inseam: 32""
    text = size_and_fit_text.lower()
    m = re.search(r'(\d{2}(?:\.\d)?)["\u201d\u2033]?\s*-?\s*(?:inch)?\s*(?:inseam)', text, re.IGNORECASE)
    if m:
        return m.group(1) + ' in'
    m = re.search(r'inseam[:\s]+(\d{2}(?:\.\d)?)["\u201d\u2033]?', text, re.IGNORECASE)
    if m:
        return m.group(1) + ' in'
    return ''


def parse_material(material_str):
    """Parse material string into cotton percentage if possible."""
    if not material_str:
        return '', ''
    clean = unescape(material_str)
    # Handle "Front: 90% Cotton, 10% Recycled Cotton | Back: 88% Cotton..." format
    # Take only the FIRST panel (before |) for cotton calculation
    first_panel = clean.split('|')[0].strip()
    cotton_pcts = re.findall(r'(\d+)%\s*(?:Recycled\s+|Organic\s+)?Cotton', first_panel, re.IGNORECASE)
    total_cotton = sum(int(p) for p in cotton_pcts) if cotton_pcts else None
    if total_cotton is None:
        organic = re.search(r'(\d+)%\s*Organic\s+Cotton', clean, re.IGNORECASE)
        if organic:
            total_cotton = int(organic.group(1))
    return clean, str(total_cotton) + '%' if total_cotton else ''


# ── AE Data Extraction ─────────────────────────────────────────────────────

def parse_ae_pdp(page_html, url):
    """
    Parse an AE PDP page. Returns a single-item list with one result dict.
    Each AE URL is one color, so always 1 row.
    """
    try:
        # Extract page text (first 10000 chars for keyword searching)
        page_text = strip_html(page_html)[:10000]

        # Extract JSON-LD Product data from script tags
        product_data = None
        scripts = re.findall(r'<script[^>]*>(.*?)</script>', page_html, re.DOTALL)
        for tag in scripts:
            t = tag.strip()
            if t.startswith('{') and '"@type"' in t and '"Product"' in t:
                try:
                    product_data = json.loads(t)
                    break
                except json.JSONDecodeError:
                    continue

        if not product_data:
            return [{'url': url, 'error': 'No JSON-LD Product data found', 'timestamp': datetime.now().isoformat()}]

        # Extract product details section (bullet points)
        product_details = ''
        details_match = re.search(r'(?:Product Details?|Details?)[:\s]*</?\w+[^>]*>(?:<[^>]+>)*\s*((?:[•\-]\s*[^<]*(?:<[^>]+>[^<]*)*\n?)+)', page_html, re.IGNORECASE)
        if details_match:
            details_text = details_match.group(1)
            product_details = strip_html(details_text).strip()

        # Extract size & fit section
        size_and_fit = ''
        sizefit_match = re.search(r'(?:Size & Fit|Size and Fit)[:\s]*</?\w+[^>]*>(?:<[^>]+>)*\s*((?:[•\-\w\s\d\-,.:""\'%]+(?:<[^>]+>[^<]*)*\n?)+)', page_html, re.IGNORECASE)
        if sizefit_match:
            sizefit_text = sizefit_match.group(1)
            size_and_fit = strip_html(sizefit_text).strip()

        # Extract fields from JSON-LD
        name = strip_html(product_data.get('name', ''))
        sku = product_data.get('sku', '')
        description = strip_html(product_data.get('description', ''))
        image = product_data.get('image', '')
        if image and image.startswith('//'):
            image = 'https:' + image
        color = product_data.get('color', '')
        material_raw = product_data.get('material', '')
        material_clean, cotton_pct = parse_material(material_raw)

        brand_data = product_data.get('brand', {})
        brand_name = brand_data.get('name', '') if isinstance(brand_data, dict) else str(brand_data)

        offers = product_data.get('offers', {})
        current_price = None
        availability = ''
        if isinstance(offers, dict):
            try:
                current_price = float(offers.get('price', 0))
            except (ValueError, TypeError):
                current_price = None
            availability = offers.get('availability', '')
        elif isinstance(offers, list) and offers:
            try:
                current_price = float(offers[0].get('price', 0))
            except (ValueError, TypeError):
                current_price = None
            availability = offers[0].get('availability', '')

        # Get list price from HTML (data-test-list-price)
        list_price = None
        list_match = re.search(r'data-test-list-price[^>]*>\s*\$?([\d.]+)', page_html)
        if list_match:
            try:
                list_price = float(list_match.group(1))
            except ValueError:
                pass

        # Get sale price from HTML
        sale_match = re.search(r'data-test-sale-price[^>]*>\s*\$?([\d.]+)', page_html)
        sale_price = None
        if sale_match:
            try:
                sale_price = float(sale_match.group(1))
            except ValueError:
                pass

        # Determine pricing
        if sale_price and list_price:
            final_current = sale_price
            final_original = list_price
            on_sale = 1
        elif current_price and list_price and current_price < list_price:
            final_current = current_price
            final_original = list_price
            on_sale = 1
        elif current_price:
            final_current = current_price
            final_original = list_price if list_price else current_price
            on_sale = 1 if list_price and list_price > current_price else 0
        else:
            final_current = list_price
            final_original = list_price
            on_sale = 0

        discount_pct = ''
        if on_sale and final_original and final_current and final_original > final_current:
            discount_pct = f"{((final_original - final_current) / final_original) * 100:.1f}%"

        # Parse attributes from product name
        rise = parse_rise(name)
        leg_shape = parse_leg_shape(name)
        fit = parse_fit(name)
        # Prefer inseam from size & fit section, fallback to description/name
        inseam_from_sizefit = parse_inseam_from_size_fit(size_and_fit)
        inseam = inseam_from_sizefit if inseam_from_sizefit else parse_inseam_from_description(description, name)
        length_hit = parse_length_hit(size_and_fit)
        category, subcategory = parse_category_from_url(url)

        # Count other color variants from image SKUs on page
        color_skus = set(re.findall(r'scene7\.com/is/image/aeo/(\d+_\d+_\d+)', page_html))
        total_colors = len(color_skus) if color_skus else 1

        return [{
            'url': url,
            'product_name': name,
            'brand': AE_BRAND,
            'brand_type': AE_BRAND_TYPE,
            'color': color,
            'current_price': final_current,
            'current_price_formatted': f"${final_current:.2f}" if final_current else '',
            'original_price': final_original,
            'original_price_formatted': f"${final_original:.2f}" if final_original else '',
            'on_sale': on_sale,
            'discount_pct': discount_pct,
            'rise': rise,
            'leg_shape': leg_shape,
            'fit': fit,
            'inseam': inseam,
            'fabric_material': material_clean,
            'cotton_pct': cotton_pct,
            'category': category,
            'subcategory': subcategory,
            'total_colors_on_page': total_colors,
            'sku': sku,
            'description': description,
            'image_url': image,
            'availability': 'InStock' if 'InStock' in str(availability) else 'OutOfStock' if 'OutOfStock' in str(availability) else str(availability),
            'page_text': page_text,
            'product_details': product_details,
            'size_and_fit': size_and_fit,
            'length_hit': length_hit,
            'timestamp': datetime.now().isoformat(),
        }]

    except Exception as e:
        return [{'url': url, 'error': f'Parse error: {str(e)[:200]}', 'timestamp': datetime.now().isoformat()}]


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 1.0
        self.current_delay = 1.0
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    Throttle UP: delay now {self.current_delay:.1f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 1.5))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'url', 'product_name', 'brand', 'brand_type', 'color',
    'current_price', 'current_price_formatted', 'original_price', 'original_price_formatted',
    'on_sale', 'discount_pct',
    'rise', 'leg_shape', 'fit', 'inseam', 'length_hit',
    'fabric_material', 'cotton_pct',
    'category', 'subcategory', 'total_colors_on_page',
    'sku', 'description', 'image_url', 'availability',
    'page_text', 'product_details', 'size_and_fit',
    'retries', 'timestamp', 'error',
]


def load_progress(d):
    """Load progress and results from files."""
    p = os.path.join(d, "ae_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "ae_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        rec[headers[i]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    """Save progress to JSON file."""
    with open(os.path.join(d, "ae_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    """Save results to Excel file with formatting."""
    out = os.path.join(d, "ae_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'AE PDP Results'

    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_HEADERS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    ws.freeze_panes = 'A2'
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Create a browser context with anti-detection measures."""
    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return ctx


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_ae_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single AE PDP page. Returns a LIST with one result dict."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
        status = resp.status if resp else 0

        if status == 403 or status == 429 or status == 503:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        page_url = page.url.lower()
        if 'captcha' in page_url or 'blocked' in page_url or 'denied' in page_url:
            await stats.inc_blocked()
            wait_time = 60 * (retries + 1) + random.uniform(10, 30)
            print(f"    Blocked on {url[:60]}... backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
            if retries < MAX_RETRIES:
                await asyncio.sleep(wait_time)
                return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
            return [{'url': url, 'error': 'Blocked/CAPTCHA', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        await asyncio.sleep(random.uniform(1.0, 2.0))
        html_content = await page.content()

        # Verify we got product data (look for JSON-LD Product)
        if '"Product"' not in html_content:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': 'No Product JSON-LD in page', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        rows = parse_ae_pdp(html_content, url)
        for row in rows:
            row['retries'] = retries
            row['url'] = url

        await stats.inc_ok(len(rows))
        await throttle.on_ok()
        return rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_ae_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    """Worker coroutine that processes URLs from the queue."""
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  W{wid}: 30-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url.split('/')[-2][:50]}...")

        await throttle.wait()
        rows = await scrape_ae_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(rows)
            progress['processed'].add(url)

        has_error = any(r.get('error') for r in rows)
        if has_error:
            print(f"    W{wid}: Error — {rows[0].get('error', '')[:80]}")
        else:
            color = rows[0].get('color', '?') if rows else '?'
            price = rows[0].get('current_price', '?') if rows else '?'
            print(f"    W{wid}: OK — {color}, ${price}")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    """Run a batch of URLs with a fresh browser and contexts."""
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   {len(results)} existing rows, {len(remaining)} URLs remaining")

    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-gpu",
    ])

    async def block_resources(route):
        await route.abort()

    all_pages = []
    all_contexts = []
    print(f"Launching {NUM_CONTEXTS} contexts x {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            # Block images/fonts to speed things up
            await pg.route('**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,ico,webp}', block_resources)
            all_pages.append(pg)

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\nBrowser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(sdir, "ae.csv")

    if not os.path.exists(csv_path):
        print(f"CSV not found at {csv_path}")
        return

    # Load URLs from CSV (second column is 'x-link-to href')
    urls = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for row in reader:
                if len(row) > 1:
                    url = row[1].strip()
                    if url.startswith('http'):
                        urls.append(url)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return

    print(f"Loaded {len(urls)} AE URLs")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 20

    crash_count = 0
    try:
        while True:
            remaining = [u for u in urls if u not in progress['processed']]
            if not remaining:
                print("All URLs processed!")
                break
            print(f"\n{'=' * 60}")
            print(f"Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
            print(f"{'=' * 60}")

            was_crash = False
            try:
                pw = await async_playwright().start()
                finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
                try:
                    await pw.stop()
                except Exception:
                    pass
                if finished:
                    break
            except Exception as e:
                print(f"\nPlaywright/browser error: {str(e)[:150]}")
                was_crash = True
                try:
                    await pw.stop()
                except Exception:
                    pass

            throttle.reset()
            save_progress(progress, sdir)
            save_to_excel(results, sdir)

            if was_crash:
                crash_count += 1
                if crash_count >= max_crashes:
                    print(f"Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                    break
                wait = min(30, 10 * crash_count)
                print(f"Crash restart in {wait}s... (crash #{crash_count})")
                await asyncio.sleep(wait)
            else:
                crash_count = 0
                print(f"Fresh restart in 5s...")
                await asyncio.sleep(5)

    except KeyboardInterrupt:
        print(f"\n\n{'=' * 60}")
        print(f"Ctrl+C detected — saving progress before exit...")
        print(f"{'=' * 60}")

    # Always save on exit (normal completion, crash, or Ctrl+C)
    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"Saved {len(results)} rows to: {out}")
    print(f"{len(progress['processed'])} URLs processed — run again to resume remaining")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
