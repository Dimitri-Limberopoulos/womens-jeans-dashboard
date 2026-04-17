#!/usr/bin/env python3
"""
Production Playwright PDP Scraper for Macy's Women's Jeans
Architecture: 5 contexts × 1 tab worker pool, Akamai bot detection handling
"""

import asyncio
import json
import logging
import os
import random
import re
import sys
import time
from asyncio import Queue
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from playwright.async_api import (
    async_playwright,
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("macys_pdp_scraper")

# ============================================================================
# CONFIGURATION
# ============================================================================

MACYS_OWNED_BRANDS = {
    "style & co",
    "i.n.c. international concepts",
    "inc international concepts",
    "and now this",
    "on 34th",
    "bar iii",
    "charter club",
    "alfani",
    "jm collection",
    "karen scott",
}

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]

NUM_WORKERS = 5
NUM_CONTEXTS = 5
RESTART_INTERVAL_MINUTES = 20
AUTOSAVE_INTERVAL_SECONDS = 120
CRASH_COUNTER_MAX = 20
BASE_DELAY = 3.0
MAX_DELAY = 30.0

EXCEL_HEADERS = [
    "url",
    "product_name",
    "brand",
    "brand_type",
    "color",
    "current_price",
    "current_price_formatted",
    "original_price",
    "original_price_formatted",
    "on_sale",
    "discount_pct",
    "rise",
    "leg_shape",
    "fit",
    "fabric_raw",
    "fabric_parsed",
    "pct_cotton",
    "pct_natural_fiber",
    "non_basic",
    "breadcrumb",
    "total_colors",
    "total_sizes",
    "sizes_list",
    "average_rating",
    "review_count",
    "image_count",
    "feature_bullets",
    "page_text",
    "product_details",
    "size_and_fit",
    "length_hit",
    "inseam",
    "retries",
    "timestamp",
    "error",
]

# ============================================================================
# HELPERS: PARSING
# ============================================================================


def parse_rise(text: str) -> Optional[str]:
    """Extract rise from product details."""
    if not text:
        return None
    text_lower = text.lower()
    if "high" in text_lower:
        return "high"
    if "mid" in text_lower:
        return "mid"
    if "low" in text_lower:
        return "low"
    return None


def parse_leg_shape(text: str) -> Optional[str]:
    """Extract leg shape from product details."""
    if not text:
        return None
    text_lower = text.lower()
    if "skinny" in text_lower:
        return "skinny"
    if "slim" in text_lower:
        return "slim"
    if "straight" in text_lower:
        return "straight"
    if "flare" in text_lower:
        return "flare"
    if "wide" in text_lower:
        return "wide-leg"
    if "crop" in text_lower:
        return "crop"
    if "boot" in text_lower:
        return "bootcut"
    return None


def parse_fit(text: str) -> Optional[str]:
    """Extract fit descriptor from product details."""
    if not text:
        return None
    text_lower = text.lower()
    if "stretch" in text_lower:
        return "stretch"
    if "fitted" in text_lower:
        return "fitted"
    if "slim" in text_lower:
        return "slim"
    if "comfort" in text_lower:
        return "comfort"
    return None


def parse_material(text: str) -> Tuple[Optional[str], Optional[float], Optional[float]]:
    """
    Parse fabric material from text.
    Returns: (fabric_desc, pct_cotton, pct_natural)
    """
    if not text:
        return (None, None, None)

    fabric = text.strip()
    pct_cotton = None
    pct_natural = None

    # Extract cotton %
    cotton_match = re.search(r"(\d+)%\s*cotton", fabric, re.IGNORECASE)
    if cotton_match:
        pct_cotton = float(cotton_match.group(1))

    # Extract polyester, spandex, etc.
    poly_match = re.search(r"(\d+)%\s*polyester", fabric, re.IGNORECASE)
    spandex_match = re.search(r"(\d+)%\s*(?:spandex|elastane)", fabric, re.IGNORECASE)

    # Natural fiber heuristic: cotton + other natural materials
    if pct_cotton:
        pct_natural = pct_cotton  # Conservative estimate

    return (fabric, pct_cotton, pct_natural)


def is_non_basic(title: str, colors: List[str]) -> bool:
    """Detect non-basic jeans (premium, distressed, etc.)."""
    if not title:
        return False

    title_lower = title.lower()
    non_basic_keywords = [
        "distressed",
        "ripped",
        "embroidered",
        "embellished",
        "premium",
        "luxury",
        "designer",
        "high-end",
        "vintage",
        "oversized",
        "low-rise",
    ]

    for keyword in non_basic_keywords:
        if keyword in title_lower:
            return True

    # Check color diversity
    unique_colors = set(c.lower() for c in colors if c)
    if len(unique_colors) > 8:
        return True

    return False


def parse_length_hit(size_and_fit_text: str) -> Optional[str]:
    """Extract length hit (e.g., 'hits at ankle', 'hits at thigh') from size & fit text."""
    if not size_and_fit_text:
        return None

    text_lower = size_and_fit_text.lower()
    length_patterns = [
        "hits at ankle",
        "hits at thigh",
        "hits at knee",
        "hits at calf",
        "hits at hip",
        "hits at mid-thigh",
        "hits at upper thigh",
        "hits at lower calf",
        "hits above ankle",
        "hits below knee",
    ]

    for pattern in length_patterns:
        if pattern in text_lower:
            return pattern

    return None


def parse_inseam(size_and_fit_text: str, title: str = "") -> Optional[str]:
    """Extract inseam measurements from size & fit text or title."""
    if not size_and_fit_text and not title:
        return None

    combined_text = (size_and_fit_text + " " + title).lower()

    # Look for patterns like "inseam: 28", "28\" inseam", "inseam 28 inches"
    inseam_match = re.search(r"inseam[:\s]+(\d+(?:\.\d+)?)\s*(?:inches?|\")?", combined_text)
    if inseam_match:
        return f"{inseam_match.group(1)} inches"

    # Look for measurement in quotes like "28""
    quote_match = re.search(r"(\d+(?:\.\d+)?)\s*[\"″]\s*inseam", combined_text)
    if quote_match:
        return f"{quote_match.group(1)} inches"

    return None


def classify_brand(brand: str) -> str:
    """Classify brand as OB (owned brand) or TP (third-party)."""
    if not brand:
        return "UNKNOWN"
    if brand.lower() in MACYS_OWNED_BRANDS:
        return "OB"
    return "TP"


# ============================================================================
# STATS & PROGRESS
# ============================================================================


class Stats:
    """Thread-safe stats counter."""

    def __init__(self):
        self.lock = asyncio.Lock()
        self.processed = 0
        self.success = 0
        self.errors = 0
        self.crashes = 0
        self.start_time = time.time()

    async def increment_processed(self):
        async with self.lock:
            self.processed += 1

    async def increment_success(self):
        async with self.lock:
            self.success += 1

    async def increment_errors(self):
        async with self.lock:
            self.errors += 1

    async def increment_crashes(self):
        async with self.lock:
            self.crashes += 1

    async def get_stats(self) -> Dict[str, Any]:
        async with self.lock:
            elapsed = time.time() - self.start_time
            rate = self.processed / elapsed if elapsed > 0 else 0
            return {
                "processed": self.processed,
                "success": self.success,
                "errors": self.errors,
                "crashes": self.crashes,
                "elapsed_seconds": int(elapsed),
                "rate_per_second": round(rate, 2),
            }


class AdaptiveThrottle:
    """Adaptive delay with exponential backoff on rate limiting."""

    def __init__(self, base_delay: float = BASE_DELAY, max_delay: float = MAX_DELAY):
        self.base_delay = base_delay
        self.max_delay = max_delay
        self.current_delay = base_delay
        self.blocked_count = 0

    async def wait(self):
        await asyncio.sleep(self.current_delay)

    def on_success(self):
        """Reduce delay slightly on success."""
        self.current_delay = max(self.base_delay, self.current_delay * 0.95)
        self.blocked_count = 0

    def on_rate_limit(self):
        """Double delay on 429/403."""
        self.blocked_count += 1
        self.current_delay = min(self.max_delay, self.current_delay * 2.0)
        logger.warning(
            f"Rate limited. Current delay: {self.current_delay:.1f}s (blocked: {self.blocked_count})"
        )

    def on_crash(self):
        """Reset to base on crash."""
        self.current_delay = self.base_delay
        self.blocked_count = 0


def load_progress(progress_file: str) -> Dict[str, Any]:
    """Load progress from JSON file."""
    if os.path.exists(progress_file):
        try:
            with open(progress_file) as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Failed to load progress: {e}")
    return {"completed": [], "failed": [], "in_progress": []}


def save_progress(progress_file: str, progress: Dict[str, Any]):
    """Save progress to JSON file."""
    try:
        with open(progress_file, "w") as f:
            json.dump(progress, f, indent=2)
    except Exception as e:
        logger.error(f"Failed to save progress: {e}")


def save_to_excel(results: List[Dict], output_file: str):
    """Save results to Excel with dark header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Write headers
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        cell.font = cell.font.copy()
        cell.font = cell.font.copy()
        cell.font = cell.font.copy()

    # Write data
    for row_idx, result in enumerate(results, start=2):
        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            value = result.get(header, "")
            ws.cell(row=row_idx, column=col_idx).value = value

    # Auto-adjust column widths
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    wb.save(output_file)
    logger.info(f"Saved {len(results)} results to {output_file}")


# ============================================================================
# PLAYWRIGHT CONTEXT SETUP
# ============================================================================


async def setup_context(
    browser, user_agent: str, context_id: int
) -> Tuple[BrowserContext, Page]:
    """Create a BrowserContext with anti-detection measures."""
    context = await browser.new_context(
        user_agent=user_agent,
        viewport={"width": 1280, "height": 720},
        locale="en-US",
        timezone_id="America/New_York",
    )

    # Disable webdriver detection
    await context.add_init_script(
        """
        Object.defineProperty(navigator, 'webdriver', {
            get: () => false,
        });
        window.chrome = {runtime: {}};
        """
    )

    # Block images to reduce bandwidth; keep JS/CSS for Akamai
    await context.route("**/*.{png,jpg,jpeg,gif,webp}", lambda route: route.abort())

    page = await context.new_page()
    page.set_default_timeout(30000)
    page.set_default_navigation_timeout(30000)

    logger.info(f"Context {context_id} initialized with UA: {user_agent[:50]}...")

    return context, page


# ============================================================================
# SCRAPING LOGIC
# ============================================================================


async def scrape_pdp(page: Page, url: str, retries: int = 0) -> Optional[Dict[str, Any]]:
    """
    Scrape a single Macy's PDP.
    Returns a list of dicts, one per color.
    """
    try:
        # Aggressive timeout and navigation
        await page.goto(url, wait_until="networkidle", timeout=30000)

        # Check for CAPTCHA/bot detection — only flag if NO product data is present.
        # Words like "robot", "verify", "challenge" appear in normal Macy's page text
        # (e.g., "Verified purchase", meta robots tags), so we check for product data first.
        page_text = await page.evaluate("() => document.documentElement.innerText")
        page_lower = page_text.lower() if page_text else ''
        has_product = (
            'add to bag' in page_lower or
            'product' in page_lower[:500] or
            bool(await page.query_selector('.product-title, h1, [data-testid*="product"]'))
        )
        if not has_product:
            real_block_indicators = ['access denied', 'are you a robot', 'press and hold', 'captcha']
            for indicator in real_block_indicators:
                if indicator in page_lower[:3000]:
                    logger.warning(f"CAPTCHA/block detected on {url}")
                    return {"error": "CAPTCHA_DETECTED", "rows": []}

        # Check for 403/429
        status = page.url
        if "403" in status or "429" in status:
            logger.warning(f"Rate limited on {url}")
            return {"error": "RATE_LIMITED", "rows": []}

        # Human-like scroll
        await page.evaluate(
            """
            async () => {
                const maxScroll = document.body.scrollHeight;
                for (let i = 0; i < maxScroll; i += 300) {
                    window.scrollTo(0, i);
                    await new Promise(r => setTimeout(r, 100));
                }
            }
            """
        )

        # Extract data
        data = await page.evaluate(
            """
            () => {
                // Title
                const titleEl = document.querySelector('.product-title') || document.querySelector('h1');
                const title = titleEl ? titleEl.textContent.trim() : '';

                // Brand
                const brandEl = document.querySelector('.product-brand') || document.querySelector('[data-testid*="brand"]');
                const brand = brandEl ? brandEl.textContent.trim() : '';

                // Breadcrumb
                const breadcrumbs = Array.from(document.querySelectorAll('.breadcrumb a, [aria-label*="breadcrumb"] a')).map(el => el.textContent.trim());
                const breadcrumb = breadcrumbs.join(' > ');

                // Price
                const priceEl = document.querySelector('.lowest-sale-price, .sale-price, [data-testid*="price"]');
                const originalPriceEl = document.querySelector('.was-price, .original-price, [data-testid*="original"]');
                const currentPrice = priceEl ? priceEl.textContent.trim() : '';
                const originalPrice = originalPriceEl ? originalPriceEl.textContent.trim() : '';

                // Colors - click each swatch
                const colorSwatches = Array.from(document.querySelectorAll('[data-testid*="swatch"], .color-swatch, .swatch-button'));
                const colors = colorSwatches.map(s => s.getAttribute('data-color') || s.textContent.trim() || s.getAttribute('aria-label') || s.title).filter(Boolean);

                // Sizes
                const sizeElements = Array.from(document.querySelectorAll('[data-testid*="size"], .size-option, [role="radio"][data-size]'));
                const sizes = sizeElements.map(s => s.textContent.trim() || s.getAttribute('data-size')).filter(Boolean);

                // Rating & reviews
                const ratingEl = document.querySelector('[data-testid*="rating"], .rating, [role*="img"][aria-label*="star"]');
                const rating = ratingEl ? ratingEl.getAttribute('aria-label') || ratingEl.textContent.trim() : '';
                const reviewCountEl = document.querySelector('[data-testid*="review"], .review-count');
                const reviewCount = reviewCountEl ? reviewCountEl.textContent.trim() : '';

                // Material/fabric
                const detailsSections = Array.from(document.querySelectorAll('[data-testid*="details"], .product-details, .accordion-section'));
                let fabric = '';
                for (let section of detailsSections) {
                    const text = section.textContent.toLowerCase();
                    if (text.includes('fabric') || text.includes('material') || text.includes('cotton')) {
                        fabric = section.textContent.trim();
                        break;
                    }
                }

                // Images
                const images = Array.from(document.querySelectorAll('img[src*="macys"], img[alt*="jeans"]')).length || 0;

                // Feature bullets
                const bullets = Array.from(document.querySelectorAll('ul li, .feature-bullet')).map(el => el.textContent.trim()).slice(0, 5);

                // Page text (first 10000 chars)
                const pageText = document.body.innerText.substring(0, 10000);

                // Product details section - extract bullet points
                let productDetails = '';
                for (let section of detailsSections) {
                    const lis = Array.from(section.querySelectorAll('li')).map(li => li.textContent.trim());
                    if (lis.length > 0) {
                        productDetails = lis.join(' | ');
                        break;
                    }
                }

                // Size & fit section
                let sizeAndFit = '';
                for (let section of detailsSections) {
                    const text = section.textContent.toLowerCase();
                    if (text.includes('size') && text.includes('fit')) {
                        sizeAndFit = section.textContent.trim();
                        break;
                    }
                }

                return {
                    title,
                    brand,
                    currentPrice,
                    originalPrice,
                    colors,
                    sizes,
                    rating,
                    reviewCount,
                    fabric,
                    breadcrumb,
                    imageCount: images,
                    featureBullets: bullets.join(' | '),
                    pageText,
                    productDetails,
                    sizeAndFit,
                };
            }
            """
        )

        # Parse extracted data
        title = data.get("title", "")
        brand = data.get("brand", "")
        colors = data.get("colors", [])
        sizes = data.get("sizes", [])
        fabric_raw = data.get("fabric", "")
        fabric_parsed, pct_cotton, pct_natural = parse_material(fabric_raw)

        # ── Phase 2: Per-color pricing via swatch clicking ──
        per_color_prices = {}
        if len(colors) > 1:
            try:
                swatch_els = await page.query_selector_all('[data-testid*="swatch"], .color-swatch, .swatch-button')
                for swatch_el in swatch_els:
                    try:
                        cn = await swatch_el.get_attribute('data-color') or ''
                        if not cn:
                            cn = await swatch_el.get_attribute('aria-label') or ''
                        if not cn:
                            cn = (await swatch_el.text_content() or '').strip()
                        if not cn:
                            cn = await swatch_el.get_attribute('title') or ''
                        cn = cn.strip()
                        if not cn:
                            continue

                        await swatch_el.click()
                        await asyncio.sleep(random.uniform(1.0, 2.0))

                        price_val = await page.evaluate("""() => {
                            var el = document.querySelector('.lowest-sale-price, .sale-price, [data-testid*="price"]');
                            if (el) {
                                var t = el.textContent.trim();
                                var m = t.match(/\\$(\\d+\\.?\\d*)/);
                                if (m) return parseFloat(m[1]);
                            }
                            return null;
                        }""")
                        if price_val and isinstance(price_val, (int, float)):
                            per_color_prices[cn.lower()] = price_val
                    except Exception:
                        continue
                if per_color_prices:
                    logger.info(f"Per-color prices captured for {len(per_color_prices)} colors")
            except Exception as e:
                logger.warning(f"Phase 2 pricing skipped: {str(e)[:80]}")

        # Generate per-color rows
        rows = []
        for color in colors:
            rise = parse_rise(fabric_raw)
            leg_shape = parse_leg_shape(fabric_raw)
            fit = parse_fit(fabric_raw)
            non_basic = is_non_basic(title, colors)
            brand_type = classify_brand(brand)

            # Parse prices - use per-color price if available
            color_lower = color.lower() if color else ''
            if color_lower in per_color_prices:
                current_price = per_color_prices[color_lower]
                current_price_str = f'{current_price:.2f}'
            else:
                current_price_str = data.get("currentPrice", "").replace("$", "").strip()
                try:
                    current_price = float(current_price_str) if current_price_str else None
                except ValueError:
                    current_price = None

            original_price_str = data.get("originalPrice", "").replace("$", "").strip()

            try:
                original_price = float(original_price_str) if original_price_str else None
            except ValueError:
                original_price = None

            on_sale = original_price and current_price and current_price < original_price
            discount_pct = (
                round(((original_price - current_price) / original_price * 100), 1)
                if on_sale
                else None
            )

            # Extract rating
            rating_text = data.get("rating", "")
            rating_match = re.search(r"(\d+\.?\d*)", rating_text)
            average_rating = float(rating_match.group(1)) if rating_match else None

            review_count_text = data.get("reviewCount", "")
            review_match = re.search(r"(\d+)", review_count_text)
            review_count = int(review_match.group(1)) if review_match else None

            # Parse new fields
            size_and_fit_text = data.get("sizeAndFit", "")
            length_hit = parse_length_hit(size_and_fit_text)
            inseam = parse_inseam(size_and_fit_text, title)

            row = {
                "url": url,
                "product_name": title,
                "brand": brand,
                "brand_type": brand_type,
                "color": color,
                "current_price": current_price,
                "current_price_formatted": data.get("currentPrice", ""),
                "original_price": original_price,
                "original_price_formatted": data.get("originalPrice", ""),
                "on_sale": on_sale,
                "discount_pct": discount_pct,
                "rise": rise,
                "leg_shape": leg_shape,
                "fit": fit,
                "fabric_raw": fabric_raw,
                "fabric_parsed": fabric_parsed,
                "pct_cotton": pct_cotton,
                "pct_natural_fiber": pct_natural,
                "non_basic": non_basic,
                "breadcrumb": data.get("breadcrumb", ""),
                "total_colors": len(colors),
                "total_sizes": len(sizes),
                "sizes_list": ",".join(sizes) if sizes else "",
                "average_rating": average_rating,
                "review_count": review_count,
                "image_count": data.get("imageCount", 0),
                "feature_bullets": data.get("featureBullets", ""),
                "page_text": data.get("pageText", ""),
                "product_details": data.get("productDetails", ""),
                "size_and_fit": size_and_fit_text,
                "length_hit": length_hit,
                "inseam": inseam,
                "retries": retries,
                "timestamp": datetime.now().isoformat(),
                "error": None,
            }
            rows.append(row)

        return {"error": None, "rows": rows}

    except PlaywrightTimeoutError:
        logger.error(f"Timeout on {url}")
        return {"error": "TIMEOUT", "rows": []}
    except Exception as e:
        logger.error(f"Error scraping {url}: {e}")
        return {"error": str(e), "rows": []}


# ============================================================================
# WORKER
# ============================================================================


async def worker(
    worker_id: int,
    queue: Queue,
    results: List[Dict],
    progress: Dict,
    stats: Stats,
    url_to_csv_brand: Dict[str, str],
):
    """Worker task: process URLs from queue."""
    throttle = AdaptiveThrottle()
    crash_count = 0

    while True:
        try:
            # Fetch URL from queue
            url = None
            try:
                url = queue.get_nowait()
            except asyncio.QueueEmpty:
                await asyncio.sleep(1)
                continue

            if url is None:
                break

            await stats.increment_processed()

            # Skip if already processed
            if url in progress["completed"] or url in progress["failed"]:
                queue.task_done()
                continue

            progress["in_progress"].append(url)

            # Create fresh browser context per worker
            async with async_playwright() as p:
                browser = await p.chromium.launch(
                    headless=False,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--disable-dev-shm-usage",
                    ],
                )

                user_agent = USER_AGENTS[worker_id % len(USER_AGENTS)]
                context, page = await setup_context(browser, user_agent, worker_id)

                try:
                    await throttle.wait()

                    result = await scrape_pdp(page, url)

                    if result and result.get("error") is None:
                        # Successful extraction
                        for row in result.get("rows", []):
                            results.append(row)

                        await stats.increment_success()
                        progress["completed"].append(url)
                        throttle.on_success()
                    elif result and result.get("error") == "RATE_LIMITED":
                        # Rate limited
                        throttle.on_rate_limit()
                        progress["in_progress"].remove(url)
                        queue.put_nowait(url)  # Re-queue
                        await stats.increment_errors()
                    else:
                        # Other error
                        error_msg = result.get("error") if result else "Unknown"
                        row = {
                            "url": url,
                            "product_name": "",
                            "brand": url_to_csv_brand.get(url, ""),
                            "brand_type": classify_brand(url_to_csv_brand.get(url, "")),
                            "color": "",
                            "current_price": None,
                            "current_price_formatted": "",
                            "original_price": None,
                            "original_price_formatted": "",
                            "on_sale": False,
                            "discount_pct": None,
                            "rise": None,
                            "leg_shape": None,
                            "fit": None,
                            "fabric_raw": "",
                            "fabric_parsed": None,
                            "pct_cotton": None,
                            "pct_natural_fiber": None,
                            "non_basic": False,
                            "breadcrumb": "",
                            "total_colors": 0,
                            "total_sizes": 0,
                            "sizes_list": "",
                            "average_rating": None,
                            "review_count": None,
                            "image_count": 0,
                            "feature_bullets": "",
                            "page_text": "",
                            "product_details": "",
                            "size_and_fit": "",
                            "length_hit": None,
                            "inseam": None,
                            "retries": 0,
                            "timestamp": datetime.now().isoformat(),
                            "error": error_msg,
                        }
                        results.append(row)
                        progress["failed"].append(url)
                        await stats.increment_errors()

                    progress["in_progress"].remove(url)

                except Exception as e:
                    logger.error(f"Worker {worker_id} crashed: {e}")
                    crash_count += 1
                    await stats.increment_crashes()
                    throttle.on_crash()

                    if url in progress["in_progress"]:
                        progress["in_progress"].remove(url)

                    if crash_count >= CRASH_COUNTER_MAX:
                        logger.error(
                            f"Worker {worker_id} exceeded crash limit ({crash_count})"
                        )
                        break

                finally:
                    await page.close()
                    await context.close()
                    await browser.close()

            queue.task_done()

        except KeyboardInterrupt:
            logger.info(f"Worker {worker_id} interrupted")
            break
        except Exception as e:
            logger.error(f"Worker {worker_id} fatal error: {e}")
            break


# ============================================================================
# MAIN
# ============================================================================


async def main():
    """Main entry point."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_file = os.path.join(script_dir, "macys.csv")
    progress_file = os.path.join(script_dir, "macys_pdp_progress.json")
    output_file = os.path.join(script_dir, "macys_pdp_results.xlsx")

    # Load URLs
    if not os.path.exists(csv_file):
        logger.error(f"CSV file not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)
    urls = df.iloc[:, 0].tolist()  # First column
    url_to_csv_brand = {url: df.iloc[i, 1] if len(df.columns) > 1 else "" for i, url in enumerate(urls)}

    logger.info(f"Loaded {len(urls)} URLs from {csv_file}")

    # Initialize
    progress = load_progress(progress_file)
    results = []
    stats = Stats()
    queue = Queue()

    # Populate queue with unprocessed URLs
    for url in urls:
        if url not in progress["completed"] and url not in progress["failed"]:
            queue.put_nowait(url)

    logger.info(
        f"Queue: {queue.qsize()} URLs ({len(progress['completed'])} completed, "
        f"{len(progress['failed'])} failed)"
    )

    # Spawn workers
    workers = [
        asyncio.create_task(
            worker(i, queue, results, progress, stats, url_to_csv_brand)
        )
        for i in range(NUM_WORKERS)
    ]

    # Auto-save loop
    async def autosave_loop():
        while True:
            await asyncio.sleep(AUTOSAVE_INTERVAL_SECONDS)
            save_progress(progress_file, progress)
            if results:
                save_to_excel(results, output_file)
            current_stats = await stats.get_stats()
            logger.info(f"Auto-save: {current_stats}")

    autosave_task = asyncio.create_task(autosave_loop())

    # Wait for queue to empty
    try:
        await asyncio.wait_for(queue.join(), timeout=3600)  # 1 hour max
    except asyncio.TimeoutError:
        logger.warning("Queue processing timeout (1 hour)")
    except KeyboardInterrupt:
        logger.info("Keyboard interrupt received")

    # Cancel workers
    for worker_task in workers:
        worker_task.cancel()

    autosave_task.cancel()

    # Final save
    save_progress(progress_file, progress)
    if results:
        save_to_excel(results, output_file)

    # Final stats
    final_stats = await stats.get_stats()
    logger.info(f"Final stats: {final_stats}")
    logger.info(f"Results saved to {output_file}")
    logger.info(f"Progress saved to {progress_file}")


if __name__ == "__main__":
    asyncio.run(main())
