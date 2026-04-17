#!/usr/bin/env python3
"""
Master Scraper Orchestrator — Cycles Between Retailers
========================================================
Instead of hammering one site continuously, this script runs each retailer
for a short batch (N URLs), then rotates to the next site. This distributes
traffic across retailers and reduces bot detection risk.

Usage:
    python3 master_scraper.py                  # run all retailers
    python3 master_scraper.py walmart amazon   # run specific ones only
    python3 master_scraper.py --batch 3        # 3 URLs per retailer per cycle

How it works:
    1. Each retailer gets its own Playwright browser + context
    2. The orchestrator runs BATCH_PER_SITE URLs on one retailer, then moves on
    3. After cycling through all retailers, it starts over
    4. Each retailer has a cooldown period between its batches
    5. Progress is saved per-retailer (same JSON/Excel files as individual scrapers)
    6. Ctrl+C saves all progress and exits cleanly

This is NOT a replacement for the individual scrapers — it imports and reuses
their extraction logic. Think of it as a scheduler that calls each scraper
in round-robin fashion.
"""

import asyncio
import importlib
import json
import os
import random
import sys
import time
import csv
from datetime import datetime
from playwright.async_api import async_playwright


# ── Configuration ───────────────────────────────────────────────────────────

BATCH_PER_SITE = 5          # URLs per retailer per rotation cycle
COOLDOWN_BETWEEN_SITES = 10  # seconds pause when switching retailers
COOLDOWN_BETWEEN_CYCLES = 30 # seconds pause after completing a full cycle
BROWSER_RESTART_EVERY = 15   # restart browser after this many batches per site
MAX_CONSECUTIVE_FAILS = 10   # skip a retailer if it fails this many in a row

# ── Retailer Definitions ───────────────────────────────────────────────────

RETAILERS = {
    'walmart': {
        'csv': 'walmart.csv',
        'csv_col': 0,           # URL column index
        'progress_json': 'walmart_pdp_progress.json',
        'results_xlsx': 'walmart_pdp_results.xlsx',
        'scraper_module': 'walmart_pdp_scraper',
        'display_name': 'Walmart',
        'base_delay': 2.0,
        'block_images_only': True,  # keep JS/CSS for Perimeter X
    },
    'amazon': {
        'csv': 'amazon.csv',
        'csv_col': 0,
        'progress_json': 'amazon_pdp_progress.json',
        'results_xlsx': 'amazon_pdp_results.xlsx',
        'scraper_module': 'amazon_pdp_scraper',
        'display_name': 'Amazon',
        'base_delay': 3.0,
        'block_images_only': False,  # block images + fonts
    },
    'ae': {
        'csv': 'ae.csv',
        'csv_col': 1,           # second column 'x-link-to href'
        'progress_json': 'ae_pdp_progress.json',
        'results_xlsx': 'ae_pdp_results.xlsx',
        'scraper_module': 'ae_pdp_scraper',
        'display_name': 'American Eagle',
        'base_delay': 2.0,
        'block_images_only': False,
    },
    'kohls': {
        'csv': 'kohls.csv',
        'csv_col': 0,
        'progress_json': 'kohls_pdp_progress.json',
        'results_xlsx': 'kohls_pdp_results.xlsx',
        'scraper_module': 'kohls_pdp_scraper',
        'display_name': "Kohl's",
        'base_delay': 2.0,
        'block_images_only': True,
    },
    'oldnavy': {
        'csv': 'oldnavy.csv',
        'csv_col': 0,
        'progress_json': 'oldnavy_pdp_progress.json',
        'results_xlsx': 'oldnavy_pdp_results.xlsx',
        'scraper_module': 'oldnavy_pdp_scraper',
        'display_name': 'Old Navy',
        'base_delay': 2.0,
        'block_images_only': False,
    },
    'macys': {
        'csv': 'macys.csv',
        'csv_col': 0,
        'progress_json': 'macys_pdp_progress.json',
        'results_xlsx': 'macys_pdp_results.xlsx',
        'scraper_module': 'macys_pdp_scraper',
        'display_name': "Macy's",
        'base_delay': 3.0,
        'block_images_only': True,  # keep JS/CSS for Akamai
    },
    'levis': {
        'csv': 'levi.csv',
        'csv_col': 0,
        'progress_json': 'levis_pdp_progress.json',
        'results_xlsx': 'levis_pdp_results.xlsx',
        'scraper_module': 'levis_pdp_scraper',
        'display_name': "Levi's",
        'base_delay': 3.0,
        'block_images_only': True,  # keep JS/CSS for Cloudflare
    },
}

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]


# ── Retailer State ─────────────────────────────────────────────────────────

class RetailerState:
    """Tracks state for a single retailer across rotation cycles."""
    def __init__(self, key, config, sdir):
        self.key = key
        self.config = config
        self.sdir = sdir
        self.display_name = config['display_name']

        # Load URLs
        self.urls = self._load_urls()
        self.total_urls = len(self.urls)

        # Load progress
        self.processed = set()
        self._load_progress()

        # Results (reloaded from Excel)
        self.results = []
        self._load_results()

        # Runtime state
        self.consecutive_fails = 0
        self.completed_count = 0
        self.failed_count = 0
        self.blocked_count = 0
        self.rows_generated = 0
        self.skipped = False
        self.skip_reason = ''

    def _load_urls(self):
        csv_path = os.path.join(self.sdir, self.config['csv'])
        if not os.path.exists(csv_path):
            return []
        urls = []
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # skip header
                col = self.config['csv_col']
                for row in reader:
                    if row and len(row) > col:
                        url = row[col].strip()
                        if url.startswith('http'):
                            urls.append(url)
        except Exception as e:
            print(f"  Error reading {csv_path}: {e}")
        return urls

    def _load_progress(self):
        p = os.path.join(self.sdir, self.config['progress_json'])
        if os.path.exists(p):
            try:
                with open(p) as f:
                    data = json.load(f)
                self.processed = set(data.get('processed', []))
            except Exception:
                pass

    def _load_results(self):
        xlsx_path = os.path.join(self.sdir, self.config['results_xlsx'])
        if os.path.exists(xlsx_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(xlsx_path, read_only=True)
                ws = wb.active
                headers = [c.value for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rec = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i]:
                            rec[headers[i]] = val if val is not None else ''
                    if rec.get('url'):
                        self.results.append(rec)
                wb.close()
            except Exception:
                pass

    def save_progress(self):
        p = os.path.join(self.sdir, self.config['progress_json'])
        with open(p, 'w') as f:
            json.dump({
                'processed': list(self.processed),
                'last_save': datetime.now().isoformat(),
                'total_processed': len(self.processed),
            }, f)

    @property
    def remaining(self):
        return [u for u in self.urls if u not in self.processed]

    @property
    def is_done(self):
        return len(self.remaining) == 0

    def next_batch(self, size):
        """Get the next batch of URLs to process."""
        return self.remaining[:size]


# ── Browser Management ─────────────────────────────────────────────────────

async def create_browser_page(pw, retailer_config, ua_index=0):
    """Create a fresh browser + context + page for a retailer."""
    browser = await pw.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-gpu",
        "--window-size=1440,900",
    ])

    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )

    # Add retailer-specific cookies
    if 'amazon' in retailer_config.get('csv', ''):
        await ctx.add_cookies([
            {"name": "sp-cdn", "value": '"L5Z9:US"', "domain": ".amazon.com", "path": "/"},
            {"name": "lc-main", "value": "en_US", "domain": ".amazon.com", "path": "/"},
            {"name": "i18n-prefs", "value": "USD", "domain": ".amazon.com", "path": "/"},
        ])

    page = await ctx.new_page()

    # Resource blocking
    async def block_resources(route):
        await route.abort()

    if retailer_config.get('block_images_only'):
        await page.route('**/*.{png,jpg,jpeg,gif,svg,ico,webp}', block_resources)
    else:
        await page.route('**/*.{png,jpg,jpeg,gif,svg,ico,webp,woff,woff2,ttf}', block_resources)

    return browser, ctx, page


async def close_browser(browser, ctx, page):
    """Safely close browser resources."""
    try:
        await page.close()
    except Exception:
        pass
    try:
        await ctx.close()
    except Exception:
        pass
    try:
        await browser.close()
    except Exception:
        pass


# ── Generic Single-URL Scraper ─────────────────────────────────────────────

async def scrape_single_url(page, url, retailer_key, base_delay):
    """
    Scrape a single URL using the retailer's module.
    Returns a list of row dicts (one per color) or error rows.
    """
    module_name = RETAILERS[retailer_key]['scraper_module']

    try:
        mod = importlib.import_module(module_name)
    except ImportError as e:
        return [{'url': url, 'error': f'Module import failed: {e}', 'timestamp': datetime.now().isoformat()}]

    # Navigate to the URL
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=25000)
        status = resp.status if resp else 0
    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str:
            raise  # Let caller handle browser crash
        return [{'url': url, 'error': f'Navigation error: {str(e)[:150]}', 'timestamp': datetime.now().isoformat()}]

    # Check for blocks
    if status in (403, 429, 503):
        return [{'url': url, 'error': f'HTTP {status}', 'timestamp': datetime.now().isoformat()}]

    # Check for CAPTCHA
    try:
        content = await page.content()
        content_lower = content.lower()[:3000]
        if 'captcha' in content_lower or 'robot' in content_lower or 'press & hold' in content_lower:
            return [{'url': url, 'error': 'CAPTCHA/bot detection', 'timestamp': datetime.now().isoformat()}]
    except Exception:
        pass

    # Wait for page to render
    await asyncio.sleep(random.uniform(base_delay, base_delay * 2))

    # Human-like scroll
    try:
        await page.evaluate('window.scrollBy(0, Math.floor(Math.random() * 400 + 200))')
        await asyncio.sleep(random.uniform(0.5, 1.5))
    except Exception:
        pass

    # Use the retailer's parse function based on which module we loaded
    # Each retailer scraper has different extraction approaches:
    try:
        if retailer_key == 'walmart':
            html_content = await page.content()
            if '__NEXT_DATA__' not in html_content:
                return [{'url': url, 'error': 'No __NEXT_DATA__', 'timestamp': datetime.now().isoformat()}]
            rows = mod.parse_walmart_pdp(html_content, url)

        elif retailer_key == 'amazon':
            # Wait for inline twister
            try:
                await page.wait_for_selector('#tp-inline-twister-dim-values-container', timeout=5000)
                await page.wait_for_selector('#tp-inline-twister-dim-values-container ul li img', timeout=3000)
            except Exception:
                pass
            await asyncio.sleep(1)

            # Extract shared data
            raw_data = await page.evaluate(mod.EXTRACT_JS)
            if not raw_data.get('title'):
                return [{'url': url, 'error': 'No product title', 'timestamp': datetime.now().isoformat()}]

            # Per-color pricing via clicking
            colors = raw_data.get('colors', [])
            per_color_prices = {}
            selected = raw_data.get('selected_color', '')
            if selected:
                per_color_prices[selected] = {
                    'current_price': raw_data.get('current_price', ''),
                    'original_price': raw_data.get('original_price', ''),
                }

            if len(colors) > 1:
                for color_name in colors:
                    if color_name in per_color_prices:
                        continue
                    try:
                        clicked = await page.evaluate("""
                            (targetColor) => {
                                const container = document.querySelector('#tp-inline-twister-dim-values-container') ||
                                                  document.getElementById('variation_color_name') ||
                                                  document.getElementById('twister');
                                if (!container) return false;
                                const imgs = container.querySelectorAll('li img, button img');
                                for (const img of imgs) {
                                    if (img.alt && img.alt.trim() === targetColor) {
                                        const clickTarget = img.closest('li') || img.closest('button') || img;
                                        clickTarget.click();
                                        return true;
                                    }
                                }
                                return false;
                            }
                        """, color_name)
                        if clicked:
                            await asyncio.sleep(random.uniform(1.0, 2.0))
                            price_data = await page.evaluate(mod.EXTRACT_PRICE_JS)
                            per_color_prices[color_name] = {
                                'current_price': price_data.get('current_price', ''),
                                'original_price': price_data.get('original_price', ''),
                            }
                        else:
                            per_color_prices[color_name] = {
                                'current_price': raw_data.get('current_price', ''),
                                'original_price': raw_data.get('original_price', ''),
                            }
                    except Exception:
                        per_color_prices[color_name] = {
                            'current_price': raw_data.get('current_price', ''),
                            'original_price': raw_data.get('original_price', ''),
                        }

            raw_data['per_color_prices'] = per_color_prices
            rows = mod.parse_amazon_pdp(raw_data, url)

        elif retailer_key == 'levis':
            # Levi's uses __NEXT_DATA__
            html_content = await page.content()
            if hasattr(mod, 'parse_levis_pdp'):
                rows = mod.parse_levis_pdp(html_content, url)
            elif hasattr(mod, 'parse_levi_pdp'):
                rows = mod.parse_levi_pdp(html_content, url)
            else:
                # Fallback: try extract_next_data + parse
                rows = [{'url': url, 'error': 'No parse function found in module', 'timestamp': datetime.now().isoformat()}]

        else:
            # For kohls, oldnavy, macys, ae — use page.evaluate with their EXTRACT_JS
            if hasattr(mod, 'EXTRACT_JS'):
                raw_data = await page.evaluate(mod.EXTRACT_JS)
                # Find the parse function
                parse_fn = None
                for name in ['parse_pdp', 'parse_kohls_pdp', 'parse_oldnavy_pdp',
                             'parse_macys_pdp', 'parse_ae_pdp', 'parse_amazon_pdp']:
                    if hasattr(mod, name):
                        parse_fn = getattr(mod, name)
                        break
                if parse_fn:
                    rows = parse_fn(raw_data, url)
                else:
                    rows = [{'url': url, 'error': 'No parse function', 'timestamp': datetime.now().isoformat()}]
            else:
                # Module might use HTML parsing instead
                html_content = await page.content()
                parse_fn = None
                for name in dir(mod):
                    if name.startswith('parse_') and 'pdp' in name.lower():
                        parse_fn = getattr(mod, name)
                        break
                if parse_fn:
                    rows = parse_fn(html_content, url)
                else:
                    rows = [{'url': url, 'error': 'No extraction method', 'timestamp': datetime.now().isoformat()}]

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str:
            raise  # Browser crashed
        rows = [{'url': url, 'error': f'Extract error: {str(e)[:150]}', 'timestamp': datetime.now().isoformat()}]

    return rows


# ── Save Results Using Retailer's Own Excel Function ───────────────────────

def save_retailer_results(retailer_key, state):
    """Save results using the retailer's own save_to_excel function."""
    try:
        mod = importlib.import_module(RETAILERS[retailer_key]['scraper_module'])
        if hasattr(mod, 'save_to_excel'):
            mod.save_to_excel(state.results, state.sdir)
    except Exception as e:
        # Fallback: basic Excel save
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            out = os.path.join(state.sdir, state.config['results_xlsx'])
            wb = Workbook()
            ws = wb.active
            ws.title = f'{state.display_name} Results'
            if state.results:
                headers = list(state.results[0].keys())
                hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
                hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = hfont
                    c.fill = hfill
                for i, data in enumerate(state.results, 2):
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=i, column=col, value=data.get(h, ''))
                ws.freeze_panes = 'A2'
            wb.save(out)
        except Exception as e2:
            print(f"    Could not save Excel for {state.display_name}: {e2}")


# ── Main Orchestrator ──────────────────────────────────────────────────────

async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))

    # Parse CLI args for retailer selection and batch size
    args = sys.argv[1:]
    batch_size = BATCH_PER_SITE
    selected_keys = []

    i = 0
    while i < len(args):
        if args[i] == '--batch' and i + 1 < len(args):
            batch_size = int(args[i + 1])
            i += 2
        else:
            if args[i] in RETAILERS:
                selected_keys.append(args[i])
            i += 1

    if not selected_keys:
        selected_keys = list(RETAILERS.keys())

    # Initialize retailer states
    states = {}
    for key in selected_keys:
        config = RETAILERS[key]
        state = RetailerState(key, config, sdir)
        if not state.urls:
            print(f"  {config['display_name']}: No CSV found or empty — skipping")
            continue
        if state.is_done:
            print(f"  {config['display_name']}: All {state.total_urls} URLs already processed — skipping")
            continue
        states[key] = state
        remaining = len(state.remaining)
        done = len(state.processed)
        print(f"  {config['display_name']}: {remaining} remaining / {state.total_urls} total ({done} already done, {len(state.results)} result rows)")

    if not states:
        print("\nAll retailers are done or have no URLs!")
        return

    active_keys = list(states.keys())
    print(f"\n{'=' * 70}")
    print(f"Master Scraper — {len(active_keys)} retailers, {batch_size} URLs per rotation")
    print(f"Order: {' → '.join(states[k].display_name for k in active_keys)}")
    print(f"Total remaining: {sum(len(s.remaining) for s in states.values())} URLs")
    print(f"{'=' * 70}\n")

    t0 = time.time()
    cycle_count = 0

    try:
        while active_keys:
            cycle_count += 1
            print(f"\n{'━' * 50}")
            print(f"Cycle {cycle_count} — {len(active_keys)} active retailers")
            print(f"{'━' * 50}")

            completed_this_cycle = False

            for retailer_key in list(active_keys):
                state = states[retailer_key]

                if state.is_done:
                    print(f"\n  ✓ {state.display_name}: COMPLETE ({len(state.results)} rows)")
                    active_keys.remove(retailer_key)
                    continue

                if state.skipped:
                    print(f"\n  ✗ {state.display_name}: Skipped — {state.skip_reason}")
                    continue

                batch = state.next_batch(batch_size)
                if not batch:
                    active_keys.remove(retailer_key)
                    continue

                remaining = len(state.remaining)
                print(f"\n  ▶ {state.display_name}: scraping {len(batch)} URLs ({remaining} remaining)")

                # Create fresh browser for this retailer batch
                pw = None
                browser = None
                try:
                    pw = await async_playwright().start()
                    browser, ctx, page = await create_browser_page(
                        pw, state.config, ua_index=cycle_count
                    )

                    for url_idx, url in enumerate(batch):
                        if url in state.processed:
                            continue

                        print(f"    [{url_idx + 1}/{len(batch)}] {url[:70]}...")

                        try:
                            # Random delay before each URL
                            delay = state.config['base_delay']
                            await asyncio.sleep(random.uniform(delay, delay * 2.5))

                            rows = await scrape_single_url(
                                page, url, retailer_key, delay
                            )

                            state.results.extend(rows)
                            state.processed.add(url)

                            has_error = any(r.get('error') for r in rows)
                            if has_error:
                                err_msg = rows[0].get('error', '')[:60]
                                print(f"      Error: {err_msg}")
                                state.failed_count += 1
                                state.consecutive_fails += 1

                                if state.consecutive_fails >= MAX_CONSECUTIVE_FAILS:
                                    state.skipped = True
                                    state.skip_reason = f"{MAX_CONSECUTIVE_FAILS} consecutive failures"
                                    print(f"      !! Skipping {state.display_name} — too many failures")
                                    break
                            else:
                                color_count = len(rows)
                                print(f"      OK — {color_count} color rows")
                                state.completed_count += 1
                                state.rows_generated += color_count
                                state.consecutive_fails = 0
                                completed_this_cycle = True

                        except Exception as e:
                            err_str = str(e).lower()
                            if 'pipe' in err_str or 'connection' in err_str:
                                print(f"      Browser crashed — ending batch")
                                break
                            else:
                                print(f"      Error: {str(e)[:80]}")
                                state.processed.add(url)
                                state.failed_count += 1

                    # Save after each retailer batch
                    state.save_progress()
                    save_retailer_results(retailer_key, state)
                    print(f"    Saved: {len(state.results)} rows, {len(state.processed)}/{state.total_urls} URLs")

                except Exception as e:
                    print(f"    Browser/Playwright error for {state.display_name}: {str(e)[:100]}")
                finally:
                    if browser:
                        await close_browser(browser, ctx, page)
                    if pw:
                        try:
                            await pw.stop()
                        except Exception:
                            pass

                # Cooldown between retailers
                if active_keys and retailer_key != active_keys[-1]:
                    wait = COOLDOWN_BETWEEN_SITES + random.uniform(0, 5)
                    print(f"    Cooling down {wait:.0f}s before next retailer...")
                    await asyncio.sleep(wait)

            # Remove completed retailers
            active_keys = [k for k in active_keys if not states[k].is_done and not states[k].skipped]

            if not active_keys:
                break

            # Cooldown between cycles
            wait = COOLDOWN_BETWEEN_CYCLES + random.uniform(0, 10)
            print(f"\n  Cycle {cycle_count} complete — cooling down {wait:.0f}s before next cycle...")
            await asyncio.sleep(wait)

    except KeyboardInterrupt:
        print(f"\n\n{'=' * 60}")
        print(f"Ctrl+C — saving all progress...")
        print(f"{'=' * 60}")

    # Final save for all retailers
    elapsed = time.time() - t0
    print(f"\n{'=' * 70}")
    print(f"Master Scraper Complete — {elapsed / 60:.1f} minutes, {cycle_count} cycles")
    print(f"{'=' * 70}")

    for key, state in states.items():
        state.save_progress()
        save_retailer_results(key, state)
        remaining = len(state.remaining)
        print(f"  {state.display_name:15s} | {state.completed_count:4d} OK | {state.failed_count:3d} fail | {state.rows_generated:5d} rows | {remaining:4d} remaining")

    print(f"\nRun again to continue where you left off.")


if __name__ == '__main__':
    asyncio.run(main())
