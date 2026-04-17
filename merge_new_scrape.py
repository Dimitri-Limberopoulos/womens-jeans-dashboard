#!/usr/bin/env python3
"""
merge_new_scrape.py — Post-scrape utility

Run AFTER the user completes scrapes on their Mac. Takes the three newly-scraped
XLSX files (macys_pdp_results.xlsx, kohls_pdp_results.xlsx, levis_pdp_results.xlsx)
and merges them into the existing v2 CSVs, deduplicated by (product_id, color) or URL.

Usage:
    python3 merge_new_scrape.py              # merge all three if new xlsx files exist
    python3 merge_new_scrape.py --retailer macys
    python3 merge_new_scrape.py --retailer kohls
    python3 merge_new_scrape.py --retailer levis

Outputs (written in place):
    macys_pdp_results_v2.csv   (appended with new rows)
    kohls_pdp_results_v2.csv   (appended with new rows)
    levis_pdp_results_v2.csv   (appended with new rows)

Also writes backups with ".bak" suffix before any modification.
"""

import argparse
import csv
import os
import re
import shutil
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Need openpyxl. Install with: pip3 install openpyxl")
    sys.exit(1)

BASE = os.path.dirname(os.path.abspath(__file__))


# ───────────────────────────── helpers ──────────────────────────────
def read_xlsx(path):
    if not os.path.exists(path):
        return None, None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, r)))
    return rows, headers


def read_csv_dicts(path):
    if not os.path.exists(path):
        return [], []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        return rows, reader.fieldnames or []


def backup(path):
    if os.path.exists(path):
        bak = f"{path}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(path, bak)
        print(f"  backup: {bak}")


def write_csv(path, rows, fieldnames):
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def safe_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ───────────────────────────── Macy's ───────────────────────────────
def transform_macys(row):
    """Map a rich macys scraper row -> v2 schema row."""
    url = row.get("url", "") or ""
    m = re.search(r"ID=(\d+)", url)
    product_id = m.group(1) if m else ""
    sale = safe_num(row.get("current_price"))
    reg = safe_num(row.get("original_price"))
    on_sale = bool(row.get("on_sale"))
    pct_off = safe_num(row.get("discount_pct")) or (
        round((reg - sale) / reg * 100, 1) if (reg and sale and reg > sale) else 0
    )
    return {
        "retailer": "Macys",
        "product_id": product_id,
        "product_name": row.get("product_name", ""),
        "brand": row.get("brand", ""),
        "color": row.get("color", ""),
        "sale_price": sale if sale is not None else "",
        "regular_price": reg if reg is not None else "",
        "on_sale": "true" if on_sale else "false",
        "percent_off": pct_off if pct_off else "",
        "url": url.split("#")[0] if url else "",
        "description": (row.get("feature_bullets") or row.get("product_details") or "")[:500],
    }


def merge_macys():
    new_rows, _ = read_xlsx(os.path.join(BASE, "macys_pdp_results.xlsx"))
    if not new_rows:
        print("[Macys] No new macys_pdp_results.xlsx found — skipping.")
        return
    v2_path = os.path.join(BASE, "macys_pdp_results_v2.csv")
    backup(v2_path)
    existing, fields = read_csv_dicts(v2_path)
    existing_keys = {(r["product_id"], r["color"]) for r in existing}

    added = 0
    for nr in new_rows:
        trow = transform_macys(nr)
        key = (trow["product_id"], trow["color"])
        if not trow["product_id"] or not trow["color"]:
            continue
        if key not in existing_keys:
            existing.append(trow)
            existing_keys.add(key)
            added += 1

    if not fields:
        fields = list(existing[0].keys()) if existing else list(trow.keys())
    write_csv(v2_path, existing, fields)
    print(f"[Macys] +{added} new rows → {v2_path} (total {len(existing)})")


# ───────────────────────────── Kohl's ───────────────────────────────
def transform_kohls(row):
    url = row.get("url", "") or ""
    m = re.search(r"prd-(\d+)", url)
    product_id = m.group(1) if m else ""
    sale = safe_num(row.get("current_price"))
    reg = safe_num(row.get("original_price"))
    on_sale = bool(row.get("on_sale"))
    return {
        "retailer": "Kohls",
        "product_id": product_id,
        "product_name": row.get("product_name", ""),
        "brand": row.get("brand", ""),
        "color": row.get("color", ""),
        "sale_price": sale if sale is not None else "",
        "regular_price": reg if reg is not None else "",
        "on_sale": "true" if on_sale else "false",
        "material": row.get("fabric_raw") or row.get("fabric_parsed") or row.get("material", ""),
        "stretch": row.get("stretch", ""),
        "rise": row.get("rise", ""),
        "fit": row.get("fit", "") or row.get("leg_shape", ""),
        "url": url.split("?")[0] if url else "",
        "description": (row.get("description") or row.get("feature_bullets", ""))[:500],
    }


def merge_kohls():
    new_rows, _ = read_xlsx(os.path.join(BASE, "kohls_pdp_results.xlsx"))
    if not new_rows:
        print("[Kohls] No new kohls_pdp_results.xlsx found — skipping.")
        return
    v2_path = os.path.join(BASE, "kohls_pdp_results_v2.csv")
    backup(v2_path)
    existing, fields = read_csv_dicts(v2_path)
    existing_keys = {(r["product_id"], r["color"]) for r in existing}

    added = 0
    for nr in new_rows:
        trow = transform_kohls(nr)
        key = (trow["product_id"], trow["color"])
        if not trow["product_id"] or not trow["color"]:
            continue
        if key not in existing_keys:
            existing.append(trow)
            existing_keys.add(key)
            added += 1

    if not fields:
        fields = list(existing[0].keys()) if existing else list(trow.keys())
    write_csv(v2_path, existing, fields)
    print(f"[Kohls] +{added} new rows → {v2_path} (total {len(existing)})")


# ───────────────────────────── Levi's ───────────────────────────────
def transform_levis(row):
    url = row.get("url", "") or ""
    sale = safe_num(row.get("current_price"))
    reg = safe_num(row.get("original_price") or row.get("regular_price"))
    on_sale = bool(row.get("on_sale")) or (sale is not None and reg is not None and sale < reg)
    return {
        "retailer": "Levis",
        "product_name": row.get("product_name", ""),
        "brand": row.get("brand", "Levi's") or "Levi's",
        "color": row.get("color", ""),
        "sale_price": sale if sale is not None else reg if reg else "",
        "regular_price": reg if reg is not None else "",
        "on_sale": "Yes" if on_sale else "No",
        "material": row.get("material", ""),
        "stretch": row.get("stretch", ""),
        "fit": row.get("fit", "") or row.get("leg_shape", ""),
        "rise": row.get("rise", ""),
        "url": url,
        "description": row.get("breadcrumbs") or row.get("description", ""),
    }


def merge_levis():
    new_rows, _ = read_xlsx(os.path.join(BASE, "levis_pdp_results.xlsx"))
    if not new_rows:
        print("[Levis] No new levis_pdp_results.xlsx found — skipping.")
        return
    v2_path = os.path.join(BASE, "levis_pdp_results_v2.csv")
    backup(v2_path)
    existing, fields = read_csv_dicts(v2_path)
    existing_keys = {r["url"] for r in existing}

    added = 0
    for nr in new_rows:
        trow = transform_levis(nr)
        key = trow["url"]
        if not key:
            continue
        if key not in existing_keys:
            existing.append(trow)
            existing_keys.add(key)
            added += 1

    if not fields:
        fields = list(existing[0].keys()) if existing else list(trow.keys())
    write_csv(v2_path, existing, fields)
    print(f"[Levis] +{added} new rows → {v2_path} (total {len(existing)})")


# ────────────────────────────── main ────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--retailer", choices=["macys", "kohls", "levis", "all"], default="all")
    args = ap.parse_args()
    if args.retailer in ("macys", "all"):
        merge_macys()
    if args.retailer in ("kohls", "all"):
        merge_kohls()
    if args.retailer in ("levis", "all"):
        merge_levis()
    print("\nAll merges complete. Next: run rebuild_dashboard.py")


if __name__ == "__main__":
    main()
