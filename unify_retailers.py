#!/usr/bin/env python3
"""
unify_retailers.py — Produce a single all-retailers CSV with a unified schema.

Reads every retailer's current output file, maps to must-have columns
(identity + pricing + product construction), writes:

    unified_retailer_data.csv

Schema (per the user's must-have selections):
    retailer, product_id, url, product_name, brand, brand_type, color,
    current_price, regular_price, on_sale, discount_pct,
    rise, leg_shape, fit, inseam,
    fabric, pct_cotton, pct_natural_fiber, stretch, closure
"""

import csv
import json
import os
import re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_CSV = os.path.join(HERE, "unified_retailer_data.csv")

# Owned brand lists (used to classify brand_type)
MACYS_OB = {"style & co", "i.n.c. international concepts", "inc international concepts",
            "and now this", "on 34th", "bar iii", "charter club", "alfani",
            "jm collection", "karen scott"}
KOHLS_OB = {"sonoma goods for life", "apt. 9", "lc lauren conrad", "so",
            "nine west", "simply vera vera wang"}
TARGET_OB = {"universal thread", "a new day", "wild fable", "knox rose",
             "auden", "goodfellow & co", "cat & jack"}
WALMART_OB = {"time and tru", "terra & sky", "no boundaries", "sofía jeans",
              "celebrity pink"}

UNIFIED_FIELDS = [
    "retailer", "product_id", "url", "product_name", "brand", "brand_type",
    "color", "current_price", "regular_price", "on_sale", "discount_pct",
    "rise", "leg_shape", "fit", "inseam",
    "fabric", "pct_cotton", "pct_natural_fiber", "stretch", "closure",
]

LEG_KEYWORDS = [
    ("wide leg", "Wide Leg"), ("wide-leg", "Wide Leg"),
    ("skinny", "Skinny"), ("jegging", "Jegging"),
    ("slim", "Slim"), ("straight", "Straight"),
    ("bootcut", "Bootcut"), ("boot-cut", "Bootcut"), ("boot cut", "Bootcut"),
    ("flare", "Flare"), ("baggy", "Baggy"), ("barrel", "Barrel"),
    ("boyfriend", "Boyfriend"), ("relaxed", "Relaxed"),
    ("crop", "Crop"), ("tapered", "Tapered"),
    ("mom", "Mom"), ("trouser", "Trouser"), ("loose", "Loose"),
]


# ───────────── helpers ─────────────
def norm_rise(s):
    if not s: return ""
    s = str(s).lower()
    if "high" in s: return "High"
    if "mid" in s: return "Mid"
    if "low" in s: return "Low"
    return str(s).strip()


def norm_leg(s, name=""):
    text = f"{s or ''} {name or ''}".lower()
    for kw, label in LEG_KEYWORDS:
        if kw in text:
            return label
    return ""


def parse_pct(s, nutrient):
    if not s: return None
    m = re.search(r"(\d+)%\s*" + nutrient, str(s), re.IGNORECASE)
    if m:
        return float(m.group(1))
    return None


def parse_cotton(s):
    return parse_pct(s, r"cotton")


def parse_natural(fabric):
    """Approx natural fiber % = cotton + linen + wool + hemp + silk + cashmere + lyocell + modal."""
    if not fabric: return None
    total = 0
    for mat in ["cotton", "linen", "wool", "hemp", "silk", "cashmere", "lyocell",
                "modal", "tencel", "ramie", "jute"]:
        v = parse_pct(fabric, mat)
        if v:
            total += v
    return total if total else None


def parse_inseam(s):
    if not s: return ""
    m = re.search(r"(\d+\.?\d*)\s*(?:in|inch|inches|\")", str(s), re.IGNORECASE)
    if m:
        return float(m.group(1))
    m = re.search(r"inseam[:\s]+(\d+\.?\d*)", str(s), re.IGNORECASE)
    if m:
        return float(m.group(1))
    return ""


def brand_type(retailer, brand):
    if not brand: return ""
    b = brand.lower().strip()
    if retailer == "Target":
        if b in TARGET_OB or any(ob in b for ob in TARGET_OB): return "Owned Brand"
        return "National Brand"
    if retailer == "Macys":
        if b in MACYS_OB or any(ob in b for ob in MACYS_OB): return "Owned Brand"
        return "National Brand"
    if retailer == "Kohls":
        if b in KOHLS_OB or any(ob in b for ob in KOHLS_OB): return "Owned Brand"
        return "National Brand"
    if retailer == "Walmart":
        if b in WALMART_OB or any(ob in b for ob in WALMART_OB): return "Owned Brand"
        return "National Brand"
    if retailer == "AE": return "National Brand"  # American Eagle is itself NB
    if retailer == "Amazon": return "Owned Brand"  # Amazon Essentials / Goodthreads owned
    if retailer == "OldNavy": return "Owned Brand"  # Old Navy is all private label
    if retailer == "Levis": return "National Brand"
    return ""


def read_xlsx(path):
    if not openpyxl or not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        out.append(dict(zip(headers, r)))
    return out


def read_csv_dicts(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def safe_f(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)):
        return float(v)
    # handle "$49.99", "49%", "$49.99 Reg.", etc.
    s = str(v).strip()
    m = re.search(r"-?\d+\.?\d*", s.replace(",", ""))
    if m:
        try: return float(m.group(0))
        except: return None
    return None


# ───────────── per-retailer transforms ─────────────
def from_target():
    rows = read_xlsx(os.path.join(HERE, "target_pdp_results.xlsx"))
    out = []
    for r in rows:
        fabric = r.get("Material") or ""
        out.append({
            "retailer": "Target",
            "product_id": r.get("Color TCIN") or r.get("Parent TCIN") or "",
            "url": r.get("Color Buy URL") or r.get("URL") or "",
            "product_name": r.get("Title", ""),
            "brand": r.get("Brand", ""),
            "brand_type": "Owned Brand" if (r.get("Owned Brand") in (True, "True", "true", 1)) else "National Brand",
            "color": r.get("Color", ""),
            "current_price": safe_f(r.get("Color Current Retail") or r.get("Color Current Price") or r.get("Current Price")),
            "regular_price": safe_f(r.get("Color Reg Retail") or r.get("Original Price")),
            "on_sale": (r.get("Price Type") or "").lower() != "regular",
            "discount_pct": safe_f(r.get("Discount %")) or safe_f(r.get("Save %")),
            "rise": norm_rise(r.get("Rise")),
            "leg_shape": norm_leg(r.get("Fit"), r.get("Title", "")),
            "fit": r.get("Fit") or "",
            "inseam": parse_inseam(r.get("Inseam Length")),
            "fabric": fabric,
            "pct_cotton": safe_f(r.get("% Cotton")) or parse_cotton(fabric),
            "pct_natural_fiber": safe_f(r.get("% Natural Fiber")) or parse_natural(fabric),
            "stretch": r.get("Stretch", "") or "",
            "closure": r.get("Closure Style", "") or "",
        })
    print(f"  Target:   {len(out)}")
    return out


def from_walmart():
    rows = read_xlsx(os.path.join(HERE, "walmart_pdp_results.xlsx"))
    out = []
    for r in rows:
        fabric = r.get("fabric_material") or ""
        out.append({
            "retailer": "Walmart",
            "product_id": re.search(r"/ip/[^/]+/(\d+)", r.get("url","") or "").group(1) if re.search(r"/ip/[^/]+/(\d+)", r.get("url","") or "") else "",
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", ""),
            "brand_type": brand_type("Walmart", r.get("brand", "")),
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("current_price")),
            "regular_price": safe_f(r.get("original_price")),
            "on_sale": bool(r.get("on_sale")),
            "discount_pct": safe_f(r.get("discount_pct")),
            "rise": norm_rise(r.get("pant_rise")),
            "leg_shape": norm_leg(r.get("pant_leg_cut") or r.get("pant_style"), r.get("product_name","")),
            "fit": r.get("clothing_fit") or "",
            "inseam": parse_inseam(r.get("inseam")),
            "fabric": fabric,
            "pct_cotton": parse_cotton(fabric) or parse_cotton(r.get("fabric_pct","")),
            "pct_natural_fiber": parse_natural(fabric) or parse_natural(r.get("fabric_pct","")),
            "stretch": "",  # Walmart doesn't expose this cleanly
            "closure": r.get("fastener_type") or "",
        })
    print(f"  Walmart:  {len(out)}")
    return out


def from_amazon():
    rows = read_xlsx(os.path.join(HERE, "amazon_pdp_results.xlsx"))
    out = []
    for r in rows:
        fabric = r.get("fabric_raw") or r.get("fabric_parsed") or ""
        out.append({
            "retailer": "Amazon",
            "product_id": r.get("asin", ""),
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", ""),
            "brand_type": r.get("brand_type") or brand_type("Amazon", r.get("brand","")),
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("current_price")),
            "regular_price": safe_f(r.get("original_price")),
            "on_sale": bool(r.get("on_sale")),
            "discount_pct": safe_f(r.get("discount_pct")),
            "rise": norm_rise(r.get("rise")),
            "leg_shape": norm_leg(r.get("leg_shape"), r.get("product_name","")),
            "fit": r.get("fit") or "",
            "inseam": "",
            "fabric": fabric,
            "pct_cotton": safe_f(r.get("pct_cotton")) or parse_cotton(fabric),
            "pct_natural_fiber": safe_f(r.get("pct_natural_fiber")) or parse_natural(fabric),
            "stretch": "",
            "closure": "",
        })
    print(f"  Amazon:   {len(out)}")
    return out


def from_ae():
    rows = read_xlsx(os.path.join(HERE, "ae_pdp_results.xlsx"))
    out = []
    for r in rows:
        fabric = r.get("fabric_material") or ""
        out.append({
            "retailer": "AE",
            "product_id": r.get("sku", ""),
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", "") or "American Eagle",
            "brand_type": "National Brand",
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("current_price")),
            "regular_price": safe_f(r.get("original_price")),
            "on_sale": bool(r.get("on_sale")),
            "discount_pct": safe_f(r.get("discount_pct")),
            "rise": norm_rise(r.get("rise")),
            "leg_shape": norm_leg(r.get("leg_shape"), r.get("product_name","")),
            "fit": r.get("fit") or "",
            "inseam": parse_inseam(r.get("inseam")),
            "fabric": fabric,
            "pct_cotton": safe_f(r.get("cotton_pct")) or parse_cotton(fabric),
            "pct_natural_fiber": parse_natural(fabric),
            "stretch": "",
            "closure": "",
        })
    print(f"  AE:       {len(out)}")
    return out


def from_oldnavy():
    rows = read_xlsx(os.path.join(HERE, "oldnavy_pdp_results.xlsx"))
    out = []
    for r in rows:
        fabric = r.get("fabric_raw") or r.get("fabric_parsed") or ""
        out.append({
            "retailer": "OldNavy",
            "product_id": re.search(r"pid=(\d+)", r.get("url","") or "").group(1) if re.search(r"pid=(\d+)", r.get("url","") or "") else "",
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", "") or "Old Navy",
            "brand_type": "Owned Brand",
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("current_price")),
            "regular_price": safe_f(r.get("original_price")),
            "on_sale": bool(r.get("on_sale")),
            "discount_pct": safe_f(r.get("discount_pct")),
            "rise": norm_rise(r.get("rise")),
            "leg_shape": norm_leg(r.get("leg_shape"), r.get("product_name","")),
            "fit": r.get("fit") or "",
            "inseam": parse_inseam(r.get("inseam")),
            "fabric": fabric,
            "pct_cotton": safe_f(r.get("pct_cotton")) or parse_cotton(fabric),
            "pct_natural_fiber": safe_f(r.get("pct_natural_fiber")) or parse_natural(fabric),
            "stretch": "",
            "closure": "",
        })
    print(f"  OldNavy:  {len(out)}")
    return out


def from_macys():
    rows = read_csv_dicts(os.path.join(HERE, "macys_pdp_results_v2.csv"))
    out = []
    for r in rows:
        desc = r.get("description", "") or ""
        fabric = ""  # not in v2 schema; rely on description
        out.append({
            "retailer": "Macys",
            "product_id": r.get("product_id", ""),
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", ""),
            "brand_type": brand_type("Macys", r.get("brand","")),
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("sale_price")),
            "regular_price": safe_f(r.get("regular_price")),
            "on_sale": (r.get("on_sale","").lower() == "true"),
            "discount_pct": safe_f(r.get("percent_off")),
            "rise": norm_rise(desc) if "rise" in desc.lower() else "",
            "leg_shape": norm_leg("", r.get("product_name","")),
            "fit": "",
            "inseam": parse_inseam(desc),
            "fabric": fabric,
            "pct_cotton": parse_cotton(desc),
            "pct_natural_fiber": parse_natural(desc),
            "stretch": "",
            "closure": "",
        })
    print(f"  Macys:    {len(out)}")
    return out


def from_kohls():
    rows = read_csv_dicts(os.path.join(HERE, "kohls_pdp_results_v2.csv"))
    out = []
    for r in rows:
        fabric = r.get("material", "") or ""
        desc = r.get("description", "") or ""
        out.append({
            "retailer": "Kohls",
            "product_id": r.get("product_id", ""),
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", ""),
            "brand_type": brand_type("Kohls", r.get("brand","")),
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("sale_price")),
            "regular_price": safe_f(r.get("regular_price")),
            "on_sale": (r.get("on_sale","").lower() == "true"),
            "discount_pct": (round((safe_f(r.get("regular_price")) - safe_f(r.get("sale_price"))) / safe_f(r.get("regular_price")) * 100, 1)
                             if safe_f(r.get("regular_price")) and safe_f(r.get("sale_price")) and safe_f(r.get("regular_price")) > safe_f(r.get("sale_price"))
                             else None),
            "rise": norm_rise(r.get("rise")),
            "leg_shape": norm_leg(r.get("fit"), r.get("product_name","")),
            "fit": r.get("fit") or "",
            "inseam": parse_inseam(desc),
            "fabric": fabric,
            "pct_cotton": parse_cotton(fabric),
            "pct_natural_fiber": parse_natural(fabric),
            "stretch": r.get("stretch", ""),
            "closure": "",
        })
    print(f"  Kohls:    {len(out)}")
    return out


def from_levis():
    rows = read_csv_dicts(os.path.join(HERE, "levis_pdp_results_v2.csv"))
    out = []
    for r in rows:
        fabric = r.get("material", "") or ""
        out.append({
            "retailer": "Levis",
            "product_id": re.search(r"/p/(\w+)", r.get("url","") or "").group(1) if re.search(r"/p/(\w+)", r.get("url","") or "") else "",
            "url": r.get("url", ""),
            "product_name": r.get("product_name", ""),
            "brand": r.get("brand", "") or "Levi's",
            "brand_type": "National Brand",
            "color": r.get("color", ""),
            "current_price": safe_f(r.get("sale_price")),
            "regular_price": safe_f(r.get("regular_price")),
            "on_sale": (str(r.get("on_sale","")).lower() in ("yes","true","1")),
            "discount_pct": None,
            "rise": norm_rise(r.get("rise")),
            "leg_shape": norm_leg(r.get("fit"), r.get("product_name","")),
            "fit": r.get("fit") or "",
            "inseam": "",
            "fabric": fabric,
            "pct_cotton": parse_cotton(fabric),
            "pct_natural_fiber": parse_natural(fabric),
            "stretch": r.get("stretch", ""),
            "closure": "",
        })
    print(f"  Levis:    {len(out)}")
    return out


def main():
    print("Building unified retailer CSV...")
    print("Source counts:")
    all_rows = []
    all_rows += from_target()
    all_rows += from_walmart()
    all_rows += from_amazon()
    all_rows += from_ae()
    all_rows += from_oldnavy()
    all_rows += from_macys()
    all_rows += from_kohls()
    all_rows += from_levis()

    with open(OUT_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=UNIFIED_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in all_rows:
            # Normalize bool -> 'true'/'false'
            if isinstance(r.get("on_sale"), bool):
                r["on_sale"] = "true" if r["on_sale"] else "false"
            w.writerow(r)

    print(f"\nWrote {OUT_CSV}")
    print(f"Total rows: {len(all_rows)}")

    # Per-retailer summary
    by_retailer = defaultdict(int)
    for r in all_rows:
        by_retailer[r["retailer"]] += 1
    print("\nRows per retailer:")
    for k in sorted(by_retailer, key=lambda x: -by_retailer[x]):
        print(f"  {k:10s} {by_retailer[k]:>5d}")


if __name__ == "__main__":
    main()
