#!/usr/bin/env python3
"""
Test All Scrapers — 5 URLs Per Retailer (No Amazon)
====================================================
Validates each retailer scraper by running 5 sample URLs using each module's
OWN browser setup, extraction JS, and parse functions — exactly as they'd run
in production.

Usage:
    python3 test_all_scrapers.py                  # test all 6 retailers
    python3 test_all_scrapers.py walmart kohls     # test specific ones
    python3 test_all_scrapers.py --skip levis      # test all except levis

Output:
    - test_results.xlsx — one tab per retailer with all extracted rows
    - Console report with pass/fail per URL + final summary table
"""

import asyncio
import csv
import importlib
import json
import os
import random
import sys
import time
import traceback
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Test URLs: 5 hand-picked per retailer (NO Amazon — already done) ────────

TEST_URLS = {
    'walmart': [
        'https://www.walmart.com/ip/Sofia-Jeans-Women-s-and-Women-s-Plus-Melisa-Flare-High-Rise-Jeans-Sizes-0-28W/5330099129',
        'https://www.walmart.com/ip/No-Boundaries-High-Rise-Skinny-Jeans-28-5-Inseam-1-or-2-Pack-Women-s-and-Women-s-Plus/5113175886',
        'https://www.walmart.com/ip/Time-and-Tru-Women-s-High-Rise-Straight-Leg-Jean-30-Inseam-Sizes-0-20-10S-16S/17842562949',
        'https://www.walmart.com/ip/No-Boundaries-Bootcut-Jean-Women-s/17896213268',
        'https://www.walmart.com/ip/HR-PATCH-POCKET-JN-S/17906057683',
    ],
    'ae': [
        'https://www.ae.com/intl/en/p/women/jeans/baggy-wide-leg-jeans/ae-dreamy-drape-strigid-low-rise-ultra-wide-leg-jean/0437_5676_471',
        'https://www.ae.com/intl/en/p/women/jeans/baggy-wide-leg-jeans/ae-dreamy-drape-strigid-low-rise-baggy-ultra-wide-leg-jean/0437_6042_898',
        'https://www.ae.com/intl/en/p/women/jeans/flare-bootcut-jeans/ae-next-level-super-high-waisted-flare-jean/1436_6024_110',
        'https://www.ae.com/intl/en/p/women/jeans/baggy-wide-leg-jeans/ae-dreamy-drape-strigid-low-rise-ultra-wide-leg-jean/0437_5676_471',
        'https://www.ae.com/intl/en/p/women/jeans/flare-bootcut-jeans/ae-next-level-super-high-waisted-flare-jean/1436_6024_110',
    ],
    'kohls': [
        'https://www.kohls.com/product/prd-7626864/womens-sonoma-goods-for-life-high-rise-straight-leg-jeans.jsp',
        'https://www.kohls.com/product/prd-6893694/womens-sonoma-goods-for-life-high-rise-skinny-jeans.jsp',
        'https://www.kohls.com/product/prd-7437759/womens-lc-lauren-conrad-super-high-rise-wide-leg-jeans.jsp',
        'https://www.kohls.com/product/prd-7626838/womens-sonoma-goods-for-life-mid-rise-relaxed-boyfriend-jeans.jsp',
        'https://www.kohls.com/product/prd-6679547/womens-lc-lauren-conrad-super-high-rise-wildflower-wide-leg-cropped-pants.jsp',
    ],
    'oldnavy': [
        'https://oldnavy.gap.com/browse/product.do?pid=901158002',
        'https://oldnavy.gap.com/browse/product.do?pid=732591002',
        'https://oldnavy.gap.com/browse/product.do?pid=734874002',
        'https://oldnavy.gap.com/browse/product.do?pid=901143002',
        'https://oldnavy.gap.com/browse/product.do?pid=894168002',
    ],
    'macys': [
        'https://www.macys.com/shop/product/style-co-petite-high-rise-cropped-wide-leg-jeans-macys-exclusive',
        'https://www.macys.com/shop/product/bar-iii-womens-high-rise-barrel-leg-jeans-macys-exclusive',
        'https://www.macys.com/shop/product/i.n.c.-international-concepts-womens-mid-rise-bootcut-jeans-created-for-macys',
        'https://www.macys.com/shop/product/now-this-womens-seamed-cuff-wide-leg-jeans-macys-exclusive',
        'https://www.macys.com/shop/product/style-co-petite-mid-rise-straight-leg-denim-jeans-exclusively-at-macys',
    ],
    'levis': [
        'https://www.levi.com/US/en_US/clothing/women/jeans/wide-leg/318-shaping-wide-leg-womens-jeans/p/001PZ0016',
        'https://www.levi.com/US/en_US/clothing/women/jeans/loose/cinch-baggy-womens-jeans/p/001UP0028',
        'https://www.levi.com/US/en_US/clothing/women/jeans/loose/low-loose-womens-jeans/p/A55660062',
        'https://www.levi.com/US/en_US/clothing/women/jeans/straight/501-original-fit-womens-jeans/p/125010632',
        'https://www.levi.com/US/en_US/clothing/women/jeans/straight/501-90s-womens-jeans/p/A19590133',
    ],
}

# Scraper module names — each one has setup_context() and scrape_*_pdp()
SCRAPER_CONFIG = {
    'walmart': {
        'module': 'walmart_pdp_scraper',
        'scrape_fn': 'scrape_walmart_pdp',   # async def scrape_walmart_pdp(page, url, stats, throttle, retries=0) -> list
        'returns': 'list',                     # returns list of dicts directly
    },
    'ae': {
        'module': 'ae_pdp_scraper',
        'scrape_fn': 'scrape_ae_pdp',
        'returns': 'list',
    },
    'kohls': {
        'module': 'kohls_pdp_scraper',
        'scrape_fn': 'scrape_kohls_pdp',
        'returns': 'list',
    },
    'oldnavy': {
        'module': 'oldnavy_pdp_scraper',
        'scrape_fn': 'scrape_oldnavy_pdp',
        'returns': 'list',
    },
    'macys': {
        'module': 'macys_pdp_scraper',
        'scrape_fn': 'scrape_pdp',
        'returns': 'dict',                     # returns {"error": ..., "rows": [...]}
    },
    'levis': {
        'module': 'levis_pdp_scraper',
        'scrape_fn': 'scrape_levis_pdp',
        'returns': 'list',
    },
}

SDIR = os.path.dirname(os.path.abspath(__file__))


# ── ANSI Colors ────────────────────────────────────────────────────────────

class C:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    BOLD = '\033[1m'
    DIM = '\033[2m'
    END = '\033[0m'


# ── Lightweight Stats & Throttle (compatible with all scrapers) ────────────

class AdaptiveThrottle:
    """Minimal throttle that satisfies each scraper's expected interface."""
    def __init__(self, base_delay=3.0):
        self.lock = asyncio.Lock()
        self.base_delay = base_delay
        self.current_delay = base_delay
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    {C.YELLOW}Throttle UP: delay now {self.current_delay:.1f}s{C.END}")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 5 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0


class Stats:
    """Minimal stats tracker matching each scraper's expected interface."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = 0
        self.completed = 0
        self.failed = 0
        self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Excel Output ──────────────────────────────────────────────────────────

def save_test_results(all_results, output_path):
    """Save results to test_results.xlsx with one tab per retailer."""
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

    for retailer_key in ['walmart', 'ae', 'kohls', 'oldnavy', 'macys', 'levis']:
        rows = all_results.get(retailer_key, [])
        if not rows:
            continue

        ws = wb.create_sheet(title=retailer_key.upper())

        # Gather all unique headers from all rows
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)

        # Write header
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Write data
        for i, row_data in enumerate(rows, 2):
            for col, h in enumerate(headers, 1):
                val = row_data.get(h, '')
                if isinstance(val, bool):
                    val = 'Yes' if val else 'No'
                ws.cell(row=i, column=col, value=val)

        # Freeze header, auto-width
        ws.freeze_panes = 'A2'
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    # Summary tab
    ws_sum = wb.create_sheet(title='SUMMARY', index=0)
    sum_headers = ['Retailer', 'URLs Tested', 'URLs OK', 'URLs Blocked', 'URLs Failed',
                   'Total Rows', 'Avg Colors/URL', 'Sample Product', 'Sample Brand', 'Sample Price']
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row_num = 2
    for retailer_key in ['walmart', 'ae', 'kohls', 'oldnavy', 'macys', 'levis']:
        rows = all_results.get(retailer_key, [])
        urls_tested = len(TEST_URLS.get(retailer_key, []))
        # Count unique URLs that produced rows
        ok_urls = len(set(r.get('url', '') for r in rows if not r.get('error')))
        blocked = len(set(r.get('url', '') for r in rows if 'BLOCKED' in str(r.get('error', '')).upper() or '403' in str(r.get('error', ''))))
        failed = urls_tested - ok_urls
        total_rows = len([r for r in rows if not r.get('error')])
        avg_colors = round(total_rows / max(ok_urls, 1), 1)

        sample_name = ''
        sample_brand = ''
        sample_price = ''
        for r in rows:
            if not r.get('error'):
                sample_name = str(r.get('product_name', r.get('name', '')))[:60]
                sample_brand = str(r.get('brand', ''))
                sample_price = str(r.get('current_price', r.get('price', '')))
                break

        ws_sum.cell(row=row_num, column=1, value=retailer_key.upper())
        ws_sum.cell(row=row_num, column=2, value=urls_tested)
        ws_sum.cell(row=row_num, column=3, value=ok_urls)
        ws_sum.cell(row=row_num, column=4, value=blocked)
        ws_sum.cell(row=row_num, column=5, value=failed)
        ws_sum.cell(row=row_num, column=6, value=total_rows)
        ws_sum.cell(row=row_num, column=7, value=avg_colors)
        ws_sum.cell(row=row_num, column=8, value=sample_name)
        ws_sum.cell(row=row_num, column=9, value=sample_brand)
        ws_sum.cell(row=row_num, column=10, value=sample_price)
        row_num += 1

    ws_sum.freeze_panes = 'A2'
    for col_idx in range(1, len(sum_headers) + 1):
        ws_sum.column_dimensions[ws_sum.cell(row=1, column=col_idx).column_letter].width = 20

    wb.save(output_path)
    return output_path


# ── Per-Retailer Test Runner ─────────────────────────────────────────────

async def test_retailer(retailer_key, urls):
    """
    Test a single retailer using its OWN module's browser setup + scrape function.
    Returns a list of result row dicts (one per color per URL).
    """
    config = SCRAPER_CONFIG[retailer_key]
    module_name = config['module']

    print(f"\n{'='*70}")
    print(f"  {C.BOLD}{C.CYAN}Testing {retailer_key.upper()}{C.END}  ({len(urls)} URLs)")
    print(f"{'='*70}")

    # ── Import the scraper module ──
    try:
        mod = importlib.import_module(module_name)
        print(f"  {C.GREEN}✓{C.END} Loaded module: {module_name}")
    except ImportError as e:
        missing = str(e)
        print(f"  {C.RED}✗ Import failed: {missing}{C.END}")
        if 'pandas' in missing:
            print(f"    {C.YELLOW}Fix: pip3 install pandas{C.END}")
        elif 'openpyxl' in missing:
            print(f"    {C.YELLOW}Fix: pip3 install openpyxl{C.END}")
        return []
    except Exception as e:
        print(f"  {C.RED}✗ Failed to import {module_name}: {e}{C.END}")
        return []

    # ── Get the scrape function ──
    scrape_fn_name = config['scrape_fn']
    scrape_fn = getattr(mod, scrape_fn_name, None)
    if scrape_fn is None:
        print(f"  {C.RED}✗ Function {scrape_fn_name} not found in {module_name}{C.END}")
        return []
    print(f"  {C.GREEN}✓{C.END} Using: {scrape_fn_name}()")

    # ── Get the module's own setup_context ──
    setup_ctx_fn = getattr(mod, 'setup_context', None)
    if setup_ctx_fn is None:
        print(f"  {C.RED}✗ setup_context() not found in {module_name}{C.END}")
        return []
    print(f"  {C.GREEN}✓{C.END} Using module's own setup_context() for stealth")

    # ── Create stats & throttle ──
    stats = Stats()
    stats.total = len(urls)
    base_delay = 3.0
    if hasattr(mod, 'AdaptiveThrottle'):
        throttle = mod.AdaptiveThrottle()
        print(f"  {C.GREEN}✓{C.END} Using module's own AdaptiveThrottle")
    else:
        throttle = AdaptiveThrottle(base_delay)

    all_rows = []

    # ── Launch browser + use module's context setup ──
    pw_instance = await async_playwright().start()
    try:
        browser = await pw_instance.chromium.launch(headless=False, args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
        ])
        # Different scrapers have different setup_context signatures:
        #   walmart, ae, kohls, oldnavy, levis: setup_context(browser, ua_index=N) → context
        #   macys: setup_context(browser, user_agent, context_id) → (context, page)
        import inspect
        sig = inspect.signature(setup_ctx_fn)
        params = list(sig.parameters.keys())

        if 'user_agent' in params:
            # Macy's-style: (browser, user_agent, context_id)
            ua = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            result = await setup_ctx_fn(browser, ua, 0)
            if isinstance(result, tuple):
                ctx = result[0]  # (context, page) tuple
            else:
                ctx = result
        else:
            ctx = await setup_ctx_fn(browser, ua_index=random.randint(0, 9))

        # Route for resource blocking — match what each scraper does
        # Walmart, Kohl's, Macy's, Levi's keep JS/CSS (anti-bot)
        keep_js_css = retailer_key in ('walmart', 'kohls', 'macys', 'levis')
        async def route_handler(route):
            rtype = route.request.resource_type
            if rtype in ('image', 'media'):
                await route.abort()
            elif rtype == 'font' and retailer_key not in ('macys',):
                await route.abort()
            elif not keep_js_css and rtype in ('stylesheet',):
                await route.abort()
            else:
                await route.continue_()

        page = await ctx.new_page()
        await page.route('**/*', route_handler)
        print(f"  {C.GREEN}✓{C.END} Browser launched (headed — visible) with {retailer_key}-specific stealth")

        for i, url in enumerate(urls, 1):
            print(f"\n  [{i}/{len(urls)}] {C.DIM}{url[:80]}...{C.END}")

            try:
                # ── Call the module's actual scrape function ──
                if retailer_key == 'macys':
                    # Macy's scrape_pdp(page, url, retries) -> {"error":..., "rows":[...]}
                    result = await scrape_fn(page, url, retries=0)
                    if result and result.get('rows'):
                        rows = result['rows']
                    elif result and result.get('error'):
                        print(f"         {C.RED}✗ {result['error']}{C.END}")
                        rows = []
                    else:
                        rows = []
                else:
                    # All others: scrape_*_pdp(page, url, stats, throttle, retries=0) -> [dicts]
                    rows = await scrape_fn(page, url, stats, throttle, retries=0)

                # ── Evaluate results ──
                if not rows:
                    print(f"         {C.RED}✗ 0 rows returned{C.END}")
                    continue

                # Check if first row has an error
                first = rows[0]
                if first.get('error') and not first.get('product_name') and not first.get('brand'):
                    err = first.get('error', 'unknown')
                    print(f"         {C.RED}✗ Error: {err}{C.END}")
                    all_rows.extend(rows)
                    continue

                # Detect hard blocks disguised as success (e.g., "Access Denied" as product name)
                name_check = str(first.get('product_name', '')).lower()
                if 'access denied' in name_check or 'blocked' in name_check or 'forbidden' in name_check or 'robot' in name_check:
                    print(f"         {C.RED}✗ HARD BLOCK — got '{first.get('product_name')}' instead of product data{C.END}")
                    print(f"         {C.RED}  Skipping remaining URLs for {retailer_key.upper()} — site is blocking us{C.END}")
                    all_rows.append({'url': url, 'error': f'Hard block: {first.get("product_name")}', 'retailer': retailer_key})
                    break  # Skip remaining URLs for this retailer

                # Success!
                name = str(first.get('product_name', first.get('name', '')))[:55]
                brand = str(first.get('brand', ''))
                colors = len(rows)
                price = first.get('current_price', first.get('price', ''))

                print(f"         {C.GREEN}✓ OK{C.END} — {name}")
                print(f"           Brand: {brand}  |  Colors/Rows: {colors}  |  Price: ${price}")

                # Show color names
                color_names = []
                for r in rows:
                    cn = r.get('color', r.get('color_name', ''))
                    if cn:
                        color_names.append(str(cn))
                if color_names:
                    print(f"           Colors: {', '.join(color_names[:10])}")

                # Show per-color pricing if prices vary
                prices = []
                for r in rows:
                    p = r.get('current_price', r.get('price', ''))
                    if p:
                        prices.append(str(p))
                unique_prices = set(prices)
                if len(unique_prices) > 1:
                    print(f"           {C.GREEN}Per-color pricing detected!{C.END} Prices: {', '.join(sorted(unique_prices)[:8])}")
                elif len(unique_prices) == 1 and colors > 1:
                    print(f"           {C.YELLOW}⚠ Same price (${prices[0]}) across all {colors} colors{C.END}")

                all_rows.extend(rows)

            except Exception as e:
                print(f"         {C.RED}✗ Exception: {str(e)[:120]}{C.END}")
                traceback.print_exc()
                all_rows.append({'url': url, 'error': str(e)[:200], 'retailer': retailer_key})

            # Delay between URLs
            if i < len(urls):
                delay = random.uniform(3.0, 6.0)
                print(f"         {C.DIM}(waiting {delay:.1f}s){C.END}")
                await asyncio.sleep(delay)

        await browser.close()
        print(f"\n  {C.BOLD}Stats:{C.END} completed={stats.completed}, failed={stats.failed}, blocked={stats.blocked}, rows={stats.rows_generated}")

    except Exception as e:
        print(f"  {C.RED}✗ Browser error: {e}{C.END}")
        traceback.print_exc()
    finally:
        await pw_instance.stop()

    return all_rows


# ── Summary Report ───────────────────────────────────────────────────────

def print_summary(all_results):
    """Print a final summary table of all test results."""
    print(f"\n\n{'='*70}")
    print(f"  {C.BOLD}FINAL TEST SUMMARY{C.END}")
    print(f"{'='*70}\n")

    header = f"  {'Retailer':<12} {'URLs':>5} {'OK':>4} {'Blocked':>8} {'Rows':>6} {'Avg Col':>8}  {'Notes'}"
    print(header)
    print(f"  {'─'*75}")

    total_urls = 0
    total_ok = 0
    total_rows = 0

    for retailer_key in ['walmart', 'ae', 'kohls', 'oldnavy', 'macys', 'levis']:
        rows = all_results.get(retailer_key, [])
        urls_tested = len(TEST_URLS.get(retailer_key, []))
        total_urls += urls_tested

        if not rows:
            print(f"  {C.RED}{retailer_key:<12}{C.END} {urls_tested:>5} {'0':>4} {'—':>8} {'0':>6} {'—':>8}  {'NOT TESTED / IMPORT FAILED'}")
            continue

        good_rows = [r for r in rows if not r.get('error')]
        error_rows = [r for r in rows if r.get('error')]
        ok_urls = len(set(r.get('url', '') for r in good_rows))
        blocked = len([r for r in error_rows if 'BLOCKED' in str(r.get('error', '')).upper() or '403' in str(r.get('error', '')) or 'CAPTCHA' in str(r.get('error', '')).upper()])
        total_good_rows = len(good_rows)
        avg_colors = round(total_good_rows / max(ok_urls, 1), 1)

        total_ok += ok_urls
        total_rows += total_good_rows

        status_color = C.GREEN if ok_urls == urls_tested else (C.YELLOW if ok_urls > 0 else C.RED)
        notes = []
        if blocked > 0:
            notes.append(f'{blocked} blocked')
        if total_good_rows > 0 and avg_colors <= 1.0:
            notes.append('⚠ only 1 color/URL?')
        note_str = ', '.join(notes) if notes else 'all good'

        print(f"  {status_color}{retailer_key:<12}{C.END} {urls_tested:>5} {ok_urls:>4} {blocked:>8} {total_good_rows:>6} {avg_colors:>8}  {note_str}")

    print(f"  {'─'*75}")
    overall_color = C.GREEN if total_ok == total_urls else (C.YELLOW if total_ok > 0 else C.RED)
    print(f"  {overall_color}{'TOTAL':<12}{C.END} {total_urls:>5} {total_ok:>4} {'':>8} {total_rows:>6}")

    print(f"\n  {C.BOLD}Verdict:{C.END} ", end='')
    if total_ok == total_urls:
        print(f"{C.GREEN}ALL SCRAPERS PASSING ✓{C.END}")
    elif total_ok >= total_urls * 0.6:
        print(f"{C.YELLOW}MOSTLY WORKING — check failures above{C.END}")
    else:
        print(f"{C.RED}SIGNIFICANT FAILURES — debug before full run{C.END}")


# ── Main ─────────────────────────────────────────────────────────────────

async def main():
    # Parse CLI args
    skip_list = set()
    only_list = set()

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == '--skip' and i + 1 < len(args):
            skip_list.add(args[i + 1].lower())
            i += 2
        elif args[i].startswith('--'):
            i += 1
        else:
            only_list.add(args[i].lower())
            i += 1

    # Determine which retailers to test
    retailer_order = ['walmart', 'ae', 'kohls', 'oldnavy', 'macys', 'levis']
    if only_list:
        retailers_to_test = [k for k in retailer_order if k in only_list]
    else:
        retailers_to_test = [k for k in retailer_order if k not in skip_list]

    if not retailers_to_test:
        print(f"{C.RED}No retailers selected for testing.{C.END}")
        print(f"Available: {', '.join(retailer_order)}")
        sys.exit(1)

    print(f"\n{C.BOLD}{'='*70}{C.END}")
    print(f"  {C.BOLD}WOMEN'S JEANS SCRAPER — TEST SUITE{C.END}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Testing: {', '.join(r.upper() for r in retailers_to_test)}")
    print(f"  URLs per retailer: 5")
    print(f"  Total URLs: {len(retailers_to_test) * 5}")
    print(f"  Note: Amazon excluded (already completed)")
    print(f"{'='*70}")

    all_results = {}  # retailer_key -> list of row dicts

    for idx, retailer_key in enumerate(retailers_to_test):
        try:
            urls = TEST_URLS[retailer_key]
            rows = await test_retailer(retailer_key, urls)
            all_results[retailer_key] = rows
        except (KeyboardInterrupt, asyncio.CancelledError):
            print(f"\n\n{C.YELLOW}Interrupted during {retailer_key}! Saving partial results...{C.END}")
            break
        except Exception as e:
            print(f"\n  {C.RED}Error testing {retailer_key}: {e}{C.END}")
            continue

    # ── Always save results (even on early exit) ──
    save_all_results(all_results)


def save_all_results(all_results):
    """Save whatever results we have to Excel + JSON."""
    if not all_results:
        print(f"\n  {C.YELLOW}No results to save.{C.END}")
        return

    print_summary(all_results)

    xlsx_path = os.path.join(SDIR, 'test_results.xlsx')
    try:
        save_test_results(all_results, xlsx_path)
        print(f"\n  {C.GREEN}Results saved to:{C.END} {xlsx_path}")
    except Exception as e:
        print(f"\n  {C.RED}Failed to save Excel: {e}{C.END}")

    json_path = os.path.join(SDIR, 'test_results.json')
    try:
        safe = {}
        for k, v in all_results.items():
            safe[k] = []
            for row in v:
                safe_row = {}
                for rk, rv in row.items():
                    if isinstance(rv, (str, int, float, bool, type(None))):
                        safe_row[rk] = rv
                    else:
                        safe_row[rk] = str(rv)
                safe[k].append(safe_row)
        with open(json_path, 'w') as f:
            json.dump(safe, f, indent=2, default=str)
        print(f"  {C.GREEN}JSON backup:{C.END} {json_path}")
    except Exception as e:
        print(f"  {C.YELLOW}JSON save failed: {e}{C.END}")


if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print(f"\n\n{C.YELLOW}Test interrupted.{C.END}")
    except SystemExit:
        pass
    except Exception as e:
        print(f"\n\n{C.RED}Fatal: {e}{C.END}")
        traceback.print_exc()
