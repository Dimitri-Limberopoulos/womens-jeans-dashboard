#!/usr/bin/env python3
"""
Fix Results Script
==================
1. Re-scrapes the handful of URLs that still have price ranges
2. Extracts color/wash from title for rows with missing color
3. Patches the Excel in place

Usage:
    python3 fix_results.py
"""

import asyncio, os, re, random
from datetime import datetime
from playwright.async_api import async_playwright
from target_pdp_scraper import (
    parse_target_pdp, extract_js_pricing, setup_context,
    save_to_excel, EXCEL_HEADERS, EXCEL_FIELDS,
)
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "target_pdp_results.xlsx")


# ── Color extraction from title ────────────────────────────────────────────

# Known color/wash words (case-insensitive matching)
COLOR_WORDS = {
    'black', 'white', 'blue', 'dark', 'medium', 'light', 'indigo', 'rinse',
    'grey', 'gray', 'charcoal', 'denim', 'wash', 'vintage', 'faded', 'stone',
    'sand', 'cream', 'ivory', 'khaki', 'olive', 'green', 'navy', 'red',
    'pink', 'brown', 'tan', 'beige', 'burgundy', 'maroon', 'coral',
    'distressed', 'acid', 'bleach', 'raw', 'deep', 'sky', 'steel',
}


def extract_color_from_title(title):
    """Extract color/wash from product title."""
    if not title:
        return ''

    # Pattern 1: "Jeans , Color , Size" or "Jeans - Color, Size"
    m = re.search(r'(?:Jeans?|Denim|Capri|Pants?)\s*[,\-]\s*([^,]+?)\s*,\s*\d', title)
    if m:
        candidate = m.group(1).strip()
        if len(candidate) < 40:
            return candidate

    # Pattern 2: "Jeans , Color , SizeP"  (petite)
    m = re.search(r'(?:Jeans?|Denim|Capri|Pants?)\s*,\s*([^,]+?)\s*,\s*\d+[PpWw]', title)
    if m:
        candidate = m.group(1).strip()
        if len(candidate) < 40:
            return candidate

    # Pattern 3: "Description - Color Size" at end
    m = re.search(r'[-–]\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2})\s+\d+[A-Z]*\s*$', title)
    if m:
        candidate = m.group(1).strip()
        skip = {'Women', 'Petite', 'Plus', 'Size', 'Length', 'Short', 'Long', 'Regular', 'Rise'}
        if candidate.split()[0] not in skip:
            return candidate

    # Pattern 4: Look for known color words after brand/product type
    # "... Jeans ColorWord ..."
    m = re.search(
        r'(?:Jeans?|Denim|Capri|Pants?)\s+'
        r'((?:(?:Dark|Medium|Light|Deep|Sky|Steel|Raw|Acid)\s+)?'
        r'(?:Black|White|Blue|Denim|Indigo|Grey|Gray|Charcoal|Navy|'
        r'Wash|Rinse|Stone|Sand|Cream|Olive|Green|Red|Pink|Brown|Tan|Khaki|Beige)(?:\s+(?:Denim|Wash))?)',
        title, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    return ''


async def rescrape_range_urls(range_urls):
    """Re-scrape URLs that have price ranges, using the updated scraper."""
    if not range_urls:
        return {}

    # Dedupe by base URL (strip preselect differences for same product)
    unique = list(set(range_urls))
    print(f"\n🔄 Re-scraping {len(unique)} URLs with price ranges...")

    pw = await async_playwright().start()
    browser = await pw.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
    ])
    ctx = await setup_context(browser, 0)
    page = await ctx.new_page()

    async def block_resources(route):
        await route.abort()
    await page.route('**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,ico,webp}', block_resources)

    # url -> list of color rows
    results = {}

    for i, url in enumerate(unique):
        print(f"\n  [{i+1}/{len(unique)}] {url[:70]}...")
        try:
            resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
            status = resp.status if resp else 0
            print(f"    HTTP {status}")

            if status >= 400:
                print(f"    ❌ HTTP error")
                continue

            await asyncio.sleep(random.uniform(2, 4))
            html = await page.content()
            js_pricing = await extract_js_pricing(page)
            print(f"    JS pricing: {len(js_pricing)} TCINs")

            color_rows = parse_target_pdp(html, url, js_pricing=js_pricing)

            # Check if ranges are resolved
            resolved = sum(1 for r in color_rows if ' - ' not in str(r.get('current_price', '')))
            print(f"    ✅ {len(color_rows)} rows, {resolved}/{len(color_rows)} prices resolved")

            for row in color_rows:
                row['url'] = url
                row['retries'] = 0

            results[url] = color_rows

        except Exception as e:
            print(f"    ❌ Error: {str(e)[:100]}")

        if i < len(unique) - 1:
            await asyncio.sleep(random.uniform(3, 5))

    await page.close()
    await ctx.close()
    await browser.close()
    await pw.stop()

    return results


def fix_excel():
    """Main fix: re-scrape ranges, extract colors from titles, patch Excel."""
    print("📊 Loading results Excel...")
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Build field index map
    idx = {h: i for i, h in enumerate(headers)}
    color_col = idx['Color'] + 1  # openpyxl is 1-indexed
    price_col = idx['Current Price'] + 1
    cprice_col = idx['Color Current Price'] + 1
    title_col = idx['Title'] + 1
    url_col = idx['URL'] + 1
    pmin_col = idx['Price Min'] + 1
    pmax_col = idx['Price Max'] + 1

    # ── Step 1: Find URLs with price ranges ──
    range_urls = set()
    for row_num in range(2, ws.max_row + 1):
        price = str(ws.cell(row=row_num, column=price_col).value or '')
        if ' - ' in price:
            range_urls.add(str(ws.cell(row=row_num, column=url_col).value or ''))

    # ── Step 2: Re-scrape range URLs ──
    rescrape_results = {}
    if range_urls:
        rescrape_results = asyncio.run(rescrape_range_urls(list(range_urls)))

    # Apply re-scraped results: replace rows for those URLs
    if rescrape_results:
        field_map = dict(zip(EXCEL_HEADERS, EXCEL_FIELDS))

        for row_num in range(ws.max_row, 1, -1):  # iterate backwards for safe deletion
            url = str(ws.cell(row=row_num, column=url_col).value or '')
            if url in rescrape_results:
                ws.delete_rows(row_num)

        # Append re-scraped rows
        for url, rows in rescrape_results.items():
            for row_data in rows:
                new_row = []
                for h, f in zip(EXCEL_HEADERS, EXCEL_FIELDS):
                    new_row.append(row_data.get(f, ''))
                ws.append(new_row)

        print(f"\n✅ Replaced rows for {len(rescrape_results)} re-scraped URLs")

    # ── Step 3: Fix remaining price ranges (use price_min) ──
    ranges_fixed = 0
    for row_num in range(2, ws.max_row + 1):
        price = str(ws.cell(row=row_num, column=price_col).value or '')
        if ' - ' in price:
            pmin = ws.cell(row=row_num, column=pmin_col).value
            if pmin:
                ws.cell(row=row_num, column=price_col, value=f"${float(pmin):.2f}")
                ws.cell(row=row_num, column=cprice_col, value=f"${float(pmin):.2f}")
                ranges_fixed += 1

    if ranges_fixed:
        print(f"✅ Fixed {ranges_fixed} remaining price ranges using Price Min")

    # ── Step 4: Extract color from title for missing colors ──
    colors_fixed = 0
    for row_num in range(2, ws.max_row + 1):
        color = ws.cell(row=row_num, column=color_col).value
        if not color or str(color).strip() == '':
            title = str(ws.cell(row=row_num, column=title_col).value or '')
            extracted = extract_color_from_title(title)
            if extracted:
                ws.cell(row=row_num, column=color_col, value=extracted)
                colors_fixed += 1

    print(f"✅ Extracted color from title for {colors_fixed} rows")

    # ── Step 5: Check remaining issues ──
    still_no_color = 0
    still_range = 0
    for row_num in range(2, ws.max_row + 1):
        color = ws.cell(row=row_num, column=color_col).value
        price = str(ws.cell(row=row_num, column=price_col).value or '')
        if not color or str(color).strip() == '':
            still_no_color += 1
        if ' - ' in price:
            still_range += 1

    print(f"\n📊 After fixes:")
    print(f"   Total rows: {ws.max_row - 1}")
    print(f"   Still missing color: {still_no_color}")
    print(f"   Still price ranges: {still_range}")

    # Save
    wb.save(XLSX_PATH)
    wb.close()
    print(f"\n💾 Saved to {XLSX_PATH}")


if __name__ == '__main__':
    fix_excel()
