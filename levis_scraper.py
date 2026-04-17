#!/usr/bin/env python3
"""
Levi's PDP Scraper — Playwright-based (1 row per color combo)
Navigates to each product-color URL, extracts structured data from:
  1. ld+json (ProductGroup) — name, color, SKU, price
  2. classifications JSON — wash, rise, leg, fit, stretch, material, closure
  3. Inline price JSON — current price, regular price, sale detection
  4. Color swatch links — discovers ALL color variants per product

Outputs: levis_pdp_results.xlsx  (1 row per product × color combo)
"""

import asyncio, csv, json, os, re, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── CONFIG ──────────────────────────────────────────────────────────
WD = os.path.dirname(os.path.abspath(__file__)) + '/'
URL_FILE       = WD + 'levis_all_color_urls.json'   # 636 color-variant URLs
URL_FILE_CSV   = WD + 'levi.csv'                     # fallback: original 243
OUTPUT_FILE    = WD + 'levis_pdp_results.xlsx'
PROGRESS_FILE  = WD + 'levis_progress.json'

DELAY_MIN    = 6.0        # was 2.5 — gentler polite-delay between page loads
DELAY_MAX    = 10.0       # was 4.5
NAV_TIMEOUT  = 30000      # ms


# ── EXTRACTION ──────────────────────────────────────────────────────
def extract_product_data(html, url):
    """Parse product data from raw HTML source."""
    result = {'url': url, 'error': None}

    try:
        # 1) ld+json blocks
        ld_blocks = re.findall(
            r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.DOTALL
        )
        for block in ld_blocks:
            try:
                parsed = json.loads(block)
                items = parsed if isinstance(parsed, list) else [parsed]
                for item in items:
                    if item.get('@type') == 'ProductGroup':
                        result['product_name'] = item.get('name', '')
                        result['productGroupID'] = item.get('productGroupID', '')
                        for v in (item.get('hasVariant') or []):
                            if isinstance(v, dict) and v.get('color'):
                                result['color'] = v['color']
                                result['sku'] = v.get('sku', '')
                                if v.get('offers'):
                                    result['ld_price'] = v['offers'].get('price')
                                break
                    if item.get('@type') == 'BreadcrumbList':
                        crumbs = [x.get('name', '') for x in (item.get('itemListElement') or [])]
                        result['breadcrumbs'] = ' > '.join(crumbs)
            except (json.JSONDecodeError, TypeError):
                pass

        # 2) classifications JSON — structured product specs
        idx = html.find('"classifications"')
        if idx > -1:
            start = html.index('[', idx)
            depth, end = 0, start
            for i in range(start, min(len(html), start + 5000)):
                if html[i] == '[': depth += 1
                if html[i] == ']': depth -= 1
                if depth == 0:
                    end = i + 1
                    break
            try:
                classif = json.loads(html[start:end])
                for group in classif:
                    for feat in (group.get('features') or []):
                        name = feat['code'].split('.')[-1]
                        vals = ', '.join(fv['value'] for fv in (feat.get('featureValues') or []))
                        result[f'spec_{name}'] = vals
            except (json.JSONDecodeError, ValueError):
                pass

        # 3) Prices from inline JSON
        val_m = re.search(r'"value"\s*:\s*([\d.]+)\s*,\s*"formattedValue"\s*:\s*"\$([\d.]+)"', html)
        reg_m = re.search(r'"regularPrice"\s*:\s*([\d.]+)', html)
        if val_m:
            result['current_price'] = float(val_m.group(1))
        if reg_m:
            result['original_price'] = float(reg_m.group(1))
        # Fallback to ld+json price
        if 'current_price' not in result and result.get('ld_price'):
            result['current_price'] = float(result['ld_price'])
        if 'original_price' not in result:
            result['original_price'] = result.get('current_price')

    except Exception as e:
        result['error'] = str(e)

    return result


def discover_color_urls(html, current_url):
    """Find all color swatch links on the page — discovers overflow colors."""
    url_prefix = current_url.rsplit('/p/', 1)[0] + '/p/' if '/p/' in current_url else ''
    if not url_prefix:
        return []

    discovered = set()

    # Pattern 1: <a href="...levi.com/.../p/CODE"> links (color swatches)
    for m in re.finditer(r'href=["\']([^"\']*?/p/\w+)["\']', html):
        href = m.group(1)
        # Must be same product path (same URL prefix before /p/)
        if url_prefix.replace('https://www.levi.com', '').replace('http://www.levi.com', '') in href.replace('https://www.levi.com', '').replace('http://www.levi.com', ''):
            # Normalize to full URL
            if href.startswith('http'):
                discovered.add(href)
            elif href.startswith('/'):
                discovered.add('https://www.levi.com' + href)

    # Pattern 2: swatch image URLs with color codes (WB_BASE-COLOR_GLO)
    base_code = ''
    bm = re.search(r'/p/(\w+)', current_url)
    if bm:
        full_code = bm.group(1)
        # Try to split into base + color (base is letters+digits, color is last 4 digits)
        cm = re.match(r'^(\w+?)(\d{4})$', full_code)
        if cm:
            base_code = cm.group(1)

    if base_code:
        for m in re.finditer(r'WB_' + re.escape(base_code) + r'-(\w+)_GLO', html):
            color_code = m.group(1)
            discovered.add(url_prefix + base_code + color_code)

    # Also catch lsco.scene7.com pattern
    if base_code:
        for m in re.finditer(r'/' + re.escape(base_code) + r'(\w+)-swatch', html):
            color_code = m.group(1)
            discovered.add(url_prefix + base_code + color_code)

    return list(discovered)


# ── LOAD URLS ───────────────────────────────────────────────────────
def load_urls():
    # Prefer the expanded color-variant list
    if os.path.exists(URL_FILE):
        with open(URL_FILE) as f:
            urls = json.load(f)
        print(f'Loaded {len(urls)} color-variant URLs from {os.path.basename(URL_FILE)}')
        return urls

    # Fallback: original CSV (1 URL per product, not per color)
    urls = []
    with open(URL_FILE_CSV, newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            urls.append(row['cell-image-link href'])
    print(f'Loaded {len(urls)} URLs from {os.path.basename(URL_FILE_CSV)} (1 per product)')
    return urls


# ── PROGRESS / RESUME ──────────────────────────────────────────────
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {'done': {}, 'results': [], 'discovered': []}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f)


# ── EXCEL OUTPUT ────────────────────────────────────────────────────
COLUMNS = [
    'url', 'product_name', 'color', 'sku', 'current_price', 'original_price',
    'spec_wash', 'spec_rise', 'spec_leg_opening', 'spec_fit', 'spec_fit_name',
    'spec_stretch', 'spec_material', 'spec_materialtype', 'spec_closure',
    'spec_fit_text', 'spec_color_family', 'spec_size_group', 'spec_gender',
    'breadcrumbs', 'productGroupID', 'error'
]

def write_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Levis PDP Data'

    # Header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_idx, rec in enumerate(results, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(col_name, ''))

    # Column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    for c in 'GHIJKLMNOPQRST':
        ws.column_dimensions[c].width = 20

    ws.freeze_panes = 'A2'
    wb.save(OUTPUT_FILE)
    print(f'  ✅ Saved {len(results)} rows → {OUTPUT_FILE}')


# ── MAIN SCRAPER ────────────────────────────────────────────────────
async def scrape_all():
    from playwright.async_api import async_playwright

    urls = load_urls()
    progress = load_progress()
    print(f'Already scraped: {len(progress["done"])}')

    # Merge in any dynamically discovered URLs
    all_urls = list(dict.fromkeys(urls + progress.get('discovered', [])))
    remaining = [u for u in all_urls if u not in progress['done']]

    if not remaining:
        print('All URLs already scraped! Writing Excel...')
        good = [r for r in progress['results'] if not r.get('error')]
        write_excel(good)
        return

    print(f'Total URLs: {len(all_urls)}, Remaining: {len(remaining)}')
    print(f'Estimated time: ~{len(remaining) * 3.5 / 60:.0f} minutes\n')

    async with async_playwright() as p:
        # ── Use Firefox (better anti-detection than Chromium) ──
        # Persistent context keeps cookies between runs
        profile_dir = WD + '.levis_browser_profile'
        os.makedirs(profile_dir, exist_ok=True)

        context = await p.firefox.launch_persistent_context(
            profile_dir,
            headless=False,
            viewport={'width': 1280, 'height': 800},
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:128.0) '
                       'Gecko/20100101 Firefox/128.0',
        )
        page = context.pages[0] if context.pages else await context.new_page()

        # ── Manual warm-up: let user solve Cloudflare challenge ──
        print('=' * 60)
        print('STEP 1: Opening levi.com — if you see a Cloudflare')
        print('challenge (checkbox, puzzle, "checking your browser"),')
        print('solve it in the Firefox window that just opened.')
        print('=' * 60)
        try:
            await page.goto('https://www.levi.com/US/en_US',
                            wait_until='domcontentloaded', timeout=60000)
        except Exception as e:
            print(f'  Homepage load warning: {e}')

        # Wait for user to solve any challenge
        input('\n👉  Press ENTER here once levi.com has loaded '
              'normally in the browser window...\n')

        # Verify we can access the site
        print('Verifying access...')
        try:
            test_resp = await page.goto(remaining[0],
                                        wait_until='domcontentloaded',
                                        timeout=NAV_TIMEOUT)
            if test_resp and test_resp.status == 403:
                print('\n❌ Still getting 403 blocked.')
                print('   Try one of these:')
                print('   • Wait 15-30 min and try again')
                print('   • Connect to a different WiFi / phone hotspot')
                print('   • In the Firefox window, navigate to levi.com')
                print('     manually and browse around for a minute')
                input('\n👉  Press ENTER to retry, or Ctrl+C to quit...\n')
            else:
                # Put the first URL back — we'll extract from it in the loop
                await page.goto('https://www.levi.com/US/en_US',
                                wait_until='domcontentloaded', timeout=NAV_TIMEOUT)
                await asyncio.sleep(1)
                print('✅ Access confirmed! Starting scrape...\n')
        except Exception as e:
            print(f'  Verification warning: {e}')
            input('\n👉  Press ENTER to continue anyway...\n')

        errors = 0
        consecutive_403 = 0
        discovered_new = 0

        i = 0
        while i < len(remaining):
            url = remaining[i]

            # Respect rate limits
            delay = random.uniform(DELAY_MIN, DELAY_MAX)
            if consecutive_403 > 0:
                delay = min(90, delay * (2 ** consecutive_403))   # was 30 — allow longer backoff
                print(f'  ⏳ Backing off {delay:.1f}s after {consecutive_403} consecutive 403s')

            if i > 0:
                await asyncio.sleep(delay)

            sku = url.split('/p/')[-1] if '/p/' in url else url[-12:]
            print(f'[{i+1}/{len(remaining)}] {sku}', end=' ... ', flush=True)

            try:
                resp = await page.goto(url, wait_until='domcontentloaded',
                                       timeout=NAV_TIMEOUT)

                if resp and resp.status == 403:
                    print('❌ 403 blocked')
                    consecutive_403 += 1
                    # NOTE: don't mark as done — let a subsequent run retry these URLs
                    if consecutive_403 >= 3:                       # was 5 — trigger pause earlier
                        print(f'\n⚠️  3 consecutive 403s — pausing 120s then aborting this run...')
                        await asyncio.sleep(120)                   # was 60
                        consecutive_403 = 0
                        break                                      # stop this run — let IP cool off
                    i += 1
                    continue

                consecutive_403 = 0
                await asyncio.sleep(1.5)

                # Get page source
                html = await page.content()

                # Extract data
                result = extract_product_data(html, url)

                # Discover additional color URLs from this page
                new_urls = discover_color_urls(html, url)
                for nu in new_urls:
                    if nu not in progress['done'] and nu not in remaining:
                        remaining.append(nu)
                        progress.setdefault('discovered', []).append(nu)
                        discovered_new += 1

                if result.get('product_name'):
                    price_str = f"${result.get('current_price', '?')}"
                    if result.get('original_price') and \
                       result['original_price'] != result.get('current_price'):
                        price_str += f" (was ${result['original_price']})"
                    extra = f' [+{len(new_urls)} colors]' if new_urls else ''
                    print(f'✅ {result["product_name"][:30]} | '
                          f'{result.get("color","?")[:22]} | {price_str}{extra}')
                else:
                    print('⚠️  No product data found')
                    result['error'] = result.get('error') or 'No product data in HTML'
                    errors += 1

                progress['done'][url] = True
                progress['results'].append(result)

            except Exception as e:
                print(f'❌ {str(e)[:60]}')
                progress['done'][url] = True
                progress['results'].append({'url': url, 'error': str(e)[:200]})
                errors += 1

            # Save progress every 10 URLs
            if (i + 1) % 10 == 0:
                save_progress(progress)
                good_so_far = sum(1 for r in progress['results'] if not r.get('error'))
                print(f'  💾 Progress: {good_so_far} good, '
                      f'{len(progress["results"]) - good_so_far} errors, '
                      f'{discovered_new} new colors discovered')

            i += 1

        # Final save
        save_progress(progress)
        await context.close()

    # Write Excel (only good results)
    good_results = [r for r in progress['results'] if not r.get('error')]
    all_results = progress['results']
    print(f'\n{"="*60}')
    print(f'DONE!')
    print(f'  Successful: {len(good_results)}')
    print(f'  Errors:     {len(all_results) - len(good_results)}')
    print(f'  New colors discovered: {discovered_new}')
    print(f'{"="*60}')
    write_excel(good_results)


if __name__ == '__main__':
    asyncio.run(scrape_all())
