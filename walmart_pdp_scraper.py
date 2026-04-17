#!/usr/bin/env python3
"""
Production Walmart PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  walmart.csv  (URLs in first column 'w-100 href')
Output: walmart_pdp_results.xlsx + walmart_pdp_progress.json

Architecture: 5 browser contexts × 2 tabs each = 10 parallel workers
              30-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

Extracts from <script id="__NEXT_DATA__"> embedded JSON to produce
one row per color variant with pricing, specs, and availability.
"""

import asyncio, json, os, random, re, time, csv, html as html_mod
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 5
TABS_PER_CTX = 1
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 5 total workers (balanced for Walmart)
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 20 * 60  # 20 minutes — restart browser with fresh session (shorter to avoid CAPTCHA buildup)


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# ── "Robot or human?" Popup Dismissal ──────────────────────────────────────

async def _dismiss_robot_popup(page):
    """
    Dismiss the "Robot or human?" popup that appears after CAPTCHA is solved.
    It has an X (close) button. This is NOT a CAPTCHA — it's a confirmation dialog.
    """
    try:
        # Look for the popup's close button — it's typically an X/close icon
        # The popup text says "Activate and hold the button to confirm that you're human"
        close_selectors = [
            'button[aria-label="Close"]',
            'button[aria-label="close"]',
            '[class*="modal"] button[class*="close"]',
            '[class*="dialog"] button[class*="close"]',
            '[role="dialog"] button',
            'button:has(svg[class*="close"])',
        ]
        for sel in close_selectors:
            try:
                btn = await page.query_selector(sel)
                if btn:
                    is_visible = await btn.is_visible()
                    if is_visible:
                        await btn.click()
                        print(f"    ✓ Dismissed 'Robot or human?' popup via {sel}")
                        await asyncio.sleep(random.uniform(0.5, 1.0))
                        return True
            except Exception:
                continue

        # Fallback: look for any visible element containing "Robot or human" and find nearby close button
        robot_el = await page.query_selector('text="Robot or human?"')
        if robot_el:
            # Find the closest ancestor dialog/modal and click its close button
            close_btn = await page.evaluate("""() => {
                var el = document.querySelector('[class*="modal"], [role="dialog"], [class*="popup"]');
                if (!el) return false;
                var btn = el.querySelector('button, [role="button"], [aria-label*="close"], [aria-label*="Close"]');
                if (btn) { btn.click(); return true; }
                // Try the X button specifically
                var svg = el.querySelector('svg');
                if (svg && svg.closest('button')) { svg.closest('button').click(); return true; }
                return false;
            }""")
            if close_btn:
                print(f"    ✓ Dismissed 'Robot or human?' popup via JS fallback")
                await asyncio.sleep(random.uniform(0.5, 1.0))
                return True
    except Exception:
        pass
    return False


# ── PerimeterX Press-and-Hold CAPTCHA Solver ───────────────────────────────

async def solve_press_and_hold(page, max_attempts=3):
    """
    Solve Walmart's PerimeterX press-and-hold CAPTCHA.
    Returns True if solved, False if failed.

    PerimeterX renders a #px-captcha element (or iframe) with a button
    that must be held down for ~5-10 seconds. We simulate a real mouse
    press-and-hold with slight jitter to appear human.
    """
    for attempt in range(1, max_attempts + 1):
        print(f"    🔒 Attempting CAPTCHA solve (attempt {attempt}/{max_attempts})...")

        try:
            # Wait for the CAPTCHA element to appear
            captcha_el = None

            # Method 1: Direct #px-captcha element
            try:
                captcha_el = await page.wait_for_selector('#px-captcha', timeout=5000)
            except Exception:
                pass

            # Method 2: Look for the press-and-hold button inside an iframe
            if not captcha_el:
                try:
                    for frame in page.frames:
                        try:
                            captcha_el = await frame.wait_for_selector(
                                '#px-captcha, [id*="captcha"], button[class*="hold"], [data-testid*="captcha"]',
                                timeout=3000
                            )
                            if captcha_el:
                                break
                        except Exception:
                            continue
                except Exception:
                    pass

            # Method 3: Any element containing "press" or "hold" text
            if not captcha_el:
                try:
                    captcha_el = await page.query_selector('[id*="captcha"], [class*="captcha"]')
                except Exception:
                    pass

            if not captcha_el:
                print(f"    ⚠️  No CAPTCHA element found — page might have refreshed")
                # Check if CAPTCHA is gone (page loaded normally)
                content = await page.content()
                if '__NEXT_DATA__' in content:
                    print(f"    ✓ Page loaded successfully — no CAPTCHA to solve")
                    return True
                await asyncio.sleep(2)
                continue

            # Get the element's bounding box for mouse positioning
            box = await captcha_el.bounding_box()
            if not box:
                print(f"    ⚠️  CAPTCHA element has no bounding box")
                await asyncio.sleep(2)
                continue

            # Calculate click position — center with slight random offset
            x = box['x'] + box['width'] / 2 + random.uniform(-10, 10)
            y = box['y'] + box['height'] / 2 + random.uniform(-5, 5)

            # Move mouse to position naturally (not instant teleport)
            await page.mouse.move(x - random.uniform(50, 150), y - random.uniform(20, 50))
            await asyncio.sleep(random.uniform(0.2, 0.5))
            await page.mouse.move(x, y)
            await asyncio.sleep(random.uniform(0.1, 0.3))

            # Press and hold — PerimeterX requires a long hold (25-35s to be safe)
            hold_time = random.uniform(25.0, 35.0)
            print(f"    🖱️  Pressing and holding for {hold_time:.1f}s...")

            await page.mouse.down()

            # Hold with micro-movements (humans don't hold perfectly still)
            elapsed = 0
            while elapsed < hold_time:
                jitter_x = x + random.uniform(-2, 2)
                jitter_y = y + random.uniform(-2, 2)
                await page.mouse.move(jitter_x, jitter_y)
                sleep_chunk = random.uniform(0.3, 0.8)
                await asyncio.sleep(sleep_chunk)
                elapsed += sleep_chunk

            await page.mouse.up()
            print(f"    🖱️  Released after {elapsed:.1f}s")

            # Wait for page to process the CAPTCHA response
            await asyncio.sleep(random.uniform(2.0, 4.0))

            # Check if CAPTCHA is solved — page should now have product data
            try:
                await page.wait_for_load_state('domcontentloaded', timeout=10000)
            except Exception:
                pass

            content = await page.content()

            # Dismiss "Robot or human?" popup if it appears after CAPTCHA solve
            await _dismiss_robot_popup(page)

            if '__NEXT_DATA__' in content:
                print(f"    ✅ CAPTCHA solved! Product data loaded.")
                return True
            elif 'press & hold' in content.lower() or 'press and hold' in content.lower():
                print(f"    ❌ CAPTCHA still present — retrying...")
                await asyncio.sleep(random.uniform(3, 6))
                continue
            else:
                # Might need to reload the original URL
                print(f"    ⏳ Page changed but no product data — checking...")
                await asyncio.sleep(2)
                await _dismiss_robot_popup(page)
                content = await page.content()
                if '__NEXT_DATA__' in content:
                    print(f"    ✅ CAPTCHA solved after reload!")
                    return True

        except Exception as e:
            print(f"    ⚠️  CAPTCHA attempt {attempt} error: {str(e)[:100]}")
            await asyncio.sleep(random.uniform(2, 5))

    print(f"    ❌ CAPTCHA not solved after {max_attempts} attempts")
    return False


# ── Walmart Price Extraction JS ────────────────────────────────────────────
WALMART_PRICE_JS = """() => {
    // Look for price in the rendered DOM after swatch click
    var el = document.querySelector('[itemprop="price"]') ||
             document.querySelector('[data-testid="price"] span') ||
             document.querySelector('.price-characteristic');
    if (el) {
        var text = el.textContent.replace(/[^0-9.]/g, '');
        return parseFloat(text) || null;
    }
    // Fallback: find any visible price
    var priceEls = document.querySelectorAll('[class*="price"], [data-automation-id*="price"]');
    for (var i = 0; i < priceEls.length; i++) {
        var t = priceEls[i].textContent.trim();
        var m = t.match(/\\$(\\d+\\.?\\d*)/);
        if (m) return parseFloat(m[1]);
    }
    return null;
}"""


# ── Walmart Page Text Extraction JS ────────────────────────────────────────
WALMART_PAGE_TEXT_JS = """() => {
    // Extract up to 10000 chars of full page text for keyword searching
    var text = document.body.innerText || '';
    return text.substring(0, 10000);
}"""


# ── Walmart Owned Brands ────────────────────────────────────────────────────
WALMART_OWNED_BRANDS = {
    'time and tru', 'terra & sky', 'no boundaries', 'free assembly',
}


def classify_brand(brand_name):
    """Returns 'OB' if Walmart owned brand, 'NB' for national brand."""
    if brand_name and brand_name.strip().lower() in WALMART_OWNED_BRANDS:
        return 'OB'
    return 'NB'


# ── Walmart Data Extraction ─────────────────────────────────────────────────

def strip_html(text):
    """Remove HTML tags from text."""
    if not text:
        return ''
    return re.sub(r'<[^>]+>', '', text)


def extract_next_data(page_html):
    """
    Extract __NEXT_DATA__ JSON from the page.
    NOTE: There are multiple __NEXT_DATA__ occurrences. The actual data is in
    the <script id="__NEXT_DATA__" type="application/json"> tag.
    Returns dict with structure: {product, idml, reviews} or None on failure.
    """
    # Use specific regex to find the actual <script> tag (not JavaScript references)
    match = re.search(
        r'<script\s+id=["\']__NEXT_DATA__["\'][^>]*type=["\']application/json["\'][^>]*>(.*?)</script>',
        page_html,
        re.DOTALL
    )
    if not match:
        # Try alternate order
        match = re.search(
            r'<script\s+type=["\']application/json["\'][^>]*id=["\']__NEXT_DATA__["\'][^>]*>(.*?)</script>',
            page_html,
            re.DOTALL
        )
    if not match:
        return None

    try:
        json_str = match.group(1).strip()
        data = json.loads(json_str)
        # Navigate the structure: props.pageProps.initialData.data
        return data.get('props', {}).get('pageProps', {}).get('initialData', {}).get('data', {})
    except (json.JSONDecodeError, KeyError, AttributeError):
        return None


def get_variant_dimension(variant_criteria, dimension_keywords):
    """
    Generic helper: extract variants from a specific dimension of variantCriteria.
    dimension_keywords: list of lowercase keywords to match dimension name (e.g., ['color', 'actual color']).
    Returns list of variant dicts from the matching dimension, or empty list.
    """
    if not isinstance(variant_criteria, list):
        return []
    for vc in variant_criteria:
        vc_name = vc.get('name', '').lower()
        for kw in dimension_keywords:
            if kw in vc_name:
                return vc.get('variantList', [])
    return []


def get_variant_colors(variant_criteria):
    """Extract color variants. Returns list of {name, swatchImageUrl, availabilityStatus, products}."""
    raw = get_variant_dimension(variant_criteria, ['color', 'actual color'])
    colors = []
    for v in raw:
        colors.append({
            'id': v.get('id', ''),                # e.g., "actual_color-blackrinse"
            'name': v.get('name', ''),
            'swatchImageUrl': v.get('swatchImageUrl', ''),
            'availabilityStatus': v.get('availabilityStatus', ''),
            'products': v.get('products', []),     # list of usItemIds for this color
            'selected': v.get('selected', False),
        })
    return colors


def get_variant_sizes(variant_criteria):
    """Extract size variant names."""
    raw = get_variant_dimension(variant_criteria, ['size'])
    # Exclude dimensions that also contain 'color' or 'length' or 'inseam'
    # But since get_variant_dimension returns the first match, we refine:
    if not isinstance(variant_criteria, list):
        return []
    sizes = []
    for vc in variant_criteria:
        vc_name = vc.get('name', '').lower()
        if 'size' in vc_name and 'color' not in vc_name and 'length' not in vc_name and 'inseam' not in vc_name:
            for v in vc.get('variantList', []):
                sizes.append(v.get('name', ''))
    return sizes


def get_variant_lengths(variant_criteria):
    """
    Extract pant leg length variants (Short, Full/Regular, Long) from variantCriteria.
    Returns list of variant dicts with {name, products, selected, ...}.
    """
    if not isinstance(variant_criteria, list):
        return []
    for vc in variant_criteria:
        vc_name = vc.get('name', '').lower()
        if 'length' in vc_name or 'inseam' in vc_name or 'pant leg' in vc_name:
            results = []
            for v in vc.get('variantList', []):
                results.append({
                    'id': v.get('id', ''),           # e.g., "pant_leg_length-short"
                    'name': v.get('name', ''),
                    'products': v.get('products', []),
                    'selected': v.get('selected', False),
                })
            return results
    return []


def detect_pack_size(name):
    """
    Detect pack size from product or variant name.
    Returns integer (1 for single items, 2 for 2-packs, etc.).
    Patterns: "2-Pack", "2 Pack", "3-Pack", "(2 pk)", etc.
    Also detects multi-color names like "Black & Medium Wash" or "Black/Dark Wash"
    as 2-pack indicators.
    """
    if not name:
        return 1
    # Explicit pack patterns: "2-Pack", "2 Pack", "3-Pack", "(2 pk)", etc.
    m = re.search(r'(\d+)\s*[-\s]?\s*(?:pack|pk|pck)', name, re.IGNORECASE)
    if m:
        return int(m.group(1))

    # Multi-color separators: "&", "and", "/" between color names
    # e.g., "Black & Medium Wash" → 2-pack
    # e.g., "Black/Dark Wash" → 2-pack
    # e.g., "Dark Wash/Light Wash" → 2-pack
    # Exclude product-name-level slashes (only split on variant/color names)
    parts = re.split(r'\s*[/&]\s*|\s+and\s+', name)
    if len(parts) >= 2:
        # Heuristic: each part should be short (< 25 chars) — looks like color names, not long product descriptions
        all_short = all(len(p.strip()) < 25 for p in parts)
        if all_short:
            return len(parts)
    return 1


def build_variant_price_map(data):
    """
    Build pricing lookups from __NEXT_DATA__ using Walmart's actual structure:

    1. variantsMap: {productId: {variants: ["actual_color-black", "pant_leg_length-short", ...],
                                  priceInfo: {currentPrice: {price: 14.98}, wasPrice: ...}}}
       → This is the primary source: every color × size × length combo has its own entry.

    2. variantProductIdMap: {"pant_leg_length-short-actual_color-black-clothing_size-8": "PRODUCT_ID"}
       → Maps dimension combos to product IDs (used to cross-reference variantsMap).

    Returns two dicts:
      variant_prices: {productId: {current_price, original_price, ...}}
      combo_prices:   {(color_slug, length_slug): {current_price, original_price, ...}}
         where color_slug = "blackrinse", length_slug = "short"/"regular"/"long"
    """
    variant_prices = {}   # productId → price info
    combo_prices = {}     # (color_slug, length_slug) → price info (deduplicated across sizes)
    product = data.get('product', {})

    # ── Primary: variantsMap (the real gold mine) ──
    variants_map = product.get('variantsMap', {})
    for pid, entry in variants_map.items():
        if not isinstance(entry, dict):
            continue
        pi = entry.get('priceInfo', {})
        if not pi:
            continue
        cp = pi.get('currentPrice', {}) or {}
        wp = pi.get('wasPrice', {}) or {}
        price_info = {
            'current_price': cp.get('price', 0),
            'current_price_formatted': cp.get('priceString', ''),
            'original_price': wp.get('price') if wp else None,
            'original_price_formatted': wp.get('priceString', '') if wp else '',
        }
        variant_prices[pid] = price_info

        # Parse dimension slugs from the variants array
        dims = entry.get('variants', [])
        color_slug = ''
        length_slug = ''
        for d in dims:
            if isinstance(d, str):
                if d.startswith('actual_color-'):
                    color_slug = d.replace('actual_color-', '')
                elif d.startswith('pant_leg_length-'):
                    length_slug = d.replace('pant_leg_length-', '')
        if color_slug and length_slug and price_info['current_price']:
            key = (color_slug, length_slug)
            # Collect ALL prices for this combo — we'll pick the mode (most common) later
            # because prices can vary by size (e.g., clearance on larger sizes)
            if key not in combo_prices:
                combo_prices[key] = []
            combo_prices[key].append(price_info)

    # ── Fallback: variantProducts (older Walmart format) ──
    if not variant_prices:
        for vp in product.get('variantProducts', []):
            uid = vp.get('usItemId', '')
            if not uid:
                continue
            pi = vp.get('priceInfo', {})
            cp = pi.get('currentPrice', {}) or {}
            wp = pi.get('wasPrice', {}) or {}
            variant_prices[uid] = {
                'current_price': cp.get('price', 0),
                'current_price_formatted': cp.get('priceString', ''),
                'original_price': wp.get('price') if wp else None,
                'original_price_formatted': wp.get('priceString', '') if wp else '',
            }

    # Post-process combo_prices: pick the MODE (most common) price for each color×length
    # This avoids clearance/markdown prices on specific sizes skewing the result
    from collections import Counter
    final_combo_prices = {}
    for key, price_list in combo_prices.items():
        price_counts = Counter(p['current_price'] for p in price_list if p['current_price'])
        if price_counts:
            mode_price = price_counts.most_common(1)[0][0]
            # Find the full price_info dict matching the mode price
            for p in price_list:
                if p['current_price'] == mode_price:
                    final_combo_prices[key] = p
                    break

    return variant_prices, final_combo_prices


def normalize_length_name(name):
    """Normalize length variant names to Short/Full/Long buckets."""
    n = name.strip().lower()
    if 'short' in n:
        return 'Short'
    elif 'long' in n:
        return 'Long'
    elif any(x in n for x in ['full', 'regular', 'medium', 'standard']):
        return 'Full'
    return name.strip()


def get_spec_value(specs, keys):
    """
    Get value from specs array by looking for any key in the keys list.
    specs: list of {name, value} dicts
    keys: list of possible key names to match
    Returns value string or empty string.
    """
    if not isinstance(specs, list):
        return ''
    for spec in specs:
        spec_name = spec.get('name', '').lower()
        for key in keys:
            if spec_name == key.lower():
                return spec.get('value', '')
    return ''


def extract_product_details(page_html):
    """
    Extract 'Product details' or 'Product information' section from page HTML.
    Looks for bullet-list sections and captures text.
    Returns a concatenated string of product details.
    """
    details = []
    # Look for product details section — often wrapped in divs with specific classes
    # Try multiple patterns
    patterns = [
        r'(?i)product\s+details.*?(?=(?:size|fit|specifications|reviews|customer|ratings|about|similar|$))',
        r'(?i)product\s+information.*?(?=(?:size|fit|specifications|reviews|customer|ratings|about|similar|$))',
        r'(?i)details.*?(?=(?:size|fit|specifications|reviews|customer|ratings|about|similar|$))',
    ]

    # Try to find bullet point lists
    bullets = re.findall(r'(?:^|\n)\s*[•\-\*]\s+(.+?)(?=\n|$)', page_html, re.MULTILINE | re.IGNORECASE)
    if bullets:
        details.extend(bullets[:20])  # Cap at 20 bullet points

    return ' '.join(details[:500]) if details else ''  # Cap at 500 chars total


def extract_size_and_fit(page_html):
    """
    Extract 'Size & fit' or 'Sizing' section from page HTML.
    Looks for text between Size/Fit headings and other sections.
    Returns a concatenated string of size and fit information.
    """
    # Look for size and fit section — often contains fit type, rise, inseam, etc.
    patterns = [
        r'(?i)size\s*[&and]\s*fit.*?(?=(?:product|details|specifications|reviews|customer|ratings|about|similar|$))',
        r'(?i)sizing.*?(?=(?:product|details|specifications|reviews|customer|ratings|about|similar|$))',
        r'(?i)fit(?:\s+details)?.*?(?=(?:product|details|specifications|reviews|customer|ratings|about|similar|$))',
    ]

    fit_text = []
    for pattern in patterns:
        match = re.search(pattern, page_html, re.DOTALL | re.IGNORECASE)
        if match:
            section = match.group(0)
            # Extract sentences and bullet points
            lines = re.findall(r'(?:^|\n)\s*(?:[•\-\*]|[0-9]+\.)\s+(.+?)(?=\n|$)', section, re.MULTILINE)
            if lines:
                fit_text.extend(lines[:15])  # Cap at 15 lines
                break

    return ' '.join(fit_text[:500]) if fit_text else ''  # Cap at 500 chars


def parse_length_hit(page_html, inseam_value):
    """
    Parse 'length hit' from the page — where on the leg the pants hit (e.g., 'hits at ankle').
    Looks for common phrases in size/fit descriptions.
    """
    hit_keywords = {
        'ankle': 'hits at ankle',
        'calf': 'hits at calf',
        'knee': 'hits at knee',
        'thigh': 'hits at thigh',
        'hip': 'hits at hip',
        'waist': 'hits at waist',
        'floor': 'hits at floor',
        'inseam': 'hits at inseam',
    }

    html_lower = page_html.lower()
    for keyword, label in hit_keywords.items():
        if keyword in html_lower:
            return label

    # Fallback: infer from inseam value if numeric
    if inseam_value:
        try:
            inseam_num = float(re.search(r'(\d+(?:\.\d+)?)', str(inseam_value)).group(1))
            if inseam_num < 28:
                return 'hits at ankle'
            elif 28 <= inseam_num < 32:
                return 'hits at calf'
            else:
                return 'hits at floor'
        except:
            pass

    return ''


def parse_walmart_pdp(page_html, url, dump_json=False):
    """
    Parse a Walmart PDP page and return a LIST of result dicts — one per color.
    If product has no color variants, returns single row with product-level color.

    New fields:
    - pack_size: integer (1, 2, 3...) detected from product/variant name
    - price_short, price_full, price_long: per-length pricing when available
    - Per-color pricing from variant price map when available
    """
    rows = []

    try:
        # Extract __NEXT_DATA__
        data = extract_next_data(page_html)
        if not data:
            return [{
                'url': url,
                'error': 'No __NEXT_DATA__ found',
                'timestamp': datetime.now().isoformat(),
            }]

        # Optionally dump full JSON for debugging
        if dump_json:
            try:
                sdir = os.path.dirname(os.path.abspath(__file__))
                dump_path = os.path.join(sdir, 'walmart_debug_json.json')
                with open(dump_path, 'w') as f:
                    json.dump(data, f, indent=2, default=str)
                print(f"    DEBUG: Full __NEXT_DATA__ dumped to {dump_path}")
                # Also dump just the keys at each level for quick inspection
                product_keys = list(data.get('product', {}).keys())
                print(f"    DEBUG: product keys = {product_keys}")
                vc = data.get('product', {}).get('variantCriteria', [])
                for dim in vc:
                    dim_name = dim.get('name', '?')
                    vlist = dim.get('variantList', [])
                    sample_keys = list(vlist[0].keys()) if vlist else []
                    print(f"    DEBUG: variantCriteria '{dim_name}' → {len(vlist)} variants, keys={sample_keys}")
                    if vlist:
                        print(f"    DEBUG:   sample variant: {json.dumps(vlist[0], default=str)[:300]}")
            except Exception as de:
                print(f"    DEBUG dump error: {de}")

        product = data.get('product', {})
        idml = data.get('idml', {})

        # Shared product-level data
        product_name = product.get('name', '')
        brand = product.get('brand', '')
        gender = product.get('gender', '')

        # Category breadcrumbs
        category_path = ''
        if 'category' in product and 'path' in product['category']:
            cat_names = [c.get('name', '') for c in product['category']['path']]
            category_path = ' > '.join(c for c in cat_names if c)

        # Product-level pricing (fallback if no per-variant pricing)
        price_info = product.get('priceInfo', {})
        current_price_obj = price_info.get('currentPrice', {})
        product_current_price = current_price_obj.get('price', 0)
        product_current_price_fmt = current_price_obj.get('priceString', '')

        was_price_obj = price_info.get('wasPrice', {})
        product_original_price = None
        product_original_price_fmt = ''
        if was_price_obj:
            product_original_price = was_price_obj.get('price')
            product_original_price_fmt = was_price_obj.get('priceString', '')
        if not product_original_price:
            product_original_price = product_current_price
            product_original_price_fmt = product_current_price_fmt

        # Ratings
        average_rating = product.get('averageRating', '')
        num_reviews = product.get('numberOfReviews', '')

        # Images
        image_info = product.get('imageInfo', {})
        all_images = image_info.get('allImages', [])
        primary_image = ''
        if all_images:
            img = all_images[0]
            if isinstance(img, dict):
                primary_image = img.get('url', '')
            elif isinstance(img, str):
                primary_image = img

        # Specs from idml
        specifications = idml.get('specifications', [])
        long_description = idml.get('longDescription', '')
        long_description_clean = strip_html(long_description)

        pant_rise = get_spec_value(specifications, ['pant rise'])
        pant_leg_cut = get_spec_value(specifications, ['pant leg cut'])
        clothing_fit = get_spec_value(specifications, ['clothing fit'])
        pant_style = get_spec_value(specifications, ['pant style'])
        jean_wash = get_spec_value(specifications, ['jean wash'])
        inseam = get_spec_value(specifications, ['inseam'])
        pant_leg_length = get_spec_value(specifications, ['pant leg length'])
        fabric_material = get_spec_value(specifications, ['fabric material name'])
        fabric_pct = get_spec_value(specifications, ['fabric material percentage'])
        clothing_size_group = get_spec_value(specifications, ['clothing size group'])
        clothing_occasion = get_spec_value(specifications, ['clothing occasion'])
        fastener_type = get_spec_value(specifications, ['fastener type'])
        fabric_care = get_spec_value(specifications, ['fabric care instructions'])
        age_group = get_spec_value(specifications, ['age group'])

        # Extract new fields from page HTML
        product_details = extract_product_details(page_html)
        size_and_fit = extract_size_and_fit(page_html)
        length_hit = parse_length_hit(page_html, inseam)
        # page_text will be captured later via JS during browser interaction
        page_text = ''  # placeholder; will be populated in scrape_walmart_pdp after JS evaluation

        # ── Variant dimensions ──
        variant_criteria = product.get('variantCriteria', [])
        color_variants = get_variant_colors(variant_criteria)
        size_variants = get_variant_sizes(variant_criteria)
        length_variant_objs = get_variant_lengths(variant_criteria)

        total_colors = len(color_variants) if color_variants else 1
        total_sizes = len(size_variants)
        length_names = [lv['name'] for lv in length_variant_objs]
        total_lengths = len(length_names)
        sizes_list = ', '.join(size_variants) if size_variants else ''
        lengths_list = ', '.join(length_names) if length_names else ''

        # ── Build per-variant price map from variantsMap ──
        variant_prices, combo_prices = build_variant_price_map(data)
        has_variant_prices = len(variant_prices) > 0
        if has_variant_prices:
            print(f"    Found {len(variant_prices)} variant prices, {len(combo_prices)} color×length combos in JSON")

        # ── Build color slug lookup: color_name_lower → color_slug ──
        # e.g., "black rinse" → "blackrinse" (from variantCriteria id field)
        color_slug_map = {}  # {color_name_lower: slug}
        for cv in color_variants:
            color_id = cv.get('id', '')  # e.g., "actual_color-blackrinse"
            if isinstance(color_id, str) and color_id.startswith('actual_color-'):
                slug = color_id.replace('actual_color-', '')
            else:
                # Fallback: lowercase and remove spaces
                slug = cv.get('name', '').lower().replace(' ', '')
            color_slug_map[cv.get('name', '').lower()] = slug

        # ── Build length slug lookup: normalized_length → length_slug ──
        length_slug_map = {}  # {"Short": "short", "Full": "regular", ...}
        for lv in length_variant_objs:
            lid = lv.get('id', '')
            if isinstance(lid, str) and 'pant_leg_length-' in lid:
                slug = lid.replace('pant_leg_length-', '')
            else:
                slug = lv['name'].lower().replace(' ', '')
            norm = normalize_length_name(lv['name'])
            length_slug_map[norm] = slug

        # ── Detect product-level pack size ──
        product_pack_size = detect_pack_size(product_name)

        # ── Build rows ──
        def make_base_row(color_name='', swatch_url='', availability='', pack_size=1, page_text_data=''):
            """Create a base row dict with shared product data."""
            return {
                'url': url,
                'product_name': product_name,
                'brand': brand,
                'brand_type': classify_brand(brand),
                'color': color_name,
                'pack_size': pack_size,
                'current_price': product_current_price,
                'current_price_formatted': product_current_price_fmt,
                'price_per_unit': '',
                'original_price': product_original_price,
                'original_price_formatted': product_original_price_fmt,
                'original_price_per_unit': '',
                'on_sale': 0,
                'discount_pct': '',
                'price_short': '',
                'price_full': '',
                'price_long': '',
                'pant_rise': pant_rise,
                'pant_leg_cut': pant_leg_cut,
                'clothing_fit': clothing_fit,
                'pant_style': pant_style,
                'jean_wash': jean_wash,
                'inseam': inseam,
                'pant_leg_length': pant_leg_length,
                'fabric_material': fabric_material,
                'fabric_pct': fabric_pct,
                'clothing_size_group': clothing_size_group,
                'clothing_occasion': clothing_occasion,
                'fastener_type': fastener_type,
                'fabric_care': fabric_care,
                'gender': gender,
                'age_group': age_group,
                'category_path': category_path,
                'total_colors': total_colors,
                'total_sizes': total_sizes,
                'sizes_list': sizes_list,
                'total_lengths': total_lengths,
                'lengths_list': lengths_list,
                'average_rating': average_rating,
                'num_reviews': num_reviews,
                'swatch_url': swatch_url,
                'primary_image': primary_image,
                'long_description': long_description_clean,
                'product_details': product_details,
                'size_and_fit': size_and_fit,
                'length_hit': length_hit,
                'page_text': page_text_data,
                'availability': availability,
                'timestamp': datetime.now().isoformat(),
            }

        def apply_pricing(row, current, original):
            """Apply pricing, compute on_sale/discount_pct, and price_per_unit."""
            row['current_price'] = current
            row['current_price_formatted'] = f'${current:.2f}' if current else ''
            if original:
                row['original_price'] = original
                row['original_price_formatted'] = f'${original:.2f}' if original else ''
            if original and current and original > current:
                row['on_sale'] = 1
                row['discount_pct'] = f"{((original - current) / original) * 100:.1f}%"
            elif current:
                row['on_sale'] = 0
                row['discount_pct'] = ''
            # Price per unit
            ps = row.get('pack_size', 1) or 1
            if current and ps > 1:
                row['price_per_unit'] = round(current / ps, 2)
            elif current:
                row['price_per_unit'] = current
            if original and ps > 1:
                row['original_price_per_unit'] = round(original / ps, 2)
            elif original:
                row['original_price_per_unit'] = original

        if not color_variants:
            # No color variants — single row
            color = get_spec_value(specifications, ['color', 'actual color'])
            pack_size = max(product_pack_size, detect_pack_size(color))
            row = make_base_row(color_name=color, pack_size=pack_size, page_text_data=page_text)
            apply_pricing(row, product_current_price, product_original_price)

            # Try per-length pricing from combo_prices
            if combo_prices:
                color_slug = color.lower().replace(' ', '')
                for norm, l_slug in length_slug_map.items():
                    key = (color_slug, l_slug)
                    if key in combo_prices:
                        price_key = 'price_' + norm.lower()
                        if price_key in row:
                            row[price_key] = combo_prices[key]['current_price']

            rows.append(row)

        else:
            # One row per color
            for color_obj in color_variants:
                color_name = color_obj.get('name', '')
                swatch_url = color_obj.get('swatchImageUrl', '')
                availability = color_obj.get('availabilityStatus', '')

                # Detect per-color pack size
                pack_size = max(product_pack_size, detect_pack_size(color_name))

                row = make_base_row(
                    color_name=color_name,
                    swatch_url=swatch_url,
                    availability=availability,
                    pack_size=pack_size,
                    page_text_data=page_text,
                )

                # ── Per-color pricing from combo_prices ──
                c_slug = color_slug_map.get(color_name.lower(), color_name.lower().replace(' ', ''))
                color_price_set = False

                if combo_prices:
                    # Find the "default" price for this color — prefer "regular" length, then any
                    default_length = length_slug_map.get('Full', 'regular')
                    key = (c_slug, default_length)
                    if key in combo_prices:
                        p = combo_prices[key]
                        apply_pricing(row, p['current_price'], p.get('original_price'))
                        color_price_set = True
                    else:
                        # Try any length for this color
                        for (cs, ls), p in combo_prices.items():
                            if cs == c_slug and p['current_price']:
                                apply_pricing(row, p['current_price'], p.get('original_price'))
                                color_price_set = True
                                break

                if not color_price_set:
                    # Fall back to product-level pricing
                    apply_pricing(row, product_current_price, product_original_price)

                # ── Per-length pricing for this color ──
                if combo_prices and length_slug_map:
                    for norm, l_slug in length_slug_map.items():
                        price_key = 'price_' + norm.lower()
                        if price_key not in row:
                            continue
                        key = (c_slug, l_slug)
                        if key in combo_prices and combo_prices[key]['current_price']:
                            row[price_key] = combo_prices[key]['current_price']
                elif length_variant_objs and not has_variant_prices:
                    # No variant prices in JSON — mark for DOM-based extraction later
                    row['_needs_length_prices'] = True

                rows.append(row)

    except Exception as e:
        rows = [{
            'url': url,
            'error': f'Parse error: {str(e)[:200]}',
            'timestamp': datetime.now().isoformat(),
        }]

    return rows


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 2.0
        self.current_delay = 2.0
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'url', 'product_name', 'brand', 'brand_type', 'color', 'pack_size',
    'current_price', 'current_price_formatted', 'price_per_unit',
    'original_price', 'original_price_formatted', 'original_price_per_unit',
    'on_sale', 'discount_pct',
    'price_short', 'price_full', 'price_long',
    'pant_rise', 'pant_leg_cut', 'clothing_fit', 'pant_style', 'jean_wash', 'inseam',
    'pant_leg_length', 'fabric_material', 'fabric_pct', 'clothing_size_group',
    'clothing_occasion', 'fastener_type', 'fabric_care', 'gender', 'age_group',
    'category_path', 'total_colors', 'total_sizes', 'sizes_list',
    'total_lengths', 'lengths_list',
    'average_rating', 'num_reviews', 'swatch_url', 'primary_image',
    'long_description', 'product_details', 'size_and_fit', 'length_hit', 'page_text',
    'availability', 'retries', 'timestamp', 'error',
]


def load_progress(d):
    """Load progress and results from files."""
    p = os.path.join(d, "walmart_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "walmart_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        rec[headers[i]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    """Save progress to JSON file."""
    with open(os.path.join(d, "walmart_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    """Save results to Excel file with formatting."""
    out = os.path.join(d, "walmart_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Walmart PDP Results'

    # Dark header styling
    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_HEADERS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    # Freeze header
    ws.freeze_panes = 'A2'

    # Column widths
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Create a browser context with anti-detection measures."""
    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return ctx


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_walmart_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single Walmart PDP page. Returns a LIST of result dicts (one per color)."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
        status = resp.status if resp else 0

        if status == 403 or status == 429 or status == 503:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        page_url = page.url.lower()
        if 'captcha' in page_url or 'blocked' in page_url or 'denied' in page_url or 'verify' in page_url:
            print(f"    🔒 Redirected to CAPTCHA page on {url[:60]}...")
            await throttle.on_block()

            # Try to solve
            solved = await solve_press_and_hold(page)
            if solved:
                throttle.reset()  # No penalty after successful solve
                # Navigate back to the original URL after solving
                try:
                    resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
                    await asyncio.sleep(random.uniform(2, 4))
                except Exception:
                    pass
                # Fall through to the rest of the scraping logic
            else:
                await stats.inc_blocked()
                wait_time = 60 * (retries + 1) + random.uniform(20, 40)
                print(f"    CAPTCHA unsolved, backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
                if retries < MAX_RETRIES:
                    await asyncio.sleep(wait_time)
                    return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
                return [{'url': url, 'error': 'Blocked/CAPTCHA redirect', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Human-like behavior: wait for page, then scroll a bit
        await asyncio.sleep(random.uniform(2.5, 4.5))

        # Dismiss "Robot or human?" popup if it appears (can show up even without CAPTCHA redirect)
        await _dismiss_robot_popup(page)

        try:
            await page.evaluate('window.scrollBy(0, Math.floor(Math.random() * 400 + 200))')
            await asyncio.sleep(random.uniform(0.5, 1.5))
        except Exception:
            pass
        html_content = await page.content()

        # Check for actual press-and-hold CAPTCHA (not just <meta name="robots">)
        html_lower = html_content.lower()
        has_captcha = (
            'press & hold' in html_lower or
            'press and hold' in html_lower or
            ('are you a robot' in html_lower) or
            ('px-captcha' in html_lower)
        )
        if has_captcha:
            print(f"    🔒 CAPTCHA detected on {url[:60]}...")
            await throttle.on_block()

            # Try to solve the press-and-hold CAPTCHA
            solved = await solve_press_and_hold(page)

            if solved:
                throttle.reset()  # No penalty after successful solve
                # Re-extract the page content after solving
                html_content = await page.content()
                # Fall through to normal parsing below
            else:
                # Could not solve — back off and retry
                await stats.inc_blocked()
                wait_time = 60 * (retries + 1) + random.uniform(20, 40)
                print(f"    CAPTCHA unsolved, backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
                if retries < MAX_RETRIES:
                    await asyncio.sleep(wait_time)
                    return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
                return [{'url': url, 'error': 'CAPTCHA/press-and-hold unsolved', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Verify we got product data
        if '__NEXT_DATA__' not in html_content:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': 'No __NEXT_DATA__ in page', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Capture page_text via JavaScript for keyword searching
        page_text_content = ''
        try:
            page_text_content = await page.evaluate(WALMART_PAGE_TEXT_JS)
            if not page_text_content:
                page_text_content = ''
        except Exception:
            # If JS evaluation fails, fallback to extracting from HTML (less ideal)
            page_text_content = html_content[:10000]

        # Enable JSON dump for the very first URL processed (debugging)
        is_first_url = (stats.completed + stats.failed) == 0
        color_rows = parse_walmart_pdp(html_content, url, dump_json=is_first_url)
        for row in color_rows:
            row['retries'] = retries
            row['url'] = url
            # Populate page_text in the result rows
            row['page_text'] = page_text_content

        # ── Phase 2: DOM-based per-color + per-length pricing (fallback) ──
        # Only needed if JSON didn't have per-variant prices
        needs_dom_pricing = (
            len(color_rows) > 1
            and not color_rows[0].get('error')
            and any(r.get('_needs_length_prices') or r.get('current_price', 0) == 0 for r in color_rows)
        )

        if needs_dom_pricing:
            try:
                per_color_prices = {}

                # Broad selector strategy — Walmart uses various DOM patterns:
                # chips, image swatches, radio buttons, links, divs
                swatch_info = await page.evaluate(r"""() => {
                    var results = [];
                    var selectors = [
                        '[data-testid*="variant"][data-testid*="olor"] button',
                        '[data-testid*="variant"][data-testid*="olor"] a',
                        '[data-testid*="variant"][data-testid*="olor"] label',
                        '[data-testid*="variant-group"] [role="radio"]',
                        '[data-testid*="Color"] button',
                        '[data-testid*="Color"] a',
                        '[aria-label*="olor"][role="radiogroup"] [role="radio"]',
                        '[aria-label*="olor"] button',
                        'section[aria-label*="olor"] button',
                        'section[aria-label*="olor"] a',
                        'div[class*="variant"] button[aria-label]',
                        'div[class*="variant"] a[aria-label]',
                        'button[data-swatch-id]',
                        '[role="radio"][aria-label]',
                        'img[data-testid*="swatch"]'
                    ];
                    var chips = [];
                    var winningSelector = '';
                    for (var s = 0; s < selectors.length; s++) {
                        chips = document.querySelectorAll(selectors[s]);
                        if (chips.length > 1) { winningSelector = selectors[s]; break; }
                    }
                    for (var i = 0; i < chips.length; i++) {
                        var el = chips[i];
                        var name = el.getAttribute('aria-label') || el.getAttribute('title') ||
                                   el.getAttribute('data-variant-name') || '';
                        name = name.replace(/^selected:?\s*/i, '').replace(/^color:?\s*/i, '')
                                   .replace(/^actual color:?\s*/i, '').trim();
                        if (!name) {
                            var img = el.querySelector('img') || (el.tagName === 'IMG' ? el : null);
                            if (img) name = (img.alt || img.getAttribute('data-name') || '').replace(/^selected:?\s*/i, '').trim();
                        }
                        if (!name) name = el.textContent.trim();
                        if (name && name.length < 80) results.push({index: i, name: name, selector: winningSelector});
                    }
                    return results;
                }""")

                print(f"    DOM swatch scan: found {len(swatch_info)} elements" +
                      (f" via '{swatch_info[0].get('selector', '')}'" if swatch_info else ''))

                for info in swatch_info:
                    try:
                        swatch_name = info['name']
                        sel = info.get('selector', '')
                        idx = info['index']

                        if sel:
                            swatches = await page.query_selector_all(sel)
                        else:
                            continue
                        if idx >= len(swatches):
                            continue

                        await swatches[idx].click()
                        await asyncio.sleep(random.uniform(1.5, 3.0))

                        price = await page.evaluate(WALMART_PRICE_JS)
                        if price and isinstance(price, (int, float)):
                            per_color_prices[swatch_name.lower()] = price
                    except Exception:
                        continue

                # Match swatch prices to color rows
                if per_color_prices:
                    matched = 0
                    for row in color_rows:
                        json_color = row.get('color', '').lower()
                        if json_color in per_color_prices:
                            new_price = per_color_prices[json_color]
                        else:
                            new_price = None
                            for sn, sp in per_color_prices.items():
                                if json_color in sn or sn in json_color:
                                    new_price = sp
                                    break
                        if new_price is not None:
                            row['current_price'] = new_price
                            row['current_price_formatted'] = f'${new_price:.2f}'
                            orig = row.get('original_price')
                            if orig and isinstance(orig, (int, float)) and orig > new_price:
                                row['on_sale'] = 1
                                row['discount_pct'] = f"{((orig - new_price) / orig) * 100:.1f}%"
                            matched += 1
                    print(f"    DOM per-color prices: {matched}/{len(color_rows)} matched from {len(per_color_prices)} swatches")

                # ── Length tab clicking for per-length pricing ──
                if any(r.get('_needs_length_prices') for r in color_rows):
                    length_tab_info = await page.evaluate(r"""() => {
                        var results = [];
                        var selectors = [
                            '[data-testid*="variant"][data-testid*="ength"] button',
                            '[data-testid*="variant"][data-testid*="ength"] a',
                            '[data-testid*="Pant Leg"] button',
                            '[data-testid*="Pant Leg"] a',
                            '[aria-label*="ength"][role="radiogroup"] [role="radio"]',
                            '[aria-label*="ength"] button',
                            'section[aria-label*="ength"] button',
                            'section[aria-label*="Pant Leg"] button'
                        ];
                        var tabs = [];
                        var winSel = '';
                        for (var s = 0; s < selectors.length; s++) {
                            tabs = document.querySelectorAll(selectors[s]);
                            if (tabs.length > 1) { winSel = selectors[s]; break; }
                        }
                        for (var i = 0; i < tabs.length; i++) {
                            var el = tabs[i];
                            var name = el.getAttribute('aria-label') || el.textContent.trim() || '';
                            name = name.replace(/^selected:?\s*/i, '').trim();
                            if (name && name.length < 30) results.push({index: i, name: name, selector: winSel});
                        }
                        return results;
                    }""")

                    print(f"    Length tabs: found {len(length_tab_info)} tabs")

                    for tab_info in length_tab_info:
                        try:
                            tab_name = tab_info['name']
                            norm = normalize_length_name(tab_name)
                            price_key = 'price_' + norm.lower()
                            sel = tab_info.get('selector', '')
                            idx = tab_info['index']

                            if not sel:
                                continue
                            tabs = await page.query_selector_all(sel)
                            if idx >= len(tabs):
                                continue

                            await tabs[idx].click()
                            await asyncio.sleep(random.uniform(1.5, 3.0))

                            price = await page.evaluate(WALMART_PRICE_JS)
                            if price and isinstance(price, (int, float)):
                                for row in color_rows:
                                    if price_key in row:
                                        row[price_key] = price
                                print(f"      {norm}: ${price:.2f}")
                        except Exception:
                            continue

            except Exception as e:
                print(f"    Phase 2 DOM pricing skipped: {str(e)[:80]}")

        # Clean up internal flags
        for row in color_rows:
            row.pop('_needs_length_prices', None)

        await stats.inc_ok(len(color_rows))
        await throttle.on_ok()
        return color_rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_walmart_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    """Worker coroutine that processes URLs from the queue."""
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  W{wid}: 30-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url[:80]}...")

        await throttle.wait()
        color_rows = await scrape_walmart_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(color_rows)
            progress['processed'].add(url)

        colors_found = len(color_rows)
        has_error = any(r.get('error') for r in color_rows)
        if has_error:
            print(f"    W{wid}: Error — {color_rows[0].get('error', '')[:80]}")
        else:
            print(f"    W{wid}: {colors_found} color rows extracted")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    """Run a batch of URLs with a fresh browser and contexts."""
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   {len(results)} existing rows, {len(remaining)} URLs remaining")

    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-gpu",
        "--window-size=1440,900",
    ])

    async def block_resources(route):
        """Block heavy resources but keep CSS/JS (Walmart needs them for bot detection)."""
        await route.abort()

    all_pages = []
    all_contexts = []
    print(f"Launching {NUM_CONTEXTS} contexts x {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            # Only block images — keep JS/CSS so Perimeter X bot detection passes
            await pg.route('**/*.{png,jpg,jpeg,gif,svg,ico,webp}', block_resources)
            all_pages.append(pg)

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\nBrowser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(sdir, "walmart.csv")

    if not os.path.exists(csv_path):
        print(f"CSV not found at {csv_path}")
        return

    # Load URLs from CSV (first column is 'w-100 href')
    urls = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for row in reader:
                if row:
                    url = row[0].strip()
                    if url.startswith('http'):
                        urls.append(url)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return

    print(f"Loaded {len(urls)} Walmart URLs")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 20

    crash_count = 0
    try:
        while True:
            remaining = [u for u in urls if u not in progress['processed']]
            if not remaining:
                print("All URLs processed!")
                break
            print(f"\n{'=' * 60}")
            print(f"Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
            print(f"{'=' * 60}")

            was_crash = False
            try:
                pw = await async_playwright().start()
                finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
                try:
                    await pw.stop()
                except Exception:
                    pass
                if finished:
                    break
            except Exception as e:
                print(f"\nPlaywright/browser error: {str(e)[:150]}")
                was_crash = True
                try:
                    await pw.stop()
                except Exception:
                    pass

            throttle.reset()
            save_progress(progress, sdir)
            save_to_excel(results, sdir)

            if was_crash:
                crash_count += 1
                if crash_count >= max_crashes:
                    print(f"Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                    break
                wait = min(30, 10 * crash_count)
                print(f"Crash restart in {wait}s... (crash #{crash_count})")
                await asyncio.sleep(wait)
            else:
                crash_count = 0
                print(f"Fresh restart in 5s...")
                await asyncio.sleep(5)

    except KeyboardInterrupt:
        print(f"\n\n{'=' * 60}")
        print(f"Ctrl+C detected — saving progress before exit...")
        print(f"{'=' * 60}")

    # Always save on exit (normal completion, crash, or Ctrl+C)
    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"Saved {len(results)} rows to: {out}")
    print(f"{len(progress['processed'])} URLs processed — run again to resume remaining")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
