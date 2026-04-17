#!/usr/bin/env python3
"""
Production Old Navy PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  oldnavy.csv  (URLs in first column 'plp_product-image href')
Output: oldnavy_pdp_results.xlsx + oldnavy_pdp_progress.json

Architecture: 5 browser contexts × 1 tab each = 5 parallel workers
              20-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

Old Navy (Gap Inc) uses window.gap JavaScript object for product data.
Extracts from CSR app via Playwright evaluate() to produce
one row per color variant with pricing, specs, ratings, and fabric.
"""

import asyncio, json, os, random, re, time, csv
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Price Extraction JS Constant ────────────────────────────────────────────
OLDNAVY_PRICE_JS = """() => {
    // Old Navy / Gap family price selectors
    var el = document.querySelector('[data-testid*="sale-price"]') ||
             document.querySelector('[data-testid*="current-price"]') ||
             document.querySelector('.product-price__sale') ||
             document.querySelector('.product-price__highlight') ||
             document.querySelector('[class*="ProductPrice"] [class*="sale"]') ||
             document.querySelector('[class*="ProductPrice"]');
    if (el) {
        var text = el.textContent.trim();
        var m = text.match(/\\$(\\d+\\.?\\d*)/);
        if (m) return parseFloat(m[1]);
    }
    // Broader fallback
    var allPrices = document.querySelectorAll('[class*="price"]');
    for (var i = 0; i < allPrices.length; i++) {
        var t = allPrices[i].textContent.trim();
        var match = t.match(/\\$(\\d+\\.?\\d*)/);
        if (match) return parseFloat(match[1]);
    }
    return null;
}"""

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 5
TABS_PER_CTX = 1
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 5 workers
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 20 * 60  # 20 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]


def classify_brand(brand_name):
    """Old Navy is always owned brand (OB)."""
    return 'OB'


# ── Non-Basic Detection ───────────────────────────────────────────────────
NON_BASIC_KEYWORDS = [
    'print', 'printed', 'graphic', 'pattern', 'floral', 'stripe', 'striped',
    'plaid', 'camo', 'camouflage', 'tie-dye', 'tie dye', 'leopard', 'logo',
    'embroidered', 'embroidery', 'novelty', 'character', 'varsity', 'retro',
    'animal', 'abstract', 'geometric', 'paisley', 'polka', 'check', 'checked',
    'tropical', 'botanical', 'ditsy', 'patchwork', 'colorblock', 'ombre',
]


def is_non_basic(title, colors_str, bullets_str):
    """Returns 1 if product is non-basic (graphic/patterned), 0 otherwise."""
    combined = f"{title} {colors_str} {bullets_str}".lower()
    for kw in NON_BASIC_KEYWORDS:
        if kw in combined:
            return 1
    return 0


# ── Rise / Leg / Fit Parsing from Title ────────────────────────────────────

def parse_rise(title):
    """Extract rise from product title."""
    t = title.lower()
    if 'super high' in t or 'super-high' in t:
        return 'Super High Rise'
    if 'high-rise' in t or 'high rise' in t:
        return 'High Rise'
    if 'mid-rise' in t or 'mid rise' in t:
        return 'Mid Rise'
    if 'low-rise' in t or 'low rise' in t:
        return 'Low Rise'
    return ''


def parse_leg_shape(title):
    """Extract leg shape from product title."""
    t = title.lower()
    if 'wide' in t and 'leg' in t:
        return 'Wide Leg'
    if 'wide' in t:
        return 'Wide Leg'
    if 'flare' in t:
        return 'Flare'
    if 'bootcut' in t or 'boot cut' in t or 'boot-cut' in t:
        return 'Bootcut'
    if 'straight' in t:
        return 'Straight'
    if 'slim' in t and 'straight' not in t:
        return 'Slim'
    if 'skinny' in t:
        return 'Skinny'
    if 'tapered' in t or 'taper' in t:
        return 'Tapered'
    if 'relaxed' in t:
        return 'Relaxed'
    if 'boyfriend' in t:
        return 'Boyfriend'
    if 'jegging' in t:
        return 'Jegging'
    if 'crop' in t:
        return 'Crop'
    if 'baggy' in t:
        return 'Baggy'
    if 'barrel' in t:
        return 'Barrel'
    return ''


def parse_fit(title):
    """Extract fit from product title."""
    t = title.lower()
    if 'relaxed' in t:
        return 'Relaxed'
    if 'slim' in t:
        return 'Slim'
    if 'curvy' in t:
        return 'Curvy'
    if 'regular' in t:
        return 'Regular'
    return ''


# ── Material Parsing ──────────────────────────────────────────────────────

def parse_material(fabric_raw):
    """
    Parse fabric composition string and return (fabric_parsed, pct_cotton, pct_natural).
    Handles: '98% Cotton, 2% Spandex' or 'Front: 90% Cotton | Back: 88% Cotton'
    Only sums first panel to avoid double-counting.
    """
    if not fabric_raw:
        return ('', '', '')

    clean = fabric_raw.strip()
    # Only take first panel if multi-panel
    first_panel = clean.split('|')[0].strip()

    cotton_pct = 0
    natural_pct = 0
    # Pattern: number% MaterialName
    matches = re.findall(r'(\d+(?:\.\d+)?)\s*%\s*([A-Za-z\s]+)', first_panel)
    for pct_str, name in matches:
        pct = float(pct_str)
        name_lower = name.strip().lower()
        if 'cotton' in name_lower:
            cotton_pct += pct
        # Natural fibers
        if any(f in name_lower for f in ['cotton', 'linen', 'silk', 'wool', 'hemp', 'bamboo', 'tencel', 'modal', 'lyocell', 'cashmere']):
            natural_pct += pct

    cotton_str = f"{cotton_pct:.0f}%" if cotton_pct > 0 else ''
    natural_str = f"{natural_pct:.0f}%" if natural_pct > 0 else ''
    return (clean, cotton_str, natural_str)


# ── Old Navy Page Extraction (via Playwright evaluate) ──────────────────────

EXTRACT_JS = """
() => {
    const result = {};

    // --- Try window.gap first (CSR product data) ---
    let gotGapData = false;
    try {
        if (window.gap && window.gap.productData) {
            const pd = window.gap.productData;
            result.title = pd.productName || '';
            result.brand = 'Old Navy';
            result.current_price = pd.price || '';
            result.original_price = pd.originalPrice || pd.regularPrice || '';
            result.product_id = pd.productId || '';
            result.rating = pd.rating || '';
            result.review_count = pd.reviewCount || '';
            result.colors = pd.colors || [];
            result.sizes = pd.sizes || [];
            result.image_count = (pd.images && pd.images.length) || 0;
            result.fabric_raw = pd.fabricContent || '';
            result.from_gap_object = true;
            gotGapData = true;
        }
    } catch (e) {}

    // --- DOM extraction (full fallback, or supplement gap object data) ---
    if (!gotGapData) {
    // Title
    result.title = '';
    const titleEl = document.querySelector('h1.productTitle, [data-testid="product-title"], h1[itemprop="name"]');
    if (titleEl) result.title = titleEl.textContent.trim();

    // Brand
    result.brand = 'Old Navy';

    // Pricing — Old Navy shows "Now $XX" and "Was $XX"
    result.current_price = '';
    result.original_price = '';
    const priceEls = document.querySelectorAll('[data-testid="current-price"], .productPrice, .price-now, span[class*="price"]');
    for (const el of priceEls) {
        const text = el.textContent.trim();
        if (text && /^\\$[\\d,.]+/.test(text) && !result.current_price) {
            result.current_price = text.match(/\\$[\\d,.]+/)[0];
            break;
        }
    }

    const origEls = document.querySelectorAll('[data-testid="original-price"], .originalPrice, .price-was, [class*="strike"], [class*="original"]');
    for (const el of origEls) {
        const text = el.textContent.trim();
        if (text && /^\\$[\\d,.]+/.test(text) && !result.original_price) {
            result.original_price = text.match(/\\$[\\d,.]+/)[0];
            break;
        }
    }

    // Color swatches — look for color name attributes/alts
    result.colors = [];
    const colorSwatches = document.querySelectorAll('[role="button"][data-color], [data-testid*="color"], .colorSwatch, [aria-label*="color"]');
    for (const swatch of colorSwatches) {
        let colorName = swatch.getAttribute('data-color') || swatch.getAttribute('aria-label') || swatch.alt || '';
        colorName = colorName.replace(/color\\s*[:–-]?\\s*/i, '').trim();
        if (colorName && result.colors.indexOf(colorName) === -1) {
            result.colors.push(colorName);
        }
    }

    // Fallback: extract color names from color buttons
    if (!result.colors.length) {
        const colorButtons = document.querySelectorAll('button[data-color], button[aria-label*="color"], li[data-color]');
        for (const btn of colorButtons) {
            let name = btn.getAttribute('data-color') || btn.getAttribute('aria-label') || btn.textContent.trim();
            name = name.replace(/color\\s*[:–-]?\\s*/i, '').trim();
            if (name && name.length > 1 && result.colors.indexOf(name) === -1) {
                result.colors.push(name);
            }
        }
    }

    // Sizes
    result.sizes = [];
    const sizeSelectors = ['[data-testid*="size"]', '.sizeButton', '[role="button"][data-size]', 'button[aria-label*="size"]'];
    for (const sel of sizeSelectors) {
        const sizeEls = document.querySelectorAll(sel);
        if (sizeEls.length > 0) {
            for (const el of sizeEls) {
                let size = el.getAttribute('data-size') || el.textContent.trim();
                size = size.replace(/size\\s*[:–-]?\\s*/i, '').trim();
                if (size && size.length <= 10 && result.sizes.indexOf(size) === -1) {
                    result.sizes.push(size);
                }
            }
            if (result.sizes.length > 0) break;
        }
    }

    // Rating
    result.rating = '';
    const ratingEl = document.querySelector('[data-testid*="rating"], [itemprop="ratingValue"], .productRating, .rating');
    if (ratingEl) {
        const m = ratingEl.textContent.match(/([\\d.]+)/);
        if (m) result.rating = m[1];
    }

    // Review count
    result.review_count = '';
    const reviewEl = document.querySelector('[data-testid*="review"], [itemprop="reviewCount"], .reviewCount');
    if (reviewEl) {
        const m = reviewEl.textContent.match(/([\\d,]+)/);
        if (m) result.review_count = m[1].replace(/,/g, '');
    }

    // Fabric/Material
    result.fabric_raw = '';
    const fabricKeys = ['Fabric', 'Material', 'Composition', 'Content'];
    const detailRows = document.querySelectorAll('[data-testid*="details"] li, .productDetails li, .specs li, dl dt, .detail-row');
    for (const row of detailRows) {
        const text = row.textContent.toLowerCase();
        for (const key of fabricKeys) {
            if (text.includes(key.toLowerCase())) {
                result.fabric_raw = row.textContent.replace(new RegExp(key, 'i'), '').trim();
                break;
            }
        }
        if (result.fabric_raw) break;
    }

    // Breadcrumb
    result.breadcrumb = '';
    const breadcrumbs = document.querySelectorAll('[data-testid*="breadcrumb"] a, .breadcrumb a, nav a[href*="/"]');
    const breadcrumbTexts = [];
    for (const el of breadcrumbs) {
        const t = el.textContent.trim();
        if (t && t.length > 0 && t.length < 100) {
            breadcrumbTexts.push(t);
        }
    }
    if (breadcrumbTexts.length > 0) {
        result.breadcrumb = breadcrumbTexts.join(' > ');
    }

    // Image count
    result.image_count = 0;
    const images = document.querySelectorAll('[data-testid*="image"], .productImage, img[alt*="product"]');
    result.image_count = images.length || 1;

    // Feature bullets / description
    result.feature_bullets = [];
    const bulletEls = document.querySelectorAll('ul[data-testid*="description"] li, .features li, [class*="bullet"]');
    for (const el of bulletEls) {
        const t = el.textContent.trim();
        if (t && t.length > 5) {
            result.feature_bullets.push(t);
        }
    }
    } // end if (!gotGapData)

    // --- Always extract these from DOM (even with gap object data) ---

    // --- Product details section ---
    // Look for the "Product details" expandable section and grab all bullet items
    result.product_details = [];
    const pdSections = document.querySelectorAll('[class*="product-details"], [data-testid*="product-details"], [class*="ProductDetails"], [id*="product-details"]');
    for (const sec of pdSections) {
        const lis = sec.querySelectorAll('li');
        for (const li of lis) {
            const t = li.textContent.trim();
            if (t && t.length > 1 && result.product_details.indexOf(t) === -1) {
                result.product_details.push(t);
            }
        }
    }
    // Fallback: look for accordion/expandable sections by heading text
    if (!result.product_details.length) {
        const allSections = document.querySelectorAll('div[class*="accordion"], div[class*="Accordion"], details, [class*="expandable"]');
        for (const sec of allSections) {
            const heading = sec.querySelector('h2, h3, h4, button, summary');
            if (heading && /product\\s*detail/i.test(heading.textContent)) {
                const lis = sec.querySelectorAll('li');
                for (const li of lis) {
                    const t = li.textContent.trim();
                    if (t && t.length > 1) result.product_details.push(t);
                }
            }
        }
    }

    // --- Size & fit section ---
    result.size_and_fit = [];
    const sfSections = document.querySelectorAll('[class*="size-fit"], [data-testid*="size-fit"], [class*="SizeFit"], [id*="size-fit"]');
    for (const sec of sfSections) {
        const lis = sec.querySelectorAll('li');
        for (const li of lis) {
            const t = li.textContent.trim();
            if (t && t.length > 1 && result.size_and_fit.indexOf(t) === -1) {
                result.size_and_fit.push(t);
            }
        }
    }
    // Fallback: accordion heading match
    if (!result.size_and_fit.length) {
        const allSections = document.querySelectorAll('div[class*="accordion"], div[class*="Accordion"], details, [class*="expandable"]');
        for (const sec of allSections) {
            const heading = sec.querySelector('h2, h3, h4, button, summary');
            if (heading && /size\\s*[&+]?\\s*fit/i.test(heading.textContent)) {
                const lis = sec.querySelectorAll('li');
                for (const li of lis) {
                    const t = li.textContent.trim();
                    if (t && t.length > 1) result.size_and_fit.push(t);
                }
            }
        }
    }

    // --- Full page text for keyword search ---
    result.page_text = '';
    try {
        // Get main product area text, skip nav/footer noise
        const main = document.querySelector('main, [role="main"], #content, .pdp-page, .product-detail');
        if (main) {
            result.page_text = main.innerText.substring(0, 10000);
        } else {
            result.page_text = document.body.innerText.substring(0, 10000);
        }
    } catch(e) {}

    result.from_gap_object = false;
    return result;
}
"""


def parse_oldnavy_pdp(raw_data, url):
    """
    Take the raw JS extraction result and build a LIST of row dicts — one per color.
    If no colors found, returns single row.
    """
    rows = []

    try:
        title = raw_data.get('title', '')
        brand = raw_data.get('brand', 'Old Navy')

        # Pricing
        current_price_str = raw_data.get('current_price', '')
        original_price_str = raw_data.get('original_price', '')

        def price_to_float(p):
            if not p:
                return None
            return float(re.sub(r'[^0-9.]', '', p)) if re.search(r'[\d]', p) else None

        current_price = price_to_float(current_price_str)
        original_price = price_to_float(original_price_str)

        on_sale = 0
        discount_pct = ''
        if original_price and current_price and original_price > current_price:
            on_sale = 1
            discount_pct = f"{((original_price - current_price) / original_price) * 100:.1f}%"

        if not original_price:
            original_price = current_price
            original_price_str = current_price_str

        # Rating
        rating = raw_data.get('rating', '')
        review_count = raw_data.get('review_count', '')

        # Category
        breadcrumb = raw_data.get('breadcrumb', '')

        # Colors / Sizes
        colors = raw_data.get('colors', [])
        sizes = raw_data.get('sizes', [])
        total_colors = len(colors) if colors else 1
        total_sizes = len(sizes)
        sizes_list = ', '.join(sizes) if sizes else ''

        # Product details & size/fit from DOM
        product_details = raw_data.get('product_details', [])
        product_details_str = ' | '.join(product_details) if product_details else ''
        size_and_fit = raw_data.get('size_and_fit', [])
        size_and_fit_str = ' | '.join(size_and_fit) if size_and_fit else ''

        # Combined text for smarter parsing
        all_detail_text = ' '.join([title] + product_details + size_and_fit).lower()

        # Rise / Leg / Fit — check title first, then product details
        rise = parse_rise(title)
        if not rise:
            rise = parse_rise(' '.join(product_details))
        leg_shape = parse_leg_shape(title)
        if not leg_shape:
            leg_shape = parse_leg_shape(' '.join(product_details))
        fit = parse_fit(title)
        if not fit:
            fit = parse_fit(' '.join(product_details))

        # Length / hit-point from size & fit (e.g. "hits at ankle", "hits at knee")
        length_hit = ''
        for item in size_and_fit:
            low = item.lower()
            if 'hits at' in low or 'hits above' in low or 'hits below' in low:
                length_hit = item.strip()
                break
        # Inseam info
        inseam = ''
        for item in size_and_fit:
            low = item.lower()
            if 'inseam' in low:
                inseam = item.strip()
                if not inseam:
                    pass
                break
        if not inseam:
            # Try title for inseam
            import re as _re
            m = _re.search(r'(\d+)["\u2033]?\s*inseam', title, _re.IGNORECASE)
            if m:
                inseam = f'{m.group(1)}" inseam'

        # Material
        fabric_raw = raw_data.get('fabric_raw', '')
        fabric_parsed, pct_cotton, pct_natural = parse_material(fabric_raw)

        # Feature bullets
        bullets = raw_data.get('feature_bullets', [])
        bullets_str = ' | '.join(bullets) if bullets else ''

        # Full page text for keyword search
        page_text = raw_data.get('page_text', '')

        # Non-basic detection
        colors_str = ', '.join(colors) if colors else ''
        non_basic = is_non_basic(title, colors_str, bullets_str)

        # Image count
        image_count = raw_data.get('image_count', 0)

        # Build rows — one per color
        def make_row(color_name):
            return {
                'url': url,
                'product_name': title,
                'brand': brand,
                'brand_type': classify_brand(brand),
                'color': color_name,
                'current_price': current_price,
                'current_price_formatted': current_price_str,
                'original_price': original_price,
                'original_price_formatted': original_price_str,
                'on_sale': on_sale,
                'discount_pct': discount_pct,
                'rise': rise,
                'leg_shape': leg_shape,
                'fit': fit,
                'length_hit': length_hit,
                'inseam': inseam,
                'fabric_raw': fabric_raw,
                'fabric_parsed': fabric_parsed,
                'pct_cotton': pct_cotton,
                'pct_natural_fiber': pct_natural,
                'non_basic': non_basic,
                'breadcrumb': breadcrumb,
                'total_colors': total_colors,
                'total_sizes': total_sizes,
                'sizes_list': sizes_list,
                'average_rating': rating,
                'review_count': review_count,
                'image_count': image_count,
                'product_details': product_details_str,
                'size_and_fit': size_and_fit_str,
                'feature_bullets': bullets_str,
                'page_text': page_text,
                'timestamp': datetime.now().isoformat(),
            }

        if colors:
            for c in colors:
                rows.append(make_row(c))
        else:
            rows.append(make_row(''))

    except Exception as e:
        rows = [{
            'url': url,
            'error': f'Parse error: {str(e)[:200]}',
            'timestamp': datetime.now().isoformat(),
        }]

    return rows


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 0.5
        self.current_delay = 0.5
        self.max_delay = 30.0
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'url', 'product_name', 'brand', 'brand_type', 'color',
    'current_price', 'current_price_formatted', 'original_price', 'original_price_formatted',
    'on_sale', 'discount_pct',
    'rise', 'leg_shape', 'fit', 'length_hit', 'inseam',
    'fabric_raw', 'fabric_parsed', 'pct_cotton', 'pct_natural_fiber', 'non_basic',
    'breadcrumb',
    'total_colors', 'total_sizes', 'sizes_list',
    'average_rating', 'review_count',
    'image_count', 'product_details', 'size_and_fit', 'feature_bullets',
    'page_text',
    'retries', 'timestamp', 'error',
]


def load_progress(d):
    """Load progress and results from files."""
    p = os.path.join(d, "oldnavy_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "oldnavy_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        rec[headers[i]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    """Save progress to JSON file."""
    with open(os.path.join(d, "oldnavy_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    """Save results to Excel file with formatting."""
    out = os.path.join(d, "oldnavy_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Old Navy PDP Results'

    # Dark header styling
    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_HEADERS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    # Freeze header
    ws.freeze_panes = 'A2'

    # Column widths
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Create a browser context with anti-detection and US locale."""
    ua = USER_AGENTS[ua_index % len(USER_AGENTS)]
    ctx = await browser.new_context(
        user_agent=ua,
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return ctx


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_oldnavy_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single Old Navy PDP page. Returns a LIST of result dicts (one per color)."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=45000)
        status = resp.status if resp else 0

        if status == 503 or status == 429:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_oldnavy_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Check for CAPTCHA or bot detection
        page_content = ''
        try:
            page_content = await page.content()
        except Exception:
            pass

        # Check for actual CAPTCHA/bot block — but ONLY if there's no real product data on page.
        # Old Navy pages often contain "captcha" in script bundles even on normal pages.
        page_lower = page_content.lower()
        has_product_data = (
            'product-title' in page_lower or
            'pdp-price' in page_lower or
            '"product"' in page_lower or
            'product__name' in page_lower or
            '<h1' in page_lower
        )
        is_real_block = (
            not has_product_data and (
                'are you a human' in page_lower[:3000] or
                'verify you are human' in page_lower[:3000] or
                'press & hold' in page_lower[:3000] or
                'press and hold' in page_lower[:3000]
            )
        )
        # "Access denied" as the MAIN page content (not buried in scripts)
        is_access_denied = (
            'access denied' in page_lower[:500] and not has_product_data
        )
        if is_real_block or is_access_denied:
            await stats.inc_blocked()
            await throttle.on_block()
            wait_time = 60 * (retries + 1) + random.uniform(15, 30)
            print(f"    CAPTCHA/block on {url[:60]}... backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
            if retries < MAX_RETRIES:
                await asyncio.sleep(wait_time)
                return await scrape_oldnavy_pdp(page, url, stats, throttle, retries + 1)
            return [{'url': url, 'error': 'CAPTCHA', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_oldnavy_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Wait for product content to load
        try:
            await page.wait_for_selector('h1, [data-testid="product-title"]', timeout=8000)
        except Exception:
            pass

        # Human-like behavior: scroll a bit, wait for dynamic content
        await asyncio.sleep(random.uniform(2, 4))
        try:
            await page.evaluate('window.scrollBy(0, Math.floor(Math.random() * 400 + 200))')
            await asyncio.sleep(random.uniform(1, 2))
        except Exception:
            pass

        # Extract product data
        try:
            raw_data = await page.evaluate(EXTRACT_JS)
        except Exception as eval_err:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_oldnavy_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'Evaluate error: {str(eval_err)[:150]}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Verify we got useful data
        if not raw_data.get('title'):
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(3, 8))
                return await scrape_oldnavy_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': 'No product title found', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        # Parse and return rows
        rows = parse_oldnavy_pdp(raw_data, url)

        # ── Phase 2: Per-color pricing via swatch clicking ──
        if len(rows) > 1 and not rows[0].get('error'):
            try:
                per_color_prices = {}
                swatches = await page.query_selector_all('[data-testid*="color"] button, [role="button"][data-color], button[aria-label*="color"]')
                if not swatches:
                    swatches = await page.query_selector_all('.swatch__button, [class*="ColorSwatch"] button, li[data-color] button')

                for swatch in swatches:
                    try:
                        color_name = await swatch.get_attribute('aria-label') or ''
                        if not color_name:
                            color_name = await swatch.get_attribute('data-color') or ''
                        if not color_name:
                            color_name = await swatch.get_attribute('title') or ''
                        color_name = color_name.strip()
                        if not color_name:
                            continue

                        await swatch.click()
                        await asyncio.sleep(random.uniform(1.0, 2.0))

                        price = await page.evaluate(OLDNAVY_PRICE_JS)
                        if price and isinstance(price, (int, float)):
                            per_color_prices[color_name.lower()] = price
                    except Exception:
                        continue

                if per_color_prices:
                    for row in rows:
                        color = row.get('color', '').lower()
                        if color in per_color_prices:
                            new_price = per_color_prices[color]
                            row['current_price'] = new_price
                            row['current_price_formatted'] = f'${new_price:.2f}'
                            orig = row.get('original_price')
                            if isinstance(orig, str):
                                try:
                                    orig = float(orig.replace('$', '').replace(',', ''))
                                except ValueError:
                                    orig = None
                            if orig and isinstance(orig, (int, float)) and orig > new_price:
                                row['on_sale'] = 1
                                row['discount_pct'] = f"{((orig - new_price) / orig) * 100:.1f}%"
                            elif orig and orig == new_price:
                                row['on_sale'] = 0
                                row['discount_pct'] = ''
                    print(f"    Per-color prices: {len(per_color_prices)} colors updated")
            except Exception as e:
                print(f"    Phase 2 pricing skipped: {str(e)[:80]}")

        await throttle.on_ok()
        await stats.inc_ok(len(rows))
        return rows

    except asyncio.TimeoutError:
        # Re-raise timeout errors so worker can track consecutive timeouts
        raise
    except Exception as e:
        print(f"    Exception scraping {url[:60]}: {str(e)[:100]}")
        await stats.inc_fail()
        return [{'url': url, 'error': f'Exception: {str(e)[:150]}', 'timestamp': datetime.now().isoformat()}]


# ── Worker ──────────────────────────────────────────────────────────────────

async def worker(worker_id, ctx, queue, results, stats, throttle, working_dir):
    """Worker processes URLs from the queue."""
    page = await ctx.new_page()
    crash_count = 0
    consecutive_timeouts = 0
    try:
        while True:
            try:
                url = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            print(f"    [W{worker_id}] Processing: {url[:70]}")

            try:
                rows = await scrape_oldnavy_pdp(page, url, stats, throttle)
                results.extend(rows)
                consecutive_timeouts = 0  # reset on success
                await throttle.wait()

            except asyncio.TimeoutError as e:
                # Track consecutive timeouts — likely rate-limited/blocked
                consecutive_timeouts += 1
                print(f"    [W{worker_id}] Timeout #{consecutive_timeouts}: {str(e)[:80]}")
                if consecutive_timeouts >= 3:
                    print(f"    [W{worker_id}] {consecutive_timeouts} consecutive timeouts — closing and recreating page")
                    try:
                        await page.close()
                    except Exception:
                        pass
                    page = await ctx.new_page()
                    consecutive_timeouts = 0
                await throttle.wait()

            except Exception as e:
                consecutive_timeouts = 0  # reset on non-timeout error
                if 'pipe' in str(e).lower() or 'connection' in str(e).lower():
                    crash_count += 1
                    print(f"    [W{worker_id}] Pipe error (crash {crash_count}/20): {str(e)[:80]}")
                    if crash_count >= 20:
                        raise BrowserCrashed(f"Worker {worker_id} pipe errors exceeded")
                    try:
                        await page.close()
                    except Exception:
                        pass
                    page = await ctx.new_page()
                else:
                    print(f"    [W{worker_id}] Error: {str(e)[:100]}")

    finally:
        try:
            await page.close()
        except Exception:
            pass


# ── Main ────────────────────────────────────────────────────────────────────

async def main():
    working_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(working_dir, "oldnavy.csv")

    if not os.path.exists(csv_path):
        print(f"ERROR: {csv_path} not found")
        return

    print(f"\n=== Old Navy PDP Scraper ===")
    print(f"Working directory: {working_dir}")
    print(f"Input: {csv_path}")

    # Load URLs
    urls = []
    try:
        with open(csv_path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = row.get('plp_product-image href', '').strip()
                if url and url.startswith('http'):
                    urls.append(url)
    except Exception as e:
        print(f"ERROR reading CSV: {e}")
        return

    print(f"Loaded {len(urls)} URLs")

    # Load progress
    progress = load_progress(working_dir)
    print(f"Resume: {len(progress['processed'])} already processed")

    # Filter out already processed
    todo_urls = [u for u in urls if u not in progress['processed']]
    print(f"Remaining: {len(todo_urls)} to process")

    if not todo_urls:
        print("All URLs already processed!")
        return

    stats = Stats()
    stats.total = len(todo_urls)
    throttle = AdaptiveThrottle()
    results = progress['results']

    batch_start = time.time()
    crash_restart_count = 0

    try:
        while todo_urls and crash_restart_count < 20:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True, args=[
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                    "--no-sandbox",
                ])
                try:
                    # Create contexts
                    contexts = []
                    for i in range(NUM_CONTEXTS):
                        ctx = await setup_context(browser, i)
                        contexts.append(ctx)

                    queue = asyncio.Queue()
                    for url in todo_urls:
                        await queue.put(url)

                    print(f"\n[Batch #{crash_restart_count + 1}] {queue.qsize()} URLs, {NUM_CONTEXTS} contexts, {TABS_PER_CTX} tab(s) each")

                    # Start workers
                    tasks = []
                    for worker_id in range(NUM_CONTEXTS):
                        t = asyncio.create_task(
                            worker(worker_id, contexts[worker_id], queue, results, stats, throttle, working_dir)
                        )
                        tasks.append(t)

                    # Auto-save every SAVE_INTERVAL seconds
                    last_save = time.time()

                    # Wait for tasks with periodic monitoring
                    while True:
                        try:
                            done, pending = await asyncio.wait(
                                tasks, timeout=30, return_when=asyncio.ALL_COMPLETED
                            )
                            # Check for exceptions in completed tasks
                            for t in done:
                                try:
                                    t.result()
                                except Exception as e:
                                    print(f"   [WORKER ERROR] {type(e).__name__}: {e}")
                            if not pending:
                                break
                        except (asyncio.CancelledError, Exception) as e:
                            print(f"   [WAIT ERROR] {type(e).__name__}: {e}")
                            break

                        # Periodic auto-save
                        if time.time() - last_save > SAVE_INTERVAL:
                            processed_now = len(progress['processed'])
                            for row in results:
                                if row.get('url') not in progress['processed']:
                                    progress['processed'].add(row['url'])
                            save_progress(progress, working_dir)
                            save_to_excel(results, working_dir)
                            new_count = len(progress['processed'])
                            if new_count > processed_now:
                                print(f"   [AUTO-SAVE] {new_count} processed, {stats.rows_generated} rows, {stats.failed} failed, {stats.blocked} blocked")
                            last_save = time.time()

                        # Check batch time limit (20 minutes)
                        if time.time() - batch_start > BATCH_TIME_LIMIT:
                            print(f"   [BATCH TIMEOUT] Restarting browser after {BATCH_TIME_LIMIT // 60:.0f} min")
                            await asyncio.gather(*tasks, return_exceptions=True)
                            break

                    # Update progress from results
                    for row in results:
                        if row.get('url') and row['url'] not in progress['processed']:
                            progress['processed'].add(row['url'])

                    # Filter todo_urls to remaining
                    todo_urls = [u for u in todo_urls if u not in progress['processed']]

                    # Close contexts
                    for ctx in contexts:
                        await ctx.close()

                except (BrowserCrashed, Exception) as be:
                    print(f"   [CRASH] {type(be).__name__}: {be}")
                    # Save whatever we have so far
                    for row in results:
                        if row.get('url') and row['url'] not in progress['processed']:
                            progress['processed'].add(row['url'])
                    save_progress(progress, working_dir)
                    save_to_excel(results, working_dir)
                    todo_urls = [u for u in todo_urls if u not in progress['processed']]
                    crash_restart_count += 1
                    throttle.reset()
                    batch_start = time.time()
                    print(f"   [RESTART] Attempt {crash_restart_count}/20, {len(todo_urls)} URLs remaining")

                finally:
                    try:
                        await browser.close()
                    except Exception:
                        pass

    except KeyboardInterrupt:
        print("\n[KEYBOARD INTERRUPT] Saving progress...")

    # Final save
    for row in results:
        if row.get('url') and row['url'] not in progress['processed']:
            progress['processed'].add(row['url'])
    save_progress(progress, working_dir)
    save_to_excel(results, working_dir)

    print(f"\n=== FINAL STATS ===")
    print(f"Processed: {len(progress['processed'])} / {stats.total}")
    print(f"Rows generated: {stats.rows_generated}")
    print(f"Failed: {stats.failed}")
    print(f"Blocked: {stats.blocked}")
    print(f"Results saved to: {os.path.join(working_dir, 'oldnavy_pdp_results.xlsx')}")


if __name__ == '__main__':
    asyncio.run(main())
