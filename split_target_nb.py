#!/usr/bin/env python3
"""
split_target_nb.py — Split the dashboard's "Target NB" group into:
  - Target NB   (1P, sold by Target)
  - Target 3P   (marketplace, everything else previously Target NB)

Target OB is NOT touched.

Matching strategy (BRAND-based)
-------------------------------
`Sold by Target.csv` lists products sold by Target. Pulling distinct brands
from that CSV gives the canonical "Target sells these themselves" set. The
RAW array has a `b` (brand) field on every entry, so brand-membership is
the natural classifier — and it's robust to the CSV having only a partial
list of products per brand (e.g., 14 Levi's URLs in the CSV when our scrape
has 59 Levi's colorways: all 59 should be NB because Levi's at Target is
never 3P).

Procedure:
  1. Extract distinct brands from `Sold by Target.csv`.
  2. Drop any brand that is a Target Owned Brand (already in Target OB and
     shouldn't be NB regardless). Detected via target_pdp_results.xlsx
     "Owned Brand" flag.
  3. For each Target NB entry in RAW, if its brand is in the resulting
     1P-NB set -> stays "Target NB". Else -> becomes "Target 3P".

Patches index.html in place — adds Target 3P to GROUPS / GROUP_LABELS / GC,
inserts a KPI tile, retags affected RAW entries. Idempotent.

Usage:
  python3 split_target_nb.py            # do the split
  python3 split_target_nb.py --dry-run  # preview
  python3 split_target_nb.py --force    # rebuild (e.g. after CSV update)
"""

import argparse
import csv
import json
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
SOLD_CSV = os.path.join(HERE, "Sold by Target.csv")
PDP_XLSX = os.path.join(HERE, "target_pdp_results.xlsx")
INDEX_HTML = os.path.join(HERE, "index.html")


# ── Inputs ─────────────────────────────────────────────────────────────────

# Brand column in Sold by Target.csv. The CSS-class header makes this fragile,
# so we tolerate variants:
SOLD_BRAND_COLS = ("styles_ndsLink__GUaai", "Brand", "brand")


def load_sold_brands():
    """Return Counter of distinct brands found in Sold by Target.csv."""
    brands = Counter()
    with open(SOLD_CSV, "r", encoding="utf-8-sig") as f:
        rdr = csv.DictReader(f)
        # Pick the first available brand column
        col = next((c for c in SOLD_BRAND_COLS if c in (rdr.fieldnames or [])), None)
        if col is None:
            raise RuntimeError(
                f"Sold by Target.csv: none of {SOLD_BRAND_COLS} present "
                f"in headers {rdr.fieldnames}"
            )
        for r in rdr:
            b = (r.get(col) or "").strip()
            if b:
                brands[b] += 1
    return brands


def load_target_owned_brands():
    """Return set of brand names flagged as Target Owned Brand in xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(PDP_XLSX, read_only=True)
    ws = wb.active
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdrs) if h}
    if "Brand" not in ix or "Owned Brand" not in ix:
        wb.close()
        return set()
    ob_brands = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ob_val = row[ix["Owned Brand"]]
        is_ob = bool(ob_val) and str(ob_val).strip().lower() in ("true", "yes", "1", "y")
        if is_ob:
            b = str(row[ix["Brand"]] or "").strip()
            if b:
                ob_brands.add(b)
    wb.close()
    return ob_brands


# ── HTML patchers ──────────────────────────────────────────────────────────

def patch_groups(html, force=False):
    """Insert 'Target 3P' right after 'Target NB' in the GROUPS array."""
    if "'Target 3P'" in html and not force:
        print("  GROUPS already contains 'Target 3P' (skip)")
        return html
    if force:
        html = re.sub(r"'Target 3P',\s*", "", html, count=1)
    # Tolerant: handles 'Target NB' followed by anything
    pat = re.compile(r"(var GROUPS = \[[^\]]*'Target NB',)")
    new = pat.sub(r"\1'Target 3P',", html, count=1)
    if new == html:
        raise RuntimeError("Could not patch GROUPS array")
    print("  GROUPS array patched (Target NB, Target 3P, ...)")
    return new


def patch_group_labels(html):
    if "'Target 3P':" in html:
        print("  GROUP_LABELS already contains Target 3P (skip)")
        return html
    new = html.replace(
        "'Target NB':'Target National Brands',",
        "'Target NB':'Target National Brands','Target 3P':'Target 3P (Marketplace)',",
        1,
    )
    if new == html:
        raise RuntimeError("Could not patch GROUP_LABELS dict")
    print("  GROUP_LABELS dict patched")
    return new


def patch_gc(html):
    """Add Target 3P color — pinkish-red to sit between Target NB red and a
    distinct hue."""
    if re.search(r"'Target 3P':\s*\{bg:", html):
        print("  GC already contains Target 3P (skip)")
        return html
    pat = re.compile(
        r"('Target NB':\s*\{bg:'#FF4444', light:'rgba\(255,68,68,0\.18\)', border:'#FF4444'\},)"
    )
    repl = (
        r"\1\n  'Target 3P':  {bg:'#FF99AA', light:'rgba(255,153,170,0.22)', border:'#FF99AA'},"
    )
    new = pat.sub(repl, html, count=1)
    if new == html:
        raise RuntimeError("Could not patch GC color dict")
    print("  GC color config patched (Target 3P = #FF99AA)")
    return new


def patch_kpi_tiles(html, nb_count, p3_count):
    """Update Target NB KPI tile count and insert Target 3P tile right after."""
    # 1. Update Target NB count
    new = re.sub(
        r'(font-weight:700">Target NB</div><div[^>]*>)\d+(</div>)',
        lambda m: m.group(1) + str(nb_count) + m.group(2),
        html, count=1,
    )
    if new == html:
        print("  WARNING: Could not update Target NB count tile")

    # 2. If Target 3P tile already present, update its count and stop
    if 'font-weight:700">Target 3P' in new:
        new = re.sub(
            r'(font-weight:700">Target 3P</div><div[^>]*>)\d+(</div>)',
            lambda m: m.group(1) + str(p3_count) + m.group(2),
            new, count=1,
        )
        print(f"  KPI tiles updated (Target NB={nb_count}, Target 3P={p3_count})")
        return new

    # 3. Insert Target 3P tile right after Target NB tile
    nb_tile_pat = re.compile(
        r'(<div style="background:var\(--bg\);border:1px solid var\(--bg3\);'
        r'border-radius:var\(--radius-sm\);padding:10px 16px;min-width:100px;'
        r'text-align:center"><div style="font-size:\.65rem;color:var\(--fg3\);'
        r'letter-spacing:\.05em;text-transform:uppercase;font-weight:700">'
        r'Target NB</div><div style="font-size:1\.3rem;font-weight:800;'
        r'color:var\(--fg\)">\d+</div><div style="font-size:\.6rem;'
        r'color:var\(--fg3\)">CCs</div></div>)'
    )
    p3_tile = (
        '<div style="background:var(--bg);border:1px solid var(--bg3);'
        'border-radius:var(--radius-sm);padding:10px 16px;min-width:100px;'
        'text-align:center"><div style="font-size:.65rem;color:var(--fg3);'
        'letter-spacing:.05em;text-transform:uppercase;font-weight:700">'
        'Target 3P</div><div style="font-size:1.3rem;font-weight:800;'
        'color:var(--fg)">' + str(p3_count) + '</div><div style="font-size:.6rem;'
        'color:var(--fg3)">CCs</div></div>'
    )
    n2 = nb_tile_pat.sub(lambda m: m.group(1) + "\n" + p3_tile, new, count=1)
    if n2 == new:
        print("  WARNING: Could not find Target NB tile to insert after — skipping tile")
        return new
    print(f"  KPI tiles: Target NB={nb_count}, Target 3P={p3_count} (inserted)")
    return n2


def split_raw(html, nb_brands, force=False):
    """Retag every Target NB entry in RAW based on brand membership.
    Brands listed in `nb_brands` keep g='Target NB'; everything else
    previously tagged Target NB becomes g='Target 3P'.
    """
    raw_start = html.find("var RAW = [")
    if raw_start < 0:
        raise RuntimeError("Could not find 'var RAW = [' in index.html")
    arr_open = raw_start + len("var RAW = ")
    arr_end = html.find("];", arr_open)
    if arr_end < 0:
        raise RuntimeError("Could not find end of RAW array")
    data = json.loads(html[arr_open:arr_end + 1])

    # If --force, restore any prior Target 3P -> Target NB so we can re-split
    # cleanly (idempotent re-runs).
    if force:
        moved = 0
        for r in data:
            if r.get("g") == "Target 3P":
                r["g"] = "Target NB"
                moved += 1
        if moved:
            print(f"  Restored {moved} Target 3P entries back to Target NB (force)")

    # Case-insensitive brand membership for resilience
    nb_brands_lc = {b.lower() for b in nb_brands}

    nb_count = 0
    p3_count = 0
    nb_per_brand = Counter()
    p3_top_brands = Counter()
    for r in data:
        if r.get("g") != "Target NB":
            continue
        b = (r.get("b") or "").strip()
        if b.lower() in nb_brands_lc:
            nb_count += 1
            nb_per_brand[b] += 1
        else:
            r["g"] = "Target 3P"
            p3_count += 1
            p3_top_brands[b] += 1

    new_arr = json.dumps(data, separators=(",", ":"))
    new_html = html[:arr_open] + new_arr + html[arr_end + 1:]
    print(f"  RAW retagged: Target NB stays={nb_count}, Target 3P new={p3_count}")
    if nb_per_brand:
        print(f"  Target NB by brand:")
        for b, n in nb_per_brand.most_common():
            print(f"    {n:>3}  {b}")
    if p3_top_brands:
        print(f"  Target 3P top brands:")
        for b, n in p3_top_brands.most_common(10):
            print(f"    {n:>3}  {b}")
    return new_html, nb_count, p3_count


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true",
                    help="rebuild Target 3P split fresh (e.g. after CSV update)")
    args = ap.parse_args()

    print(f"Splitting Target NB -> Target NB (1P) + Target 3P  "
          f"(force={args.force}, dry_run={args.dry_run})\n")

    if not os.path.exists(SOLD_CSV):
        print(f"ERROR: missing {SOLD_CSV}")
        return 1
    if not os.path.exists(PDP_XLSX):
        print(f"ERROR: missing {PDP_XLSX}")
        return 1

    sold_brands = load_sold_brands()
    print(f"  Brands in Sold by Target.csv:   {len(sold_brands)}")
    for b, n in sold_brands.most_common():
        print(f"    {n:>3}  {b}")

    ob_brands = load_target_owned_brands()
    print(f"\n  Target Owned Brands (excluded): {sorted(ob_brands)}")

    nb_brands = set(sold_brands.keys()) - ob_brands
    print(f"\n  Target NB (1P) brand set:       {sorted(nb_brands)}")

    if not nb_brands:
        print("\nERROR: After excluding Owned Brands, no NB brands remain.")
        return 1

    print(f"\nLoading {INDEX_HTML} ...")
    with open(INDEX_HTML, "r", encoding="utf-8") as f:
        html = f.read()
    orig_size = len(html)

    new_html = patch_groups(html, force=args.force)
    new_html = patch_group_labels(new_html)
    new_html = patch_gc(new_html)
    new_html, nb_count, p3_count = split_raw(new_html, nb_brands,
                                             force=args.force)
    new_html = patch_kpi_tiles(new_html, nb_count, p3_count)

    if args.dry_run:
        print(f"\n[dry-run] would write {len(new_html):,} chars (was {orig_size:,})")
        return 0

    backup = INDEX_HTML + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(INDEX_HTML, backup)
    print(f"\nBackup written: {backup}")

    with open(INDEX_HTML, "w", encoding="utf-8") as f:
        f.write(new_html)
    print(f"index.html updated: {orig_size:,} -> {len(new_html):,} chars")
    print(f"Final: Target NB={nb_count}, Target 3P={p3_count}, "
          f"Target OB unchanged")
    return 0


if __name__ == "__main__":
    sys.exit(main())
