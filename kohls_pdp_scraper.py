#!/usr/bin/env python3
"""
Production Kohl's PDP Scraper — 1 Row per Product × Color
===========================================================
Input:  kohls.csv  (URLs in first column 'flex href')
Output: kohls_pdp_results.xlsx + kohls_pdp_progress.json

Architecture: 5 browser contexts × 1 tab each = 5 parallel workers
              20-minute auto-restart for fresh sessions
              Adaptive throttle, crash recovery, progress resume

Extracts from Kohl's PDP pages via Playwright evaluate() to produce
one row per color variant with pricing, specs, ratings, and fabric.
"""

import asyncio, json, os, random, re, time
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Price Extraction JS ──────────────────────────────────────────────────────
KOHLS_PRICE_JS = """() => {
    var el = document.querySelector('[data-testid="product-price"]') ||
             document.querySelector('.product-price') ||
             document.querySelector('[class*="Price"]');
    if (el) {
        var text = el.textContent.trim();
        var m = text.match(/\\$(\\d+\\.?\\d*)/);
        if (m) return parseFloat(m[1]);
    }
    return null;
}"""

# ── Configuration ───────────────────────────────────────────────────────────
NUM_CONTEXTS = 1
TABS_PER_CTX = 1
CONCURRENCY = NUM_CONTEXTS * TABS_PER_CTX  # 1 worker (was 5) — gentler on Akamai
SAVE_INTERVAL = 120          # auto-save every 2 minutes
MAX_RETRIES = 3
BATCH_TIME_LIMIT = 20 * 60  # 20 minutes — restart browser with fresh session


class BrowserCrashed(Exception):
    """Raised when pipe/connection errors persist — signals browser is dead."""
    pass


USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# ── Kohl's Owned Brands ────────────────────────────────────────────────────
KOHLS_OWNED_BRANDS = {
    'sonoma goods for life', 'lc lauren conrad', 'apt. 9',
    'simply vera vera wang', 'croft & barrow', 'jumping beans',
    'tek gear', 'so', 'urban pipeline',
}


def classify_brand(brand_name):
    """Returns 'OB' if Kohl's owned brand, 'NB' for national brand."""
    if brand_name and brand_name.strip().lower() in KOHLS_OWNED_BRANDS:
        return 'OB'
    return 'NB'


# ── Non-Basic Detection ───────────────────────────────────────────────────
NON_BASIC_KEYWORDS = [
    'print', 'printed', 'graphic', 'pattern', 'floral', 'stripe', 'striped',
    'plaid', 'camo', 'camouflage', 'tie-dye', 'tie dye', 'leopard', 'logo',
    'embroidered', 'embroidery', 'novelty', 'character', 'varsity', 'retro',
    'animal', 'abstract', 'geometric', 'paisley', 'polka', 'check', 'checked',
    'tropical', 'botanical', 'ditsy', 'patchwork', 'colorblock', 'ombre',
    'destroy', 'distressed', 'destructed',
]


def is_non_basic(title, colors_str, bullets_str):
    """Returns 1 if product is non-basic (graphic/patterned), 0 otherwise."""
    combined = f"{title} {colors_str} {bullets_str}".lower()
    for kw in NON_BASIC_KEYWORDS:
        if kw in combined:
            return 1
    return 0


# ── Rise / Leg / Fit Parsing from Title ────────────────────────────────────

def parse_rise(title):
    """Extract rise from product title."""
    t = title.lower()
    if 'super high' in t or 'super-high' in t:
        return 'Super High Rise'
    if 'high-rise' in t or 'high rise' in t or 'high' in t and 'rise' in t:
        return 'High Rise'
    if 'mid-rise' in t or 'mid rise' in t or 'mid' in t and 'rise' in t:
        return 'Mid Rise'
    if 'low-rise' in t or 'low rise' in t or 'low' in t and 'rise' in t:
        return 'Low Rise'
    return ''


def parse_leg_shape(title):
    """Extract leg shape from product title."""
    t = title.lower()
    if 'straight' in t:
        return 'Straight Leg'
    if 'skinny' in t:
        return 'Skinny'
    if 'slim' in t:
        return 'Slim'
    if 'bootcut' in t or 'boot cut' in t or 'boot-cut' in t:
        return 'Bootcut'
    if 'flare' in t:
        return 'Flare'
    if 'wide' in t and 'leg' in t:
        return 'Wide Leg'
    if 'boyfriend' in t:
        return 'Boyfriend'
    return ''


def parse_fit(title):
    """Extract fit descriptor from product title."""
    t = title.lower()
    if 'slim' in t:
        return 'Slim'
    if 'skinny' in t:
        return 'Skinny'
    if 'fitted' in t:
        return 'Fitted'
    if 'relaxed' in t:
        return 'Relaxed'
    if 'straight' in t:
        return 'Straight'
    return ''


def parse_length_hit(size_and_fit_text, title_text):
    """
    Extract length hit descriptor (e.g., 'hits at ankle', 'hits above knee') from
    size & fit section or product title.
    """
    if not size_and_fit_text and not title_text:
        return ''

    combined = f"{size_and_fit_text} {title_text}".lower()

    # Common length hit patterns
    patterns = [
        'hits at ankle', 'hit at ankle',
        'hits ankle', 'ankle length',
        'hits at knee', 'hit at knee',
        'hits knee', 'knee length',
        'hits above knee', 'hit above knee',
        'hits mid-calf', 'hit mid-calf',
        'mid-calf', 'midcalf',
        'hits at calf', 'hit at calf',
        'hits calf', 'calf length',
        'hits at thigh', 'hit at thigh',
        'hits thigh', 'thigh length',
        'full length', 'full-length',
        'cropped', 'crop length',
    ]

    for pattern in patterns:
        if pattern in combined:
            return pattern.title()

    return ''


def parse_inseam(size_and_fit_text, product_name):
    """
    Extract inseam measurement (e.g., '32 inch', '28"', '32in') from
    size & fit section or product name.
    """
    if not size_and_fit_text and not product_name:
        return ''

    combined = f"{size_and_fit_text} {product_name}"

    # Match inseam patterns: "32 inch", "32\"", "32in", "32\" inseam", etc.
    patterns = [
        r'(\d+)\s*(?:inch|")\s*(?:inseam)?',
        r'inseam[:\s]+(\d+)\s*(?:inch|")?',
    ]

    for pattern in patterns:
        m = re.search(pattern, combined, re.IGNORECASE)
        if m:
            inches = m.group(1)
            return f"{inches}\""

    return ''


def parse_material(fabric_raw):
    """
    Extract fabric composition from raw fabric string.
    Returns (fabric_parsed, pct_cotton, pct_natural_fiber).
    """
    if not fabric_raw:
        return '', '0%', '0%'

    # Extract percentage composition
    fabric_parsed = ''
    m = re.search(r'(\d+\s*%\s*\w+(?:\s*,\s*\d+\s*%\s*[\w\s-]+)*)', fabric_raw, re.IGNORECASE)
    if m:
        fabric_parsed = m.group(1).strip()
    else:
        fabric_parsed = fabric_raw[:150]

    # Calculate cotton percentage
    cotton_pct = '0%'
    cm = re.search(r'(\d+)\s*%\s*(?:Recycled\s+)?Cotton', fabric_raw, re.IGNORECASE)
    if cm:
        cotton_pct = f"{cm.group(1)}%"

    # Calculate natural fiber percentage
    natural_pct = '0%'
    ns = re.findall(r'(\d+)\s*%\s*(?:Recycled\s+)?(?:Cotton|Wool|Silk|Linen|Hemp|Lyocell|Tencel)', fabric_raw, re.IGNORECASE)
    if ns:
        total = sum(int(n) for n in ns)
        natural_pct = f"{total}%"

    return fabric_parsed, cotton_pct, natural_pct


# ── Adaptive Throttle ───────────────────────────────────────────────────────

class AdaptiveThrottle:
    """Shared throttle — when ANY worker hits a block, ALL workers slow down."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.base_delay = 6.0      # was 2.0 — gentler polite-delay
        self.current_delay = 6.0   # was 2.0
        self.max_delay = 60.0      # was 30.0 — allow longer backoff on blocks
        self.consecutive_ok = 0
        self.cooldown_until = 0

    async def on_block(self):
        async with self.lock:
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            self.cooldown_until = time.time() + self.current_delay * 3
            self.consecutive_ok = 0
            print(f"    🐌 Throttle UP: delay now {self.current_delay:.1f}s, all workers pausing {self.current_delay * 3:.0f}s")

    async def on_ok(self):
        async with self.lock:
            self.consecutive_ok += 1
            if self.consecutive_ok > 20 and self.current_delay > self.base_delay:
                self.current_delay = max(self.current_delay * 0.8, self.base_delay)
                self.consecutive_ok = 0
                print(f"    🏎️  Throttle DOWN: delay now {self.current_delay:.1f}s")

    async def wait(self):
        now = time.time()
        async with self.lock:
            cd = self.cooldown_until
            delay = self.current_delay
        if now < cd:
            await asyncio.sleep(cd - now)
        await asyncio.sleep(random.uniform(delay, delay * 2))

    def reset(self):
        self.current_delay = self.base_delay
        self.consecutive_ok = 0
        self.cooldown_until = 0
        print(f"    🔄 Throttle RESET to {self.base_delay}s")


# ── Stats ───────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self):
        self.lock = asyncio.Lock()
        self.total = self.completed = self.failed = self.blocked = 0
        self.rows_generated = 0

    async def inc_ok(self, row_count=1):
        async with self.lock:
            self.completed += 1
            self.rows_generated += row_count

    async def inc_fail(self):
        async with self.lock:
            self.failed += 1

    async def inc_blocked(self):
        async with self.lock:
            self.blocked += 1


# ── Progress / Excel ────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'URL', 'Product Name', 'Brand', 'Brand Type', 'Color',
    'Current Price', 'Current Price Formatted', 'Original Price', 'Original Price Formatted',
    'On Sale', 'Discount %',
    'Rise', 'Leg Shape', 'Fit',
    'Fabric Raw', 'Fabric Parsed', '% Cotton', '% Natural Fiber',
    'Non-Basic',
    'Breadcrumb', 'Department',
    'Total Colors', 'Total Sizes', 'Sizes List',
    'Average Rating', 'Review Count',
    'Image Count',
    'Feature Bullets', 'Product Details Raw', 'Page Text', 'Size and Fit',
    'Length Hit', 'Inseam',
    'Retries', 'Timestamp', 'Error',
]

EXCEL_FIELDS = [
    'url', 'product_name', 'brand', 'brand_type', 'color',
    'current_price', 'current_price_formatted', 'original_price', 'original_price_formatted',
    'on_sale', 'discount_pct',
    'rise', 'leg_shape', 'fit',
    'fabric_raw', 'fabric_parsed', 'pct_cotton', 'pct_natural_fiber',
    'non_basic',
    'breadcrumb', 'department',
    'total_colors', 'total_sizes', 'sizes_list',
    'average_rating', 'review_count',
    'image_count',
    'feature_bullets', 'product_details_raw', 'page_text', 'size_and_fit',
    'length_hit', 'inseam',
    'retries', 'timestamp', 'error',
]


def load_progress(d):
    p = os.path.join(d, "kohls_pdp_progress.json")
    processed = set()
    if os.path.exists(p):
        with open(p) as f:
            data = json.load(f)
        processed = set(data.get('processed', []))

    results = []
    xlsx_path = os.path.join(d, "kohls_pdp_results.xlsx")
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook as lwb
            wb = lwb(xlsx_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]]
            field_map = dict(zip(EXCEL_HEADERS, EXCEL_FIELDS))
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i] in field_map:
                        rec[field_map[headers[i]]] = val if val is not None else ''
                if rec.get('url'):
                    results.append(rec)
            wb.close()
            print(f"   📂 Reloaded {len(results)} previous result rows from Excel")
        except Exception as e:
            print(f"   ⚠️  Could not reload Excel results: {e}")
    return {'processed': processed, 'results': results}


def save_progress(progress, d):
    with open(os.path.join(d, "kohls_pdp_progress.json"), 'w') as f:
        json.dump({
            'processed': list(progress['processed']),
            'last_save': datetime.now().isoformat(),
            'total_processed': len(progress['processed']),
        }, f)


def save_to_excel(results, d):
    out = os.path.join(d, "kohls_pdp_results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Kohls PDP Results'

    hfill = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, data in enumerate(results, 2):
        for col, field in enumerate(EXCEL_FIELDS, 1):
            val = data.get(field, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            ws.cell(row=i, column=col, value=val)

    ws.freeze_panes = 'A2'
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    wb.save(out)
    return out


# ── Browser Setup ───────────────────────────────────────────────────────────

async def setup_context(browser, ua_index=0):
    """Create a browser context. Works with both Chromium and Firefox."""
    ctx = await browser.new_context(
        viewport={'width': 1920, 'height': 1080},
        locale="en-US",
    )
    # Only override webdriver for Chromium (Firefox doesn't expose it the same way)
    try:
        await ctx.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
    except Exception:
        pass
    return ctx


# ── Kohl's Data Extraction ──────────────────────────────────────────────────

async def extract_kohls_data(page):
    """
    Extract product data from rendered Kohl's page using JavaScript.
    Returns a dict with product info, colors, sizes, pricing, etc.
    """
    try:
        data = await page.evaluate('''() => {
            const result = {
                title: '',
                brand: '',
                colors: [],
                sizes: [],
                current_price: '',
                original_price: '',
                rating: '',
                review_count: '',
                breadcrumb: '',
                image_count: 0,
                product_details: '',
                feature_bullets: '',
                fabric_raw: '',
                page_text: '',
                size_and_fit: '',
            };

            // Product title — usually h1 or .pdp-title
            const titleEl = document.querySelector('h1') || document.querySelector('[data-testid="pdp-title"]') || document.querySelector('.product-title');
            if (titleEl) {
                result.title = titleEl.textContent.trim();
            }

            // Brand name — look for brand link or text near title
            const brandEl = document.querySelector('[data-testid="brand-name"]') ||
                          document.querySelector('.brand-name') ||
                          document.querySelector('a[href*="/brand/"]');
            if (brandEl) {
                result.brand = brandEl.textContent.trim();
            }

            // Current and original prices
            const priceEl = document.querySelector('[data-testid="product-price"]') ||
                          document.querySelector('.product-price') ||
                          document.querySelector('[class*="Price"]');
            if (priceEl) {
                result.current_price = priceEl.textContent.trim();
            }

            const originalPriceEl = document.querySelector('[data-testid="original-price"]') ||
                                   document.querySelector('.original-price') ||
                                   document.querySelector('[class*="original"]');
            if (originalPriceEl) {
                result.original_price = originalPriceEl.textContent.trim();
            }

            // Color swatches — clickable color options
            const colorSwatches = document.querySelectorAll('[data-testid*="color"]') ||
                                 document.querySelectorAll('.color-swatch') ||
                                 document.querySelectorAll('[class*="ColorSwatch"]');
            colorSwatches.forEach(swatch => {
                const colorName = swatch.getAttribute('aria-label') ||
                                swatch.getAttribute('title') ||
                                swatch.textContent.trim();
                if (colorName && !result.colors.includes(colorName)) {
                    result.colors.push(colorName);
                }
            });

            // Sizes — size selector buttons
            const sizeButtons = document.querySelectorAll('[data-testid*="size"]') ||
                               document.querySelectorAll('.size-button') ||
                               document.querySelectorAll('[class*="SizeButton"]');
            sizeButtons.forEach(btn => {
                const sizeText = btn.textContent.trim();
                if (sizeText && !result.sizes.includes(sizeText)) {
                    result.sizes.push(sizeText);
                }
            });

            // Rating and review count
            const ratingEl = document.querySelector('[data-testid="rating"]') ||
                            document.querySelector('.rating') ||
                            document.querySelector('[class*="Rating"]');
            if (ratingEl) {
                result.rating = ratingEl.textContent.trim();
            }

            const reviewEl = document.querySelector('[data-testid="review-count"]') ||
                            document.querySelector('.review-count') ||
                            document.querySelector('[class*="ReviewCount"]');
            if (reviewEl) {
                result.review_count = reviewEl.textContent.trim();
            }

            // Breadcrumb navigation
            const breadcrumbEl = document.querySelector('[data-testid="breadcrumb"]') ||
                                document.querySelector('.breadcrumb') ||
                                document.querySelector('nav[aria-label*="breadcrumb"]');
            if (breadcrumbEl) {
                const items = breadcrumbEl.querySelectorAll('a, span');
                result.breadcrumb = Array.from(items).map(el => el.textContent.trim()).join(' > ');
            }

            // Image count — count product images
            const images = document.querySelectorAll('[data-testid*="product-image"]') ||
                          document.querySelectorAll('.product-image') ||
                          document.querySelectorAll('[class*="ProductImage"]');
            result.image_count = images.length || 1;

            // Product details / specs — usually in a table or accordion
            const detailsEl = document.querySelector('[data-testid="product-details"]') ||
                             document.querySelector('.product-details') ||
                             document.querySelector('[class*="ProductDetails"]');
            if (detailsEl) {
                result.product_details = detailsEl.textContent.trim().substring(0, 500);
            }

            // Feature bullets
            const bulletsEl = document.querySelector('[data-testid="feature-bullets"]') ||
                             document.querySelector('.feature-bullets') ||
                             document.querySelector('[class*="Features"]');
            if (bulletsEl) {
                result.feature_bullets = bulletsEl.textContent.trim().substring(0, 500);
            }

            // Fabric/material composition
            const materialEl = document.querySelector('[data-testid*="material"]') ||
                              document.querySelector('.material') ||
                              document.querySelector('[class*="Material"]');
            if (materialEl) {
                result.fabric_raw = materialEl.textContent.trim();
            }

            // Page text — capture full page text up to 10000 chars for keyword searching
            const bodyEl = document.querySelector('body');
            if (bodyEl) {
                result.page_text = bodyEl.textContent.trim().substring(0, 10000);
            }

            // Size & Fit section — look for dedicated section or accordion
            const sizeAndFitEl = document.querySelector('[data-testid*="size-fit"]') ||
                                document.querySelector('[class*="SizeFit"]') ||
                                document.querySelector('[id*="size"]') ||
                                document.querySelector('[class*="SizeGuide"]');
            if (sizeAndFitEl) {
                result.size_and_fit = sizeAndFitEl.textContent.trim().substring(0, 1000);
            } else {
                // Fallback: search for size & fit section in product details
                const detailsText = (detailsEl ? detailsEl.textContent : '') + (bulletsEl ? bulletsEl.textContent : '');
                const match = detailsText.match(/(?:size\s*[&]?\s*fit|sizing|measurements|fit[^a-z]*:)(.*?)(?=\n\n|$)/is);
                if (match) {
                    result.size_and_fit = match[1].trim().substring(0, 1000);
                }
            }

            return result;
        }''')
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"    JS evaluation error: {str(e)[:100]}")
        return {}


def parse_price_str(price_str):
    """Parse a price string like '$69.99' into float."""
    if not price_str:
        return None
    m = re.search(r'\$?([\d,.]+)', str(price_str))
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            return None
    return None


def parse_kohls_pdp(page_html, url, js_data=None):
    """
    Parse Kohl's PDP and return a LIST of result dicts — one per color.
    Falls back to regex parsing if JS extraction doesn't work.
    """
    rows = []

    try:
        # Initialize parent-level data
        parent = {
            'url': url,
            'product_name': '',
            'brand': '',
            'brand_type': '',
            'breadcrumb': '',
            'department': '',
            'total_colors': 0,
            'total_sizes': 0,
            'sizes_list': '',
            'current_price': '',
            'current_price_formatted': '',
            'original_price': '',
            'original_price_formatted': '',
            'on_sale': False,
            'discount_pct': '',
            'average_rating': '',
            'review_count': '',
            'image_count': 0,
            'feature_bullets': '',
            'product_details_raw': '',
            'page_text': '',
            'size_and_fit': '',
            'fabric_raw': '',
            'fabric_parsed': '',
            'pct_cotton': '0%',
            'pct_natural_fiber': '0%',
            'rise': '',
            'leg_shape': '',
            'fit': '',
            'length_hit': '',
            'inseam': '',
            'non_basic': 0,
            'retries': 0,
            'timestamp': datetime.now().isoformat(),
            'error': '',
        }

        # Use JS-extracted data if available, otherwise fall back to regex
        if js_data:
            parent['product_name'] = js_data.get('title', '')
            parent['brand'] = js_data.get('brand', '')
            parent['current_price_formatted'] = js_data.get('current_price', '')
            parent['original_price_formatted'] = js_data.get('original_price', '')
            parent['breadcrumb'] = js_data.get('breadcrumb', '')
            parent['image_count'] = js_data.get('image_count', 0)
            parent['feature_bullets'] = js_data.get('feature_bullets', '')
            parent['product_details_raw'] = js_data.get('product_details', '')
            parent['page_text'] = js_data.get('page_text', '')
            parent['size_and_fit'] = js_data.get('size_and_fit', '')
            parent['fabric_raw'] = js_data.get('fabric_raw', '')
            parent['average_rating'] = js_data.get('rating', '')
            parent['review_count'] = js_data.get('review_count', '')

            # Extract numeric price
            price_num = parse_price_str(parent['current_price_formatted'])
            if price_num:
                parent['current_price'] = price_num

            # Extract numeric original price
            orig_num = parse_price_str(parent['original_price_formatted'])
            if orig_num:
                parent['original_price'] = orig_num

            # Calculate on_sale and discount
            if orig_num and price_num and orig_num > price_num:
                parent['on_sale'] = True
                discount = ((orig_num - price_num) / orig_num) * 100
                parent['discount_pct'] = f"{discount:.1f}%"

            colors = js_data.get('colors', [])
            sizes = js_data.get('sizes', [])
        else:
            colors = []
            sizes = []

        # Fallback regex parsing if JS didn't capture
        if not parent['product_name']:
            title_m = re.search(r'<title[^>]*>(.*?)</title>', page_html, re.IGNORECASE)
            if title_m:
                parent['product_name'] = title_m.group(1).replace(' | Kohl\'s', '').strip()

        if not parent['brand']:
            brand_m = re.search(r'(?:data-testid|class)="?brand[^>]*>([^<]+)', page_html, re.IGNORECASE)
            if brand_m:
                parent['brand'] = brand_m.group(1).strip()

        if not parent['current_price_formatted']:
            price_m = re.search(r'(\$[\d,.]+)', page_html)
            if price_m:
                parent['current_price_formatted'] = price_m.group(1)
                parent['current_price'] = parse_price_str(price_m.group(1))

        # Parse attributes from title
        parent['rise'] = parse_rise(parent['product_name'])
        parent['leg_shape'] = parse_leg_shape(parent['product_name'])
        parent['fit'] = parse_fit(parent['product_name'])

        # Parse size/fit attributes
        parent['length_hit'] = parse_length_hit(parent['size_and_fit'], parent['product_name'])
        parent['inseam'] = parse_inseam(parent['size_and_fit'], parent['product_name'])

        # Parse fabric
        if parent['fabric_raw']:
            parsed, pct_c, pct_n = parse_material(parent['fabric_raw'])
            parent['fabric_parsed'] = parsed
            parent['pct_cotton'] = pct_c
            parent['pct_natural_fiber'] = pct_n

        # Non-basic detection
        parent['brand_type'] = 'OB' if classify_brand(parent['brand']) == 'OB' else 'NB'
        bullets_text = parent['feature_bullets'] + ' ' + parent['product_details_raw']
        parent['non_basic'] = is_non_basic(parent['product_name'], ' '.join(colors), bullets_text)

        # If no colors extracted, create single row
        if not colors:
            row = {**parent}
            row['color'] = ''
            rows.append(row)
        else:
            # One row per color
            parent['total_colors'] = len(colors)
            parent['sizes_list'] = ', '.join(sizes) if sizes else ''
            parent['total_sizes'] = len(sizes)

            for color in colors:
                row = {**parent}
                row['color'] = color
                rows.append(row)

    except Exception as e:
        rows = [{
            'url': url,
            'error': f'Parse error: {str(e)[:200]}',
            'timestamp': datetime.now().isoformat(),
        }]

    return rows


# ── Scraping ────────────────────────────────────────────────────────────────

async def scrape_kohls_pdp(page, url, stats, throttle, retries=0):
    """Scrape a single Kohl's PDP page. Returns a LIST of result dicts (one per color)."""
    try:
        resp = await page.goto(url, wait_until='domcontentloaded', timeout=20000)
        status = resp.status if resp else 0

        # ── Akamai / CDN hard block detection — bail IMMEDIATELY, no retries ──
        # Kohl's uses Akamai/Edgesuite which returns 403 with "Access Denied" page.
        # Retrying is pointless — the IP is blocked at the CDN level.
        if status == 403:
            try:
                body_text = await page.text_content('body')
                body_lower = (body_text or '').lower()[:1000]
            except Exception:
                body_lower = ''
            if 'access denied' in body_lower or 'akamai' in body_lower or 'edgesuite' in body_lower or 'reference #' in body_lower:
                print(f"    AKAMAI HARD BLOCK — Access Denied (HTTP 403). Skipping immediately.")
                await stats.inc_blocked()
                return [{'url': url, 'product_name': 'Access Denied', 'error': 'Akamai CDN hard block (403)', 'retries': 0, 'timestamp': datetime.now().isoformat()}]
            # Non-Akamai 403 — could be transient, allow limited retry
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < 1:  # Only 1 retry for non-Akamai 403
                await asyncio.sleep(random.uniform(5, 10))
                return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP 403', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status == 429 or status == 503:
            await throttle.on_block()
            await stats.inc_blocked()
            if retries < MAX_RETRIES:
                await throttle.wait()
                return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        page_url = page.url.lower()
        if 'captcha' in page_url or 'blocked' in page_url or 'denied' in page_url:
            await stats.inc_blocked()
            # Check if it's an Akamai page — bail immediately
            try:
                body_text = await page.text_content('body')
                body_lower = (body_text or '').lower()[:1000]
            except Exception:
                body_lower = ''
            if 'access denied' in body_lower or 'akamai' in body_lower or 'reference #' in body_lower:
                print(f"    AKAMAI HARD BLOCK — redirected to Access Denied. Skipping immediately.")
                return [{'url': url, 'product_name': 'Access Denied', 'error': 'Akamai CDN hard block', 'retries': 0, 'timestamp': datetime.now().isoformat()}]
            wait_time = 30 + random.uniform(5, 15)
            print(f"    Blocked on {url[:60]}... backing off {wait_time:.0f}s (retry {retries + 1}/{MAX_RETRIES})")
            if retries < MAX_RETRIES:
                await asyncio.sleep(wait_time)
                return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
            return [{'url': url, 'error': 'Blocked/CAPTCHA', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        if status >= 400:
            if retries < MAX_RETRIES:
                await asyncio.sleep(random.uniform(5, 15))
                return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
            await stats.inc_fail()
            return [{'url': url, 'error': f'HTTP {status}', 'retries': retries, 'timestamp': datetime.now().isoformat()}]

        await asyncio.sleep(random.uniform(1.5, 3.0))

        # Wait for product title to load
        try:
            await page.wait_for_selector('h1', timeout=5000)
        except Exception:
            pass

        # Wait for color container if it exists
        try:
            await page.wait_for_selector('[data-testid*="color"], .color-swatch, [class*="ColorSwatch"]', timeout=3000)
        except Exception:
            pass

        html_content = await page.content()

        # Extract via JavaScript
        js_data = await extract_kohls_data(page)

        # Parse the page
        color_rows = parse_kohls_pdp(html_content, url, js_data=js_data)
        for row in color_rows:
            row['retries'] = retries
            row['url'] = url

        # ── Phase 2: Per-color pricing via swatch clicking ──
        if len(color_rows) > 1 and not color_rows[0].get('error'):
            try:
                per_color_prices = {}
                swatches = await page.query_selector_all('[data-testid*="color"], .color-swatch, [class*="ColorSwatch"]')
                if not swatches:
                    swatches = await page.query_selector_all('button[aria-label*="color"], [role="radio"][data-color]')

                for swatch in swatches:
                    try:
                        color_name = await swatch.get_attribute('aria-label') or ''
                        if not color_name:
                            color_name = await swatch.get_attribute('title') or ''
                        if not color_name:
                            color_name = (await swatch.text_content() or '').strip()
                        color_name = color_name.strip()
                        if not color_name:
                            continue

                        await swatch.click()
                        await asyncio.sleep(random.uniform(1.0, 2.0))

                        price = await page.evaluate(KOHLS_PRICE_JS)
                        if price and isinstance(price, (int, float)):
                            per_color_prices[color_name.lower()] = price
                    except Exception:
                        continue

                if per_color_prices:
                    for row in color_rows:
                        color = row.get('color', '').lower()
                        if color in per_color_prices:
                            new_price = per_color_prices[color]
                            row['current_price'] = new_price
                            row['current_price_formatted'] = f'${new_price:.2f}'
                            orig = row.get('original_price')
                            if isinstance(orig, str):
                                try:
                                    orig = float(orig.replace('$', '').replace(',', ''))
                                except ValueError:
                                    orig = None
                            if orig and isinstance(orig, (int, float)) and orig > new_price:
                                row['on_sale'] = True
                                row['discount_pct'] = f"{((orig - new_price) / orig) * 100:.1f}%"
                            elif orig and orig == new_price:
                                row['on_sale'] = False
                                row['discount_pct'] = ''
                    print(f"    Per-color prices: {len(per_color_prices)} colors updated")
            except Exception as e:
                print(f"    Phase 2 pricing skipped: {str(e)[:80]}")

        await stats.inc_ok(len(color_rows))
        await throttle.on_ok()
        return color_rows

    except Exception as e:
        err_str = str(e).lower()
        if 'pipe' in err_str or 'connection' in err_str or 'reset' in err_str or 'aborted' in err_str:
            if retries < 1:
                await asyncio.sleep(2)
                return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
            raise BrowserCrashed(f"Pipe/connection error after {retries + 1} attempts: {str(e)[:100]}")
        elif retries < MAX_RETRIES:
            await asyncio.sleep(random.uniform(5, 15))
            return await scrape_kohls_pdp(page, url, stats, throttle, retries + 1)
        await stats.inc_fail()
        return [{'url': url, 'error': str(e)[:200], 'retries': retries, 'timestamp': datetime.now().isoformat()}]


async def worker(wid, page, queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start):
    while True:
        if time.time() - batch_start > BATCH_TIME_LIMIT:
            print(f"  ⏰ W{wid}: 20-min timer hit, stopping for restart")
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break
        if url in progress['processed']:
            queue.task_done()
            continue

        done = stats.completed + stats.failed
        print(f"  [{done + 1}/{stats.total}] W{wid}: {url[:80]}...")

        await throttle.wait()
        color_rows = await scrape_kohls_pdp(page, url, stats, throttle)

        async with rlock:
            results.extend(color_rows)
            progress['processed'].add(url)

        colors_found = len(color_rows)
        has_error = any(r.get('error') for r in color_rows)
        if has_error:
            print(f"    ❌ W{wid}: Error — {color_rows[0].get('error', '')[:80]}")
        else:
            print(f"    ✅ W{wid}: {colors_found} color rows extracted")

        now = time.time()
        if now - last_save['time'] > SAVE_INTERVAL:
            async with rlock:
                save_progress(progress, sdir)
                save_to_excel(results, sdir)
                last_save['time'] = now
                print(f"  💾 Auto-saved {len(results)} rows ({len(progress['processed'])} URLs) at {datetime.now().strftime('%H:%M:%S')}")
        queue.task_done()


async def run_batch(p, urls, results, progress, stats, throttle, sdir):
    remaining = [u for u in urls if u not in progress['processed']]
    if not remaining:
        return True

    queue = asyncio.Queue()
    for u in remaining:
        await queue.put(u)
    stats.total = len(remaining)
    print(f"   📊 {len(results)} existing rows, {len(remaining)} URLs remaining")

    # Use Chromium in headed mode with stealth flags
    browser = await p.chromium.launch(headless=False, args=[
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--no-sandbox",
    ])

    async def block_resources(route):
        await route.abort()

    all_pages = []
    all_contexts = []
    print(f"🌐 Launching {NUM_CONTEXTS} contexts × {TABS_PER_CTX} tabs = {CONCURRENCY} workers")

    for ci in range(NUM_CONTEXTS):
        ctx = await setup_context(browser, ci)
        all_contexts.append(ctx)
        for ti in range(TABS_PER_CTX):
            pg = await ctx.new_page()
            await pg.route('**/*.{png,jpg,jpeg,gif,svg,ico,webp}', block_resources)
            all_pages.append(pg)

    rlock = asyncio.Lock()
    last_save = {'time': time.time()}
    batch_start = time.time()

    try:
        tasks = [
            asyncio.create_task(
                worker(i, all_pages[i], queue, results, stats, throttle, progress, sdir, rlock, last_save, batch_start)
            )
            for i in range(CONCURRENCY)
        ]
        await asyncio.gather(*tasks)
        timed_out = (time.time() - batch_start) >= BATCH_TIME_LIMIT
        finished = not timed_out
    except Exception as e:
        print(f"\n⚠️  Browser crashed: {str(e)[:100]}")
        finished = False

    save_progress(progress, sdir)
    save_to_excel(results, sdir)
    print(f"  💾 Saved {len(results)} rows ({len(progress['processed'])} URLs processed)")

    try:
        for pg in all_pages:
            await pg.close()
        for ctx in all_contexts:
            await ctx.close()
        await browser.close()
    except Exception:
        pass

    return finished


async def main():
    sdir = os.path.dirname(os.path.abspath(__file__))
    uf = os.path.join(sdir, "kohls.csv")
    if not os.path.exists(uf):
        print(f"❌ kohls.csv not found at {sdir}")
        print(f"   Create kohls.csv with one Kohl's URL per line in first column (or CSV with 'flex href' header)")
        return

    with open(uf, encoding="utf-8-sig") as f:
        lines = [l.strip() for l in f if l.strip()]

    # Try to find URLs in first column (for CSV with headers like 'flex href')
    urls = []
    for i, line in enumerate(lines):
        if i == 0 and ('href' in line.lower() or 'url' in line.lower()):
            # Skip header
            continue
        # Extract URL from CSV or plain line
        if line.startswith('http'):
            urls.append(line)
        elif ',' in line:
            # CSV: try first column
            parts = line.split(',')
            if parts[0].strip().startswith('http'):
                urls.append(parts[0].strip())

    print(f"🎯 Loaded {len(urls)} Kohl's URLs")

    progress = load_progress(sdir)
    results = progress.get('results', [])
    if not isinstance(results, list):
        results = []
    stats = Stats()
    throttle = AdaptiveThrottle()
    t0 = time.time()
    max_crashes = 20

    crash_count = 0
    while True:
        remaining = [u for u in urls if u not in progress['processed']]
        if not remaining:
            print("✅ All URLs processed!")
            break
        print(f"\n{'=' * 60}")
        print(f"▶️  Starting batch — {len(remaining)} URLs left" + (f" (restart #{crash_count})" if crash_count else ""))
        print(f"{'=' * 60}")

        was_crash = False
        try:
            pw = await async_playwright().start()
            finished = await run_batch(pw, urls, results, progress, stats, throttle, sdir)
            try:
                await pw.stop()
            except Exception:
                pass
            if finished:
                break
        except Exception as e:
            print(f"\n⚠️  Playwright/browser error: {str(e)[:150]}")
            was_crash = True
            try:
                await pw.stop()
            except Exception:
                pass

        throttle.reset()
        save_progress(progress, sdir)
        save_to_excel(results, sdir)

        if was_crash:
            crash_count += 1
            if crash_count >= max_crashes:
                print(f"❌ Browser crashed {max_crashes} times in a row, giving up. Run again to resume.")
                break
            wait = min(30, 10 * crash_count)
            print(f"🔄 Crash restart in {wait}s... (crash #{crash_count})")
            await asyncio.sleep(wait)
        else:
            crash_count = 0
            print(f"🔄 Fresh restart in 5s...")
            await asyncio.sleep(5)

    elapsed = time.time() - t0
    save_progress(progress, sdir)
    out = save_to_excel(results, sdir)
    print(f"\n{'=' * 80}")
    print(f"✅ Done in {elapsed / 60:.1f}min | URLs: {stats.completed} | Rows: {stats.rows_generated} | Fail: {stats.failed} | Blocked: {stats.blocked}")
    print(f"   Results: {out}")
    print(f"{'=' * 80}")


if __name__ == '__main__':
    asyncio.run(main())
