#!/usr/bin/env python3
"""
add_walmart_nb.py — Inject "Walmart NB" group into the dashboard.

Pulls NB rows from two sources:
  1. walmart_pdp_results.xlsx  — legacy scrape; rows held back from the dashboard
                                 (everything that's NOT a Walmart-owned brand)
  2. walmart_nb_pdp_results.xlsx — new scrape produced by running
                                   walmart_pdp_scraper.py with
                                   WALMART_INPUT=walmart_nb_urls.json
                                   WALMART_OUTPUT_PREFIX=walmart_nb_pdp

Dedupes by (url, color), maps to the RAW dashboard schema, and patches
index.html in place to add the new group + KPI tile + RAW entries.

Idempotent: re-running detects the prior injection and refuses unless --force.

Usage:
  python3 add_walmart_nb.py            # do the injection
  python3 add_walmart_nb.py --dry-run  # show what would change
  python3 add_walmart_nb.py --force    # re-inject (rebuilds Walmart NB block)
"""

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime

# Reuse mapping helpers from the existing dashboard updater
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from update_dashboard import (              # noqa: E402
    classify_wash, normalize_rise, parse_leg_shape, map_fit_style,
    parse_cotton_pct, cotton_pct_range, safe_float, parse_bool,
)

# Brands considered Walmart Owned (already in dashboard's "Walmart OB" group).
# Anything ELSE in walmart_pdp_results.xlsx is treated as held-back NB.
WALMART_OB_BRANDS = {
    "Time and Tru", "No Boundaries", "Terra & Sky", "Free Assembly",
}

LEGACY_XLSX = os.path.join(HERE, "walmart_pdp_results.xlsx")
NEW_XLSX = os.path.join(HERE, "walmart_nb_pdp_results.xlsx")
INDEX_HTML = os.path.join(HERE, "index.html")


def load_xlsx_rows(path):
    if not os.path.exists(path):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: (row[i] if i < len(row) else "") for i, h in enumerate(headers) if h}
        if rec.get("url"):
            rows.append(rec)
    wb.close()
    return rows


def transform_walmart_nb(rows):
    """Map Walmart PDP rows to the dashboard RAW schema.

    Walmart PDP cols (selected): url, product_name, brand, color,
      current_price, original_price, on_sale, discount_pct,
      pant_rise, pant_leg_cut, clothing_fit, jean_wash, pant_leg_length,
      fabric_material, fabric_pct, pack_size.
    """
    out = []
    seen = set()  # (url, color) dedupe
    for r in rows:
        url = (r.get("url") or "").strip()
        color = (r.get("color") or "").strip()
        key = (url.split("?")[0], color.lower())
        if key in seen:
            continue
        seen.add(key)

        name = (r.get("product_name") or "").strip()
        brand = (r.get("brand") or "").strip()
        if not brand:
            continue

        cur = safe_float(r.get("current_price"))
        orig = safe_float(r.get("original_price"))
        if orig <= 0:
            orig = cur
        if cur <= 0:
            cur = orig

        on_sale = parse_bool(r.get("on_sale"))
        pct_off = safe_float(r.get("discount_pct"))
        if pct_off <= 0 and on_sale and orig > 0 and cur > 0 and cur < orig:
            pct_off = round((orig - cur) / orig * 100, 1)
        if not on_sale:
            pct_off = 0

        material = (r.get("fabric_material") or "").strip()
        pct_field = (r.get("fabric_pct") or "").strip()
        # parse_cotton_pct prefers "<n>% cotton" patterns; combine the two fields
        cotton_src = (pct_field + " " + material).strip()

        wash_explicit = (r.get("jean_wash") or "").strip()
        wash = classify_wash(wash_explicit) if wash_explicit else classify_wash(color)
        # Fallback through name if still unclassified
        if (not wash) or wash == "Unclassified":
            wash = classify_wash(name) or wash or "Unclassified"

        rise = normalize_rise(r.get("pant_rise") or "", name, "")
        leg = parse_leg_shape(name, r.get("clothing_fit") or "", r.get("pant_leg_cut") or "")
        fit = map_fit_style(leg, name, r.get("clothing_fit") or "")
        cot = parse_cotton_pct(cotton_src)
        ln = (r.get("pant_leg_length") or "").strip()

        entry = {
            "g": "Walmart NB",
            "n": name,
            "b": brand,
            "p": round(cur, 2),
            "o": round(orig, 2),
            "s": 1 if on_sale else 0,
            "d": round(pct_off, 1),
            "w": wash or "Unclassified",
            "c": color,
            "ri": rise,
            "le": leg,
            "fi": fit,
            "ln": ln,
        }
        if material:
            entry["mat"] = material
        if cot is not None:
            entry["cot"] = cot
            entry["cp"] = cot
            entry["cpr"] = cotton_pct_range(cot)
        pack = r.get("pack_size") or ""
        pack = str(pack).strip()
        if pack:
            entry["pk"] = pack
        out.append(entry)
    return out


def collect_walmart_nb_entries():
    legacy = load_xlsx_rows(LEGACY_XLSX)
    new = load_xlsx_rows(NEW_XLSX)
    print(f"  Legacy walmart_pdp_results.xlsx       : {len(legacy):>5} rows")
    print(f"  New    walmart_nb_pdp_results.xlsx    : {len(new):>5} rows")
    legacy_nb = [r for r in legacy if (r.get("brand") or "").strip() not in WALMART_OB_BRANDS]
    print(f"  Legacy NB-eligible (not Walmart OB)   : {len(legacy_nb):>5} rows")

    combined = legacy_nb + new
    entries = transform_walmart_nb(combined)
    print(f"  After dedupe by (url, color)          : {len(entries):>5} entries")
    return entries


# ── HTML patchers ──────────────────────────────────────────────────────────

def patch_groups(html, force=False):
    """Insert 'Walmart NB' into GROUPS list, right after 'Walmart OB'.
    Tolerant of intervening groups like 'Target 3P' between Target NB and
    Walmart OB.
    """
    if "'Walmart NB'" in html and not force:
        print("  GROUPS already contains 'Walmart NB' (skip)")
        return html, False
    if force:
        # Strip any existing 'Walmart NB', entry first
        html = re.sub(r"'Walmart NB',", "", html, count=1)
    # Find the Walmart OB entry inside the GROUPS line and insert after it
    groups_line_pat = re.compile(r"(var GROUPS = \[[^\]]*'Walmart OB',)")
    new = groups_line_pat.sub(r"\1'Walmart NB',", html, count=1)
    if new == html:
        raise RuntimeError("Could not patch GROUPS array")
    print("  GROUPS array patched")
    return new, True


def patch_group_labels(html):
    if "'Walmart NB':'Walmart National Brands'" in html:
        print("  GROUP_LABELS already contains Walmart NB (skip)")
        return html
    new = html.replace(
        "'Walmart OB':'Walmart Owned Brands',",
        "'Walmart OB':'Walmart Owned Brands','Walmart NB':'Walmart National Brands',",
        1,
    )
    if new == html:
        raise RuntimeError("Could not patch GROUP_LABELS dict")
    print("  GROUP_LABELS dict patched")
    return new


def patch_gc(html):
    """Add Walmart NB color config — lighter Walmart blue tone."""
    if "'Walmart NB':" in html and re.search(r"'Walmart NB':\s*\{bg:", html):
        print("  GC already contains Walmart NB (skip)")
        return html
    pat = re.compile(
        r"('Walmart OB': \{bg:'#0071DC', light:'rgba\(0,113,220,0\.18\)', border:'#0071DC'\},)"
    )
    repl = (
        r"\1\n  'Walmart NB': {bg:'#FFC220', light:'rgba(255,194,32,0.18)', border:'#FFC220'},"
    )
    new = pat.sub(repl, html, count=1)
    if new == html:
        raise RuntimeError("Could not patch GC color dict")
    print("  GC color config patched (Walmart NB = Walmart yellow #FFC220)")
    return new


def patch_kpi_tile(html, count):
    """Add a KPI tile between Walmart OB and the next tile (Amazon OB)."""
    if 'Walmart NB</div>' in html and 'font-weight:700">Walmart NB' in html:
        # Update count if present
        new = re.sub(
            r"(font-weight:700\">Walmart NB</div><div[^>]*>)\d+(</div>)",
            lambda m: m.group(1) + str(count) + m.group(2),
            html, count=1,
        )
        if new != html:
            print(f"  KPI tile updated count -> {count}")
            return new
        print("  KPI tile present (no count change)")
        return html

    # Find the Walmart OB tile block and insert a new tile right after it
    # The tile block is one self-contained <div>...CCs</div></div>
    walmart_ob_tile_pat = re.compile(
        r"(<div style=\"background:var\(--bg\);border:1px solid var\(--bg3\);"
        r"border-radius:var\(--radius-sm\);padding:10px 16px;min-width:100px;"
        r"text-align:center\"><div style=\"font-size:\.65rem;color:var\(--fg3\);"
        r"letter-spacing:\.05em;text-transform:uppercase;font-weight:700\">"
        r"Walmart OB</div><div style=\"font-size:1\.3rem;font-weight:800;"
        r"color:var\(--fg\)\">\d+</div><div style=\"font-size:\.6rem;"
        r"color:var\(--fg3\)\">CCs</div></div>)"
    )
    nb_tile = (
        '<div style="background:var(--bg);border:1px solid var(--bg3);'
        'border-radius:var(--radius-sm);padding:10px 16px;min-width:100px;'
        'text-align:center"><div style="font-size:.65rem;color:var(--fg3);'
        'letter-spacing:.05em;text-transform:uppercase;font-weight:700">'
        'Walmart NB</div><div style="font-size:1.3rem;font-weight:800;'
        'color:var(--fg)">' + str(count) + '</div><div style="font-size:.6rem;'
        'color:var(--fg3)">CCs</div></div>'
    )
    new = walmart_ob_tile_pat.sub(lambda m: m.group(1) + "\n" + nb_tile, html, count=1)
    if new == html:
        print("  WARNING: could not find Walmart OB KPI tile to insert after — skipping tile")
        return html
    print(f"  KPI tile inserted after Walmart OB (count = {count})")
    return new


def patch_raw(html, entries, force=False):
    """Append new entries to the var RAW = [...] array.
    If `force`, first remove any existing 'Walmart NB' entries from the array.
    """
    raw_start = html.find("var RAW = [")
    if raw_start < 0:
        raise RuntimeError("Could not find 'var RAW = [' in index.html")
    arr_open = raw_start + len("var RAW = ")
    arr_end = html.find("];", arr_open)
    if arr_end < 0:
        raise RuntimeError("Could not find end of RAW array")
    arr_text = html[arr_open:arr_end + 1]
    data = json.loads(arr_text)
    pre_count = len(data)
    if force:
        data = [r for r in data if r.get("g") != "Walmart NB"]
        print(f"  Removed {pre_count - len(data)} pre-existing Walmart NB rows from RAW")
    elif any(r.get("g") == "Walmart NB" for r in data):
        existing = sum(1 for r in data if r.get("g") == "Walmart NB")
        print(f"  RAW already has {existing} Walmart NB entries (use --force to rebuild)")
        return html, existing
    data.extend(entries)
    new_arr = json.dumps(data, separators=(",", ":"))
    new_html = html[:arr_open] + new_arr + html[arr_end + 1:]
    print(f"  RAW: {pre_count} -> {len(data)} entries (+{len(entries)})")
    return new_html, len(entries)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true",
                    help="rebuild Walmart NB block (overwrite existing)")
    args = ap.parse_args()

    print(f"Walmart NB integration  (force={args.force}, dry_run={args.dry_run})\n")
    print("Sources:")
    entries = collect_walmart_nb_entries()
    if not entries:
        print("\nNo entries to inject — make sure the new scrape file exists "
              "(walmart_nb_pdp_results.xlsx) or there are NB rows in the legacy file.")
        return 1

    print(f"\nLoading {INDEX_HTML} ...")
    with open(INDEX_HTML, "r", encoding="utf-8") as f:
        html = f.read()
    orig_size = len(html)

    new_html, _ = patch_groups(html, force=args.force)
    new_html = patch_group_labels(new_html)
    new_html = patch_gc(new_html)
    new_html, added = patch_raw(new_html, entries, force=args.force)
    # KPI count = total Walmart NB rows after RAW patch
    kpi_count = sum(1 for r in re.finditer(r'"g":"Walmart NB"', new_html))
    new_html = patch_kpi_tile(new_html, kpi_count)

    if args.dry_run:
        print(f"\n[dry-run] would write {len(new_html):,} chars (was {orig_size:,})")
        return 0

    backup = INDEX_HTML + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(INDEX_HTML, backup)
    print(f"\nBackup written: {backup}")

    with open(INDEX_HTML, "w", encoding="utf-8") as f:
        f.write(new_html)
    print(f"index.html updated: {orig_size:,} -> {len(new_html):,} chars")
    print(f"Walmart NB total entries in RAW: {kpi_count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
